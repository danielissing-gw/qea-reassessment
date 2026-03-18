"""Validate a BOTEC .xlsx file for common errors.

Checks:
  - Circular references among cell formulas
  - Formula syntax errors (unbalanced parens, unknown functions)
  - References to non-existent sheets
  - Hyperlink integrity (display text present, target looks like a URL)
  - Structure: single-sheet with section headers, formulas in col B, labels in col A

Usage:
    python qa/validate_botec.py <path_to_xlsx>

Writes JSON report to outputs/qa_reports/{stem}_botec_qa.json
"""

import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# Excel functions recognised by openpyxl / common usage
KNOWN_FUNCTIONS = {
    "ABS", "ACOS", "ACOSH", "ADDRESS", "AND", "ASIN", "ASINH", "ATAN",
    "ATAN2", "ATANH", "AVERAGE", "AVERAGEIF", "AVERAGEIFS",
    "CEILING", "CHAR", "CHOOSE", "CLEAN", "CODE", "COLUMN", "COLUMNS",
    "COMBIN", "CONCATENATE", "COS", "COSH", "COUNT", "COUNTA",
    "COUNTBLANK", "COUNTIF", "COUNTIFS",
    "DATE", "DATEVALUE", "DAY", "DAYS", "DAYS360",
    "EDATE", "EOMONTH", "EVEN", "EXACT", "EXP",
    "FACT", "FALSE", "FIND", "FIXED", "FLOOR",
    "HLOOKUP", "HOUR", "HYPERLINK",
    "IF", "IFERROR", "IFNA", "IFS", "INDEX", "INDIRECT", "INT",
    "ISBLANK", "ISERR", "ISERROR", "ISEVEN", "ISLOGICAL", "ISNA",
    "ISNUMBER", "ISODD", "ISTEXT",
    "LARGE", "LEFT", "LEN", "LN", "LOG", "LOG10", "LOOKUP", "LOWER",
    "MATCH", "MAX", "MAXIFS", "MEDIAN", "MID", "MIN", "MINIFS",
    "MINUTE", "MOD", "MONTH",
    "NOT", "NOW", "NPV",
    "ODD", "OFFSET", "OR",
    "PERCENTILE", "PI", "PMT", "POWER", "PRODUCT", "PROPER", "PV",
    "QUOTIENT",
    "RAND", "RANDBETWEEN", "RANK", "RATE", "REPLACE", "REPT", "RIGHT",
    "ROUND", "ROUNDDOWN", "ROUNDUP", "ROW", "ROWS",
    "SEARCH", "SECOND", "SIGN", "SIN", "SINH", "SMALL", "SORT",
    "SQRT", "STDEV", "SUBSTITUTE", "SUBTOTAL", "SUM", "SUMIF",
    "SUMIFS", "SUMPRODUCT",
    "TAN", "TANH", "TEXT", "TIME", "TIMEVALUE", "TODAY", "TRIM",
    "TRUE", "TRUNC", "TYPE",
    "UNIQUE", "UPPER",
    "VALUE", "VAR", "VLOOKUP",
    "WEEKDAY", "WEEKNUM",
    "XIRR", "XLOOKUP", "XNPV",
    "YEAR", "YEARFRAC",
}

# Regex to find cell references like B2, $B$2, Sheet1!A1, 'Sheet Name'!A1
_CELL_REF = re.compile(
    r"(?:'[^']+'\!|[A-Za-z_]\w*\!)?[\$]?[A-Z]{1,3}[\$]?\d+"
)
_SHEET_REF = re.compile(r"(?:'([^']+)'\!|([A-Za-z_]\w*)\!)")
_FUNC_NAME = re.compile(r"([A-Z][A-Z0-9_.]+)\s*\(")


def _parse_refs(formula: str) -> list[str]:
    """Return list of cell references (with sheet prefix if any) from a formula."""
    return _CELL_REF.findall(formula.upper())


def _normalise_ref(ref: str, default_sheet: str) -> tuple[str, str]:
    """Return (sheet, cell) tuple, filling in default_sheet if no sheet prefix."""
    ref = ref.replace("$", "")
    if "!" in ref:
        sheet_part, cell_part = ref.split("!", 1)
        sheet_part = sheet_part.strip("'")
        return (sheet_part, cell_part)
    return (default_sheet, ref)


def check_circular_refs(wb: openpyxl.Workbook) -> list[dict]:
    """Build dependency graph across all sheets and detect cycles via DFS."""
    errors = []
    # Build adjacency list: node = "Sheet!Cell", edges = depends on
    graph: dict[str, list[str]] = defaultdict(list)
    formula_cells: set[str] = set()

    for ws in wb.worksheets:
        sheet_name = ws.title.upper()
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    node = f"{sheet_name}!{cell.coordinate}"
                    formula_cells.add(node)
                    for ref in _parse_refs(val):
                        sheet, cell_ref = _normalise_ref(ref, sheet_name)
                        dep = f"{sheet}!{cell_ref}"
                        graph[node].append(dep)

    # DFS cycle detection
    WHITE, GRAY, BLACK = 0, 1, 2
    color: dict[str, int] = defaultdict(int)
    path: list[str] = []

    def dfs(u: str) -> bool:
        color[u] = GRAY
        path.append(u)
        for v in graph.get(u, []):
            if color[v] == GRAY:
                cycle_start = path.index(v)
                cycle = path[cycle_start:] + [v]
                errors.append({
                    "check": "circular_reference",
                    "severity": "error",
                    "cycle": cycle,
                    "message": f"Circular reference detected: {' -> '.join(cycle)}"
                })
                path.pop()
                color[u] = BLACK
                return True
            if color[v] == WHITE:
                if dfs(v):
                    pass  # continue checking for more cycles
        path.pop()
        color[u] = BLACK
        return False

    for node in formula_cells:
        if color[node] == WHITE:
            dfs(node)

    return errors


def check_formulas(wb: openpyxl.Workbook) -> list[dict]:
    """Check formula syntax: parens balance, known functions, valid sheet refs."""
    errors = []
    sheet_names = {ws.title.upper() for ws in wb.worksheets}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue

                formula = val[1:]  # strip leading =
                loc = f"{ws.title}!{cell.coordinate}"

                # Unbalanced parentheses
                if formula.count("(") != formula.count(")"):
                    errors.append({
                        "check": "unbalanced_parens",
                        "severity": "error",
                        "cell": loc,
                        "formula": val,
                        "message": f"Unbalanced parentheses in {loc}: {val}"
                    })

                # Unknown functions
                for match in _FUNC_NAME.finditer(formula.upper()):
                    fname = match.group(1)
                    if fname not in KNOWN_FUNCTIONS:
                        errors.append({
                            "check": "unknown_function",
                            "severity": "warning",
                            "cell": loc,
                            "function": fname,
                            "formula": val,
                            "message": f"Possibly unknown function '{fname}' in {loc}: {val}"
                        })

                # References to non-existent sheets
                for m in _SHEET_REF.finditer(formula):
                    ref_sheet = (m.group(1) or m.group(2)).upper()
                    if ref_sheet not in sheet_names:
                        errors.append({
                            "check": "missing_sheet_ref",
                            "severity": "error",
                            "cell": loc,
                            "referenced_sheet": m.group(1) or m.group(2),
                            "formula": val,
                            "message": f"Reference to non-existent sheet '{m.group(1) or m.group(2)}' in {loc}: {val}"
                        })

    return errors


def check_hyperlinks(wb: openpyxl.Workbook) -> list[dict]:
    """Check that hyperlinks have display text and plausible URL targets."""
    errors = []
    url_pattern = re.compile(r"^https?://")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                hl = cell.hyperlink
                if hl is None:
                    continue

                target = hl.target or ""
                display = cell.value
                loc = f"{ws.title}!{cell.coordinate}"

                if not display and not hl.display:
                    errors.append({
                        "check": "hyperlink_no_display",
                        "severity": "warning",
                        "cell": loc,
                        "target": target,
                        "message": f"Hyperlink in {loc} has no display text"
                    })

                if target and not url_pattern.match(target):
                    errors.append({
                        "check": "hyperlink_bad_target",
                        "severity": "warning",
                        "cell": loc,
                        "target": target,
                        "message": f"Hyperlink target in {loc} doesn't look like a URL: {target[:80]}"
                    })

    return errors


def check_structure(wb: openpyxl.Workbook) -> list[dict]:
    """Check BOTEC structure: single sheet preferred, section headers, formulas, labels."""
    errors = []

    # Single-sheet check (warning, not error — old format had multiple)
    if len(wb.sheetnames) > 1:
        errors.append({
            "check": "multiple_sheets",
            "severity": "warning",
            "sheets": wb.sheetnames,
            "message": f"BOTEC has {len(wb.sheetnames)} sheets ({', '.join(wb.sheetnames)}). "
                       f"Expected single-sheet format."
        })

    # Check for section headers and formulas on the first sheet
    ws = wb.worksheets[0]
    has_formula = False
    has_section_header = False
    section_keywords = {"costs", "infections", "benefits", "final", "ce", "result",
                        "parameters", "calculation", "prevented", "averted"}
    has_labels_in_a = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # Check col A for labels
        cell_a = row[0] if len(row) > 0 else None
        if cell_a and cell_a.value and isinstance(cell_a.value, str):
            has_labels_in_a = True
            if any(kw in cell_a.value.lower() for kw in section_keywords):
                has_section_header = True

        # Check col B for formulas
        cell_b = row[1] if len(row) > 1 else None
        if cell_b and isinstance(cell_b.value, str) and cell_b.value.startswith("="):
            has_formula = True

    if not has_section_header:
        errors.append({
            "check": "no_section_headers",
            "severity": "warning",
            "message": "No section headers found in column A (expected keywords like "
                       "'Costs', 'Benefits', 'Final CE', etc.)"
        })

    if not has_formula:
        errors.append({
            "check": "no_formulas",
            "severity": "error",
            "message": "No formulas found in column B. BOTEC should use cell formulas, "
                       "not pre-computed values."
        })

    if not has_labels_in_a:
        errors.append({
            "check": "no_labels",
            "severity": "error",
            "message": "No parameter labels found in column A."
        })

    return errors


def validate(xlsx_path: str) -> dict:
    """Run all checks and return a combined report."""
    path = Path(xlsx_path)
    if not path.exists():
        return {
            "file": str(path),
            "status": "error",
            "errors": [{"check": "file_not_found", "severity": "error",
                        "message": f"File not found: {path}"}],
            "warnings": []
        }

    try:
        wb = openpyxl.load_workbook(str(path))
    except Exception as e:
        return {
            "file": str(path),
            "status": "error",
            "errors": [{"check": "load_failed", "severity": "error",
                        "message": f"Failed to load workbook: {e}"}],
            "warnings": []
        }

    all_issues = []
    all_issues.extend(check_circular_refs(wb))
    all_issues.extend(check_formulas(wb))
    all_issues.extend(check_hyperlinks(wb))
    all_issues.extend(check_structure(wb))

    errors = [i for i in all_issues if i["severity"] == "error"]
    warnings = [i for i in all_issues if i["severity"] == "warning"]

    return {
        "file": str(path),
        "status": "fail" if errors else ("warn" if warnings else "pass"),
        "errors": errors,
        "warnings": warnings,
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python qa/validate_botec.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(2)

    xlsx_path = sys.argv[1]
    report = validate(xlsx_path)

    # Determine output path
    stem = Path(xlsx_path).stem
    project_root = Path(__file__).resolve().parent.parent
    out_dir = project_root / "outputs" / "qa_reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{stem}_botec_qa.json"

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)

    # Print summary
    print(f"BOTEC QA: {report['status'].upper()}")
    if report["errors"]:
        print(f"  {len(report['errors'])} error(s):")
        for e in report["errors"]:
            print(f"    - {e['message']}")
    if report["warnings"]:
        print(f"  {len(report['warnings'])} warning(s):")
        for w in report["warnings"]:
            print(f"    - {w['message']}")
    if not report["errors"] and not report["warnings"]:
        print("  All checks passed.")

    print(f"Report: {out_path}")
    sys.exit(1 if report["errors"] else 0)


if __name__ == "__main__":
    main()
