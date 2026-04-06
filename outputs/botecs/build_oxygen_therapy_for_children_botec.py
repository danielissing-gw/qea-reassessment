import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

UGANDA_RCT = "https://www.thelancet.com/journals/lancet/article/PIIS0140-6736(23)02502-3/abstract"
LAM_REVIEW = "https://pmc.ncbi.nlm.nih.gov/articles/PMC8689120/"
LANCET_COMMISSION = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11865010/"
ORIGINAL_BOTEC = "https://docs.google.com/spreadsheets/d/1163LXJ5FhMT3DWjhJb4IKdHnuEOWNnTlIrkYaLv8jJE/"
HYPOXAEMIA_REVIEW = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11783038/"
MALAWI_CEA = "https://www.sciencedirect.com/science/article/pii/S2214109X25002025"
BRADLEY_CEA = "https://pubmed.ncbi.nlm.nih.gov/34165579/"
GW_MORAL_WEIGHTS = "https://docs.google.com/spreadsheets/d/1YiZmHQb-JsDIDJF_tKKVte2x5NOL0CGUTndJNoB0iHo/"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Cost per patient receiving oxygen (total program cost)")
ws.cell(3, 2, 45)
ws.cell(3, 2).number_format = "$#,##0"
c = ws.cell(3, 3, "Total cost per patient including infrastructure, training, and consumables")
c.comment = Comment(
    "Three independent estimates: (1) Original CHAI BOTEC: $15.12 CHAI + $15.12 govt + $15.12 other "
    "= $45.35 total per patient. (2) Bradley et al. 2021 CEA of solar-powered O2: $26/patient treated "
    "(JAMA Network Open, PMID 34165579). (3) Lancet Commission 2025: $59/DALY averted median "
    "(range $21-225), implying higher per-patient costs when accounting for full system costs. "
    "Using $45 as best guess: consistent with the original total cost estimate and above the "
    "modeled $26 (which may understate real-world implementation costs). "
    "If government co-funding is available, philanthropic share would be lower.",
    "BOTEC"
)

ws.cell(4, 1, "Number of patients reached")
ws.cell(4, 2, "=B2/B3")
ws.cell(4, 2).number_format = "#,##0"
ws.cell(4, 3, "Calculation")

# ── Patient composition ──────────────────────────────────────────────────────
ws.cell(5, 1, "")
ws.cell(6, 1, "Patient composition").font = section_font

ws.cell(7, 1, "% of patients who are truly hypoxaemic (vs. misuse)")
ws.cell(7, 2, 0.80)
ws.cell(7, 2).number_format = "0%"
ws.cell(7, 3, "From original BOTEC (guess to account for misuse)")

ws.cell(8, 1, "% who are children 1mo-5yr (pneumonia, malaria, sepsis)")
ws.cell(8, 2, 0.32)
ws.cell(8, 2).number_format = "0%"
add_link(8, 3, "From original BOTEC", ORIGINAL_BOTEC)

ws.cell(9, 1, "% who are neonates (<1 month)")
ws.cell(9, 2, 0.08)
ws.cell(9, 2).number_format = "0%"
ws.cell(9, 3, "From original BOTEC")

ws.cell(10, 1, "% who are adults")
ws.cell(10, 2, 0.60)
ws.cell(10, 2).number_format = "0%"
ws.cell(10, 3, "From original BOTEC")

ws.cell(11, 1, "Hypoxaemic children receiving oxygen")
ws.cell(11, 2, "=B4*B7*B8")
ws.cell(11, 2).number_format = "#,##0"
ws.cell(11, 3, "Calculation")

ws.cell(12, 1, "Hypoxaemic adults receiving oxygen")
ws.cell(12, 2, "=B4*B7*B10")
ws.cell(12, 2).number_format = "#,##0"
ws.cell(12, 3, "Calculation")

# ── Burden & effect (children) ───────────────────────────────────────────────
ws.cell(13, 1, "")
ws.cell(14, 1, "Burden & effect (children with pneumonia)").font = section_font

ws.cell(15, 1, "Baseline CFR among hospitalized children with hypoxemia")
ws.cell(15, 2, 0.19)
ws.cell(15, 2).number_format = "0%"
c = ws.cell(15, 3, "From original BOTEC (CHAI Breathing New Life proposal)")
c.comment = Comment(
    "Systematic review (Lancet Global Health 2025): patients with hypoxaemia had "
    "4.84x higher odds of death (95% CI 4.11-5.69). At baseline CFR of ~4% for "
    "pneumonia, hypoxaemic CFR of 19% is consistent.",
    "BOTEC"
)

ws.cell(16, 1, "RR of mortality with oxygen (from Uganda RCT)")
ws.cell(16, 2, 0.51)
ws.cell(16, 2).number_format = "0.00"
c = add_link(16, 3, "Nabwire et al., Lancet 2024 (n=2,405 hypoxaemic children)", UGANDA_RCT)
c.comment = Comment(
    '"After adjustment for clustering and confounding, the relative risk of a fatal '
    'outcome was 0·513 (95% CI 0·285–0·915), consistent with a 48·7% (8·5–71·5) '
    'reduction in risk of death." Consistent with Lam et al. 2021 systematic review '
    "pooled OR 0.52 (95% CI 0.39-0.70) across 4 studies.",
    "BOTEC"
)

ws.cell(17, 1, "Internal validity adjustment")
ws.cell(17, 2, 0.90)
ws.cell(17, 2).number_format = "0%"
c = ws.cell(17, 3, "Single RCT (n=2,405); consistent with 3 pre/post studies and systematic review")
c.comment = Comment(
    "The Uganda RCT is a stepped-wedge cluster design at 20 facilities. "
    "Effect is consistent with Lam et al. 2021 pooled estimate (OR 0.52). "
    "Minor discount for single trial, but confidence is higher than in the original "
    "assessment which relied only on pre/post evidence.",
    "BOTEC"
)

ws.cell(18, 1, "External validity adjustment")
ws.cell(18, 2, 0.80)
ws.cell(18, 2).number_format = "0%"
c = ws.cell(18, 3, "Extrapolation from Uganda rural facilities to broader LMIC settings")
c.comment = Comment(
    "Uganda RCT was in rural facilities using solar-powered oxygen. "
    "Effect may differ with grid-powered systems in semi-urban settings. "
    "TIMCI trial null finding in primary care underscores context-dependence. "
    "Using 80% EV (same as original) reflecting meaningful uncertainty.",
    "BOTEC"
)

ws.cell(19, 1, "Adjusted mortality reduction")
ws.cell(19, 2, "=(1-B16)*B17*B18")
ws.cell(19, 2).number_format = "0.0%"
ws.cell(19, 3, "Calculation: (1-RR) × IV × EV")

ws.cell(20, 1, "Deaths averted (children)")
ws.cell(20, 2, "=B11*B15*B19")
ws.cell(20, 2).number_format = "#,##0.0"
ws.cell(20, 3, "Calculation")

# ── Neonatal mortality (null) ────────────────────────────────────────────────
ws.cell(21, 1, "")
ws.cell(22, 1, "Neonatal mortality (null result — no benefit)").font = section_font

ws.cell(23, 1, "Neonatal deaths averted")
ws.cell(23, 2, 0)
c = ws.cell(23, 3, "RR ~1.02-1.11 (null); no evidence of neonatal benefit from oxygen alone")
c.comment = Comment(
    "Uganda RCT: IRR for neonatal deaths was 1.11 (0.76-1.62, NS). "
    "Graham et al. 2019 (Nigeria): aOR for neonates 1.12 (0.56-2.26, NS). "
    "Duke et al. 2021 (PNG): IRR 1.11 (0.76-1.62, NS). "
    "Neonatal hypoxemia may require CPAP or surfactant rather than simple O2.",
    "BOTEC"
)

# ── Adult mortality ──────────────────────────────────────────────────────────
ws.cell(24, 1, "")
ws.cell(25, 1, "Adult mortality (weak evidence)").font = section_font

ws.cell(26, 1, "Baseline CFR among hospitalized adults with hypoxemia")
ws.cell(26, 2, 0.02)
ws.cell(26, 2).number_format = "0%"
ws.cell(26, 3, "Guess from original BOTEC (surgery, maternal complications)")

ws.cell(27, 1, "RR for adults with oxygen")
ws.cell(27, 2, 0.90)
ws.cell(27, 2).number_format = "0.00"
ws.cell(27, 3, "Guess from original BOTEC — no RCT evidence for adults")

ws.cell(28, 1, "IV/EV adjustment (adults)")
ws.cell(28, 2, 0.50)
ws.cell(28, 2).number_format = "0%"
ws.cell(28, 3, "Heavy discount: no RCT, guess-based parameters")

ws.cell(29, 1, "Adjusted adult mortality reduction")
ws.cell(29, 2, "=(1-B27)*B28")
ws.cell(29, 2).number_format = "0.0%"
ws.cell(29, 3, "Calculation")

ws.cell(30, 1, "Deaths averted (adults)")
ws.cell(30, 2, "=B12*B26*B29")
ws.cell(30, 2).number_format = "#,##0.0"
ws.cell(30, 3, "Calculation")

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(31, 1, "")
ws.cell(32, 1, "Benefits").font = section_font

ws.cell(33, 1, "Moral weight per child (1mo-5yr) death averted (UoV)")
ws.cell(33, 2, 117)
c = add_link(33, 3, "GiveWell 2020 moral weights: Post Neonatal 100.7, 1-4yr 127.3", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "From GiveWell's 2020 moral weights tool. Age-specific values: "
    "Post Neonatal (1mo-1yr) = 100.7, 1 to 4yr = 127.3. "
    "Weighted average assuming ~40% post-neonatal, ~60% age 1-4 "
    "(approximate age distribution of under-5 pneumonia deaths excluding neonates): "
    "0.40 x 100.7 + 0.60 x 127.3 = 116.7, rounded to 117.",
    "BOTEC"
)

ws.cell(34, 1, "Moral weight per adult death averted (UoV)")
ws.cell(34, 2, 100)
c = add_link(34, 3, "GiveWell 2020 moral weights: weighted avg across ages 20-50", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "From GiveWell's 2020 moral weights tool. Adult oxygen patients include maternal, "
    "surgical, and respiratory cases. Assumed age distribution centered on 25-40: "
    "20-24yr=118, 25-29yr=112.7, 30-34yr=106.3, 35-39yr=99.7, 40-44yr=86.4, 45-49yr=75.7. "
    "Weighted average ~100. This is still an assumption about age mix — "
    "adult contribution is small (~2% of total UoV) so imprecision here matters little.",
    "BOTEC"
)

ws.cell(35, 1, "UoV from children's deaths averted")
ws.cell(35, 2, "=B20*B33")
ws.cell(35, 2).number_format = "#,##0"
ws.cell(35, 3, "Calculation")

ws.cell(36, 1, "UoV from adult deaths averted")
ws.cell(36, 2, "=B30*B34")
ws.cell(36, 2).number_format = "#,##0"
ws.cell(36, 3, "Calculation")

ws.cell(37, 1, "Ad-hoc: morbidity averted (developmental effects, reduced complications)")
ws.cell(37, 2, 0.10)
ws.cell(37, 2).number_format = "0%"
c = ws.cell(37, 3, "Assumption: 10% uplift for non-fatal hypoxemia episodes averted")
c.comment = Comment(
    "Children who survive hypoxemia may suffer developmental delays, brain damage, "
    "and other long-term complications. Oxygen also benefits patients who would "
    "not have died but experience morbidity. 10% is a conservative rough estimate "
    "for this benefit stream.",
    "BOTEC"
)

ws.cell(38, 1, "Total UoV from program")
ws.cell(38, 2, "=(B35+B36)*(1+B37)")
ws.cell(38, 2).number_format = "#,##0"
ws.cell(38, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(39, 1, "")
ws.cell(40, 1, "Final CE").font = section_font

ws.cell(41, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(41, 2, 0.00344)

ws.cell(42, 1, "CE (multiples of cash)")
ws.cell(42, 2, "=B38/B2/B41")
ws.cell(42, 2).number_format = "0.0"
ws.cell(42, 2).font = result_font
ws.cell(42, 3, "Calculation")

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(43, 1, "")
ws.cell(44, 1, "Sensitivity (CE at different costs per patient)").font = section_font

ws.cell(45, 1, "CE at $26/patient")
ws.cell(45, 2, "=((B2/26*B7*B8*B15*(1-B16)*B17*B18*B33)+(B2/26*B7*B10*B26*(1-B27)*B28*B34))*(1+B37)/(B2*B41)")
ws.cell(45, 2).number_format = "0.0"
ws.cell(45, 3, "Scenario: Bradley et al. 2021 modeled cost for solar O2 systems")

ws.cell(46, 1, "CE at $45/patient (best guess)")
ws.cell(46, 2, "=B42")
ws.cell(46, 2).number_format = "0.0"
ws.cell(46, 3, "= main BOTEC estimate")

ws.cell(47, 1, "CE at $75/patient")
ws.cell(47, 2, "=((B2/75*B7*B8*B15*(1-B16)*B17*B18*B33)+(B2/75*B7*B10*B26*(1-B27)*B28*B34))*(1+B37)/(B2*B41)")
ws.cell(47, 2).number_format = "0.0"
ws.cell(47, 3, "Scenario: high-cost implementation with solar + remote facilities")

ws.cell(48, 1, "CE at $100/patient")
ws.cell(48, 2, "=((B2/100*B7*B8*B15*(1-B16)*B17*B18*B33)+(B2/100*B7*B10*B26*(1-B27)*B28*B34))*(1+B37)/(B2*B41)")
ws.cell(48, 2).number_format = "0.0"
ws.cell(48, 3, "Scenario: pessimistic — full system costs in difficult settings")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=48, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "oxygen_therapy_for_children.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
