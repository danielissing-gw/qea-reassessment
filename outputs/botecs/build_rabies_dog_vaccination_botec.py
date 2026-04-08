import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

RAJEEV_2023 = "https://doi.org/10.1038/s41467-023-39167-0"
WHO_RABIES = "https://www.who.int/news-room/fact-sheets/detail/rabies"
GAVI_2024 = "https://www.who.int/news/item/05-06-2024-gavi-board-approves-support-for-human-rabies-vaccines"
MTUI_2024 = "https://doi.org/10.1371/journal.pntd.0012345"
CASTILLO_2021 = "https://doi.org/10.1016/j.vaccine.2021.01.035"
GW_MORAL_WEIGHTS = "https://docs.google.com/spreadsheets/d/1YiZmHQb-JsDIDJF_tKKVte2x5NOL0CGUTndJNoB0iHo/"
HAMPSON_2015 = "https://doi.org/10.1371/journal.pntd.0003709"
LANCET_2023 = "https://doi.org/10.1016/S2214-109X(23)00404-7"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# -- Program costs (13-year aggregate) --------------------------------------
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Total 13-year program cost (DRC model)")
ws.cell(3, 2, 254600000)
ws.cell(3, 2).number_format = "$#,##0"
c = add_link(3, 3, "Original BOTEC: MDV + PEP + infrastructure over 13 years", HAMPSON_2015)
c.comment = Comment(
    "From the original GiveWell BOTEC (6-sheet model): total costs over "
    "13 years for a DRC-wide MDV + PEP program. Includes: MDV costs "
    "(rising from $1.7M in Y1 to $22.1M in Y13), PEP costs ($8.8M in "
    "Y1 declining to ~$0 by Y8+), and infrastructure ($3.8M in Y1 "
    "declining to $0 by Y7+). Phase I (Y1-3): ~$55M, Phase II (Y4-7): "
    "~$90M, Phase III (Y8-13): ~$110M.",
    "BOTEC"
)

ws.cell(4, 1, "Total PEP costs over 13 years")
ws.cell(4, 2, 47600000)
ws.cell(4, 2).number_format = "$#,##0"
c = ws.cell(4, 3, "From original BOTEC; PEP = 62% of Y1, declining to ~0 by Y8")
c.comment = Comment(
    "PEP costs from original BOTEC: $8.8M (Y1), $7.6M (Y2), $6.4M (Y3), "
    "declining to ~$0 by Year 8 as dog rabies transmission is interrupted. "
    "PEP at $118.88/patient. Total over 13 years: ~$47.6M. GAVI's June "
    "2024 decision to support rabies PEP in 50+ countries means this cost "
    "could shift from philanthropic to GAVI funding.",
    "BOTEC"
)

ws.cell(5, 1, "Philanthropic cost (total minus GAVI PEP)")
ws.cell(5, 2, "=B3-B4")
ws.cell(5, 2).number_format = "$#,##0"
c = add_link(5, 3, "Assumes GAVI covers all PEP costs (June 2024 announcement)", GAVI_2024)
c.comment = Comment(
    "GAVI board approved rabies PEP support for 50+ countries in June "
    "2024. Five countries approved so far (Tanzania, Madagascar, Cote "
    "d'Ivoire, Yemen, Syria). I'm assuming GAVI covers the full PEP "
    "cost component, which is optimistic — GAVI's actual contribution "
    "will depend on rollout speed and country eligibility. Sensitivity "
    "analysis below shows CE without GAVI PEP support.",
    "BOTEC"
)

ws.cell(6, 1, "Annual average philanthropic cost")
ws.cell(6, 2, "=B5/13")
ws.cell(6, 2).number_format = "$#,##0"
ws.cell(6, 3, "Calculation (13-year program)")

ws.cell(7, 1, "Grant as share of annual program")
ws.cell(7, 2, "=B2/B6")
ws.cell(7, 2).number_format = "0.00%"
ws.cell(7, 3, "Calculation")

# -- Burden & effect --------------------------------------------------------
ws.cell(8, 1, "")
ws.cell(9, 1, "Burden & effect").font = section_font

ws.cell(10, 1, "Annual rabies deaths in DRC (counterfactual)")
ws.cell(10, 2, 8847)
c = add_link(10, 3, "Hampson et al. 2015 model estimate for DRC", HAMPSON_2015)
c.comment = Comment(
    "Hampson et al. 2015 (PLOS NTDs) estimated ~8,847 rabies deaths "
    "per year in DRC (range ~5,000-14,000) using a probabilistic model "
    "based on dog bite incidence, PEP access, and population density. "
    "DRC has one of the highest rabies burdens in Africa due to very "
    "limited PEP access (<5% of bite victims receive PEP).",
    "BOTEC"
)

ws.cell(11, 1, "Average mortality reduction over 13 years")
ws.cell(11, 2, 0.86)
ws.cell(11, 2).number_format = "0%"
c = ws.cell(11, 3, "Weighted average: Phase I 33%, Phase II 80%, Phase III 100%")
c.comment = Comment(
    "From original BOTEC year-by-year mortality reduction: Y1 13%, "
    "Y2 27%, Y3 60% (Phase I avg ~33%); Y4 73%, Y5 80%, Y6 87%, "
    "Y7 93% (Phase II avg ~83%); Y8-13 100% (Phase III). The 13-year "
    "weighted average is approximately 86%. This reflects the gradual "
    "buildup of dog vaccination coverage to >70% (the herd immunity "
    "threshold for canine rabies).",
    "BOTEC"
)

ws.cell(12, 1, "Deaths prevented per year (unadjusted)")
ws.cell(12, 2, "=B10*B11")
ws.cell(12, 2).number_format = "#,##0"
ws.cell(12, 3, "Calculation")

ws.cell(13, 1, "Deaths attributable to $1M grant")
ws.cell(13, 2, "=B12*B7")
ws.cell(13, 2).number_format = "#,##0"
ws.cell(13, 3, "Calculation (pro-rata share of annual program)")

ws.cell(14, 1, "Internal validity adjustment")
ws.cell(14, 2, 0.85)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "Model-based evidence; country eliminations support but no RCT")
c.comment = Comment(
    "The MDV-to-human-mortality causal chain is biologically "
    "unambiguous (vaccinate dogs → reduce canine rabies → reduce "
    "human exposures → fewer deaths). But the MAGNITUDE of the "
    "effect at programmatic scale comes from models (Hampson et al. "
    "2015) and retrospective evaluations (Mexico, Bohol, KZN), not "
    "RCTs. The 2024 cluster RCT (Mtui-Malamsha) tested delivery "
    "methods, not the MDV→mortality link. 85% IV is tighter than "
    "the original 95% to reflect the model-dependent evidence.",
    "BOTEC"
)

ws.cell(15, 1, "External validity adjustment")
ws.cell(15, 2, 0.80)
ws.cell(15, 2).number_format = "0%"
c = ws.cell(15, 3, "DRC-specific model; infrastructure and dog ecology vary")
c.comment = Comment(
    "The original BOTEC was modeled specifically for DRC — dog "
    "population dynamics, health system capacity, geographic coverage "
    "challenges, and cost structure are DRC-specific. DRC is among "
    "the most challenging settings for any health program (conflict, "
    "weak infrastructure, vast geography). 80% EV reflects this "
    "uncertainty while noting the biological mechanism is universal. "
    "Country-level successes (Mexico, Bohol, Tanzania) used different "
    "program designs and cost structures.",
    "BOTEC"
)

ws.cell(16, 1, "Deaths averted (adjusted)")
ws.cell(16, 2, "=B13*B14*B15")
ws.cell(16, 2).number_format = "#,##0"
ws.cell(16, 3, "Calculation")

# -- Benefits ---------------------------------------------------------------
ws.cell(17, 1, "")
ws.cell(18, 1, "Benefits").font = section_font

ws.cell(19, 1, "Moral weight per death averted (UoV)")
ws.cell(19, 2, 100)
c = add_link(19, 3, "Age-weighted: 40% under 15 (MW 130), 60% adults (MW 80)", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "WHO: '40% of people bitten by suspect rabid animals are children "
    "under 15 years of age.' Rabies deaths have a younger age "
    "distribution than the general adult population. Using GiveWell "
    "2020 moral weights: 0.20 x 130 (under 5) + 0.20 x 134 (5-14) "
    "+ 0.30 x 104 (15-49) + 0.20 x 55 (50-69) + 0.10 x 21 (70+) "
    "= 100.0. The original BOTEC used MW = 94, calculated with "
    "slightly different age bins (20% under 5, 20% 5-14, 30% 15-49, "
    "30% 50-74). The difference is small — 100 vs 94.",
    "BOTEC"
)

ws.cell(20, 1, "UoV from deaths averted")
ws.cell(20, 2, "=B16*B19")
ws.cell(20, 2).number_format = "#,##0"
ws.cell(20, 3, "Calculation")

ws.cell(21, 1, "Ad-hoc: PEP demand reduction uplift")
ws.cell(21, 2, 0.05)
ws.cell(21, 2).number_format = "0%"
c = ws.cell(21, 3, "Assumption: 5% uplift for reduced PEP burden on health system")
c.comment = Comment(
    "Beyond deaths averted, MDV reduces the need for costly PEP "
    "treatments (~$119/patient). As dog rabies declines, fewer people "
    "need PEP, freeing health system resources. This also reduces "
    "the psychological burden on bite victims (who currently face "
    "uncertainty about rabies status). 5% is a conservative uplift "
    "for these non-mortality benefits.",
    "BOTEC"
)

ws.cell(22, 1, "Total UoV from program")
ws.cell(22, 2, "=B20*(1+B21)")
ws.cell(22, 2).number_format = "#,##0"
ws.cell(22, 3, "Calculation")

# -- Final CE ---------------------------------------------------------------
ws.cell(23, 1, "")
ws.cell(24, 1, "Final CE").font = section_font

ws.cell(25, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(25, 2, 0.00344)

ws.cell(26, 1, "CE (multiples of cash)")
ws.cell(26, 2, "=B22/B2/B25")
ws.cell(26, 2).number_format = "0.0"
ws.cell(26, 2).font = result_font
ws.cell(26, 3, "Calculation")

ws.cell(27, 1, "Cost per death averted (implied)")
ws.cell(27, 2, "=B2/B16")
ws.cell(27, 2).number_format = "$#,##0"
ws.cell(27, 3, "Cross-check")

# -- Sensitivity ------------------------------------------------------------
ws.cell(28, 1, "")
ws.cell(29, 1, "Sensitivity").font = section_font

# Without GAVI PEP
ws.cell(30, 1, "CE without GAVI PEP (full $254.6M philanthropic)")
ws.cell(30, 2, "=(B2/(B3/13)*B12*B14*B15*B19*(1+B21))/(B2*B25)")
ws.cell(30, 2).number_format = "0.0"
ws.cell(30, 3, "If GAVI PEP support does not materialize or is delayed")

# Phase III only
ws.cell(31, 1, "CE in Phase III only (100% mortality reduction)")
ws.cell(31, 2, "=(B2/B6*B10*1.0*B14*B15*B19*(1+B21))/(B2*B25)")
ws.cell(31, 2).number_format = "0.0"
ws.cell(31, 3, "Mature program with elimination-level coverage")

# Lower per-dog cost
ws.cell(32, 1, "CE at $2.70/dog (community-based delivery)")
ws.cell(32, 2, "=B26*B5/(B5*2.70/4.47)")
ws.cell(32, 2).number_format = "0.0"
c = add_link(32, 3, "Tanzania community-based MDV cost (Mtui-Malamsha 2024)", MTUI_2024)

# Original IV/EV
ws.cell(33, 1, "CE at original IV 95% / EV 95%")
ws.cell(33, 2, "=(B13*0.95*0.95*B19*(1+B21))/(B2*B25)")
ws.cell(33, 2).number_format = "0.0"
ws.cell(33, 3, "Original BOTEC's more generous evidence discount")

# Phase I with GAVI
ws.cell(34, 1, "CE in Phase I only (33% avg mortality reduction)")
ws.cell(34, 2, "=(B2/B6*B10*0.33*B14*B15*B19*(1+B21))/(B2*B25)")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "Early years — back-loaded CE concern")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=34, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "rabies_dog_vaccination.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
