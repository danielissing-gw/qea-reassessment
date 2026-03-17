import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ── PARAMETERS SHEET ──────────────────────────────────────────────────────────
ws_p = wb.active
ws_p.title = "Parameters"

header_font = Font(bold=True)
input_fill = PatternFill("solid", fgColor="DDEEFF")   # blue = input
calc_fill  = PatternFill("solid", fgColor="F0F0F0")   # grey = computed

headers = ["Parameter", "Original value", "Updated value", "Source", "Notes"]
for col, h in enumerate(headers, 1):
    cell = ws_p.cell(row=1, column=col, value=h)
    cell.font = header_font

rows = [
    # (Parameter, Original, Updated, Source, Notes)
    ("Effect size: % reduction in RHD risk with prevention program",
     0.89, 0.89,
     "original",
     "From Cuba observational study (Pinar del Rio 1986-1996); unchanged — no comparable new observational data found."),

    ("Internal validity adjustment",
     0.70, 0.70,
     "original",
     "Unchanged — evidence base remains observational, not RCT."),

    ("External validity adjustment",
     0.70, 0.70,
     "original",
     "Unchanged — Cuba→India extrapolation still uncertain; 2024 WHO guidelines provide policy support but do not resolve efficacy in India context."),

    ("Combined validity-adjusted effect (= row 2 × row 3 × row 4)",
     None, None,
     "calculation",
     "Computed in Calculation sheet."),

    ("% of over-20 deaths caused by RHD in India",
     0.0150, 0.0166,
     "https://www.ahajournals.org/doi/10.1161/circ.150.suppl_1.4144731",
     "Updated: GBD 2021 data indicates ~166,017 India RHD deaths vs ~129,000 in original (AHA abstract; medium confidence). "
     "India all-cause deaths ~10M/yr → 166k/10M ≈ 1.66%. Original used 1.50%."),

    ("Lifetime RHD deaths per cohort of 100,000",
     1500, 1660,
     "calculation",
     "= row 6 × 100,000. Updated proportionally with revised burden estimate."),

    ("Units of value per RHD death averted (GiveWell moral weights)",
     52.5, 52.5,
     "original",
     "Unchanged — no new GiveWell moral weight revision found."),

    ("Annual program cost in 2010 USD (Watkins et al. 2015, Cuba study)",
     20289, 20289,
     "original",
     "No new cost data found. This remains the key uncertainty. GiveWell itself considered this likely a gross underestimate."),

    ("CPI inflation factor (2010 → year of assessment)",
     1.21, 1.46,
     "assumption",
     "Original used 2010→2021 factor of 1.21. Updated to approximate 2010→2026 CPI factor of ~1.46 (I'm estimating this; "
     "the BLS CPI calculator should be used to verify: https://www.bls.gov/data/inflation_calculator.htm)."),

    ("Number covered by program annually (Watkins et al. 2015)",
     273933, 273933,
     "original",
     "Unchanged."),

    ("Duration of intervention required per beneficiary (years)",
     19, 19,
     "original",
     "Unchanged — intervention covers ages 5–24."),

    ("GiveDirectly units of value per $100,000 donated",
     344, 344,
     "original",
     "Unchanged."),
]

for r, (param, orig, upd, src, notes) in enumerate(rows, 2):
    ws_p.cell(row=r, column=1, value=param)
    if orig is not None:
        ws_p.cell(row=r, column=2, value=orig)
        ws_p.cell(row=r, column=2).fill = input_fill if src != "calculation" else calc_fill
    if upd is not None:
        ws_p.cell(row=r, column=3, value=upd)
        ws_p.cell(row=r, column=3).fill = input_fill if src != "calculation" else calc_fill
    ws_p.cell(row=r, column=4, value=src)
    ws_p.cell(row=r, column=5, value=notes)

# Row for combined effect (row 5, parameter index 4 → data row 5)
combined_row = 5
ws_p.cell(row=combined_row, column=2, value="=B2*B3*B4")
ws_p.cell(row=combined_row, column=3, value="=C2*C3*C4")
ws_p.cell(row=combined_row, column=2).fill = calc_fill
ws_p.cell(row=combined_row, column=3).fill = calc_fill

# Column widths
ws_p.column_dimensions["A"].width = 55
ws_p.column_dimensions["B"].width = 18
ws_p.column_dimensions["C"].width = 18
ws_p.column_dimensions["D"].width = 55
ws_p.column_dimensions["E"].width = 80
for col in ["D", "E"]:
    for row in range(1, len(rows) + 2):
        ws_p.cell(row=row, column=ord(col)-64).alignment = Alignment(wrap_text=True)

# ── CALCULATION SHEET ──────────────────────────────────────────────────────────
ws_c = wb.create_sheet("Calculation")

calc_rows = [
    ("Step", "Description", "Original", "Updated", "Notes"),
    (1,  "Combined validity-adjusted effect size",
         "=Parameters!B5",  "=Parameters!C5",  "= effect × internal validity × external validity"),
    (2,  "Lifetime RHD deaths per 100,000 cohort",
         "=Parameters!B7",  "=Parameters!C7",  "Based on % of over-20 deaths from RHD in India × 100,000"),
    (3,  "Lifetime deaths averted per 100,000 cohort (= step 1 × step 2)",
         "=C3*C4",           "=D3*D4",           "Input"),
    (4,  "Units of value per death averted",
         "=Parameters!B8",  "=Parameters!C8",  "GiveWell moral weights for LMIC adult death"),
    (5,  "Total units of value per 100,000 cohort (= step 3 × step 4)",
         "=C5*C6",           "=D5*D6",           ""),
    (6,  "Annual program cost in inflation-adjusted USD",
         "=Parameters!B9*Parameters!B10", "=Parameters!C9*Parameters!C10", "= raw cost × CPI factor"),
    (7,  "Number covered annually",
         "=Parameters!B11", "=Parameters!C11", "From original study"),
    (8,  "Annual cost per beneficiary",
         "=C8/C9",          "=D8/D9",           ""),
    (9,  "Total cost per beneficiary over full program duration (= step 8 × duration)",
         "=C10*Parameters!B12", "=D10*Parameters!C12", ""),
    (10, "Total cost per 100,000 cohort (= step 9 × 100,000)",
         "=C11*100000",     "=D11*100000",      ""),
    (11, "Units of value per $100,000 donated (= step 5 / step 10 × 100,000)",
         "=C7/C12*100000",  "=D7/D12*100000",   ""),
    (12, "GiveDirectly units of value per $100,000",
         "=Parameters!B13", "=Parameters!C13",  "Benchmark"),
    ("",  "── RESULT ──", "", "", ""),
    (13, "MULTIPLES OF CASH (= step 11 / step 12)",
         "=C13/C14",        "=D13/D14",         "Main cost-effectiveness output"),
]

for r, row_data in enumerate(calc_rows, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws_c.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = Font(bold=True)
        if r == len(calc_rows) and c in (3, 4):
            cell.font = Font(bold=True, size=12)

ws_c.column_dimensions["A"].width = 6
ws_c.column_dimensions["B"].width = 55
ws_c.column_dimensions["C"].width = 20
ws_c.column_dimensions["D"].width = 20
ws_c.column_dimensions["E"].width = 45

# ── NOTES SHEET ───────────────────────────────────────────────────────────────
ws_n = wb.create_sheet("Notes")
notes_text = [
    ("Key caveats and uncertainties", ""),
    ("", ""),
    ("1. Cost is the pivotal unknown.",
     "The $0.09/person-year figure from the Cuba study (Watkins et al. 2015) is almost certainly an underestimate "
     "for India. The original GiveWell BOTEC itself stated this is 'possibly a gross underestimate.' Downstream medical "
     "costs (drug costs for treating identified GAS cases, secondary prevention injections) are entirely excluded. "
     "GiveWell's own note: the intervention would still be 10x cash even at 6x cost increase."),
    ("", ""),
    ("2. Effect size from Cuba may not apply to India.",
     "The 89% RHD reduction came from Pinar del Rio, Cuba — a setting with universal free healthcare and strong "
     "primary care infrastructure. The 0.70 external validity adjustment already attempts to correct for this, "
     "but India's health system infrastructure differs substantially."),
    ("", ""),
    ("3. New evidence since 2021 (not yet in model):",
     "- WHO 2024 guideline: strong recommendation for antibiotic treatment of clinically suspected GAS in "
     "high-risk LMIC settings — could enable government co-financing and lower marginal costs. "
     "- India Lancet 2023 CEA: secondary prevention estimated at ~$30/QALY in India — if correct, implies "
     "~100x+ cash benchmark; different program design than this BOTEC. "
     "- GBD 2021: India burden updated to ~166k deaths (vs ~129k assumed); updated in this BOTEC."),
    ("", ""),
    ("4. Missing benefit streams.",
     "Morbidity (from heart failure, valve damage), perinatal outcomes (RHD is leading heart condition in "
     "pregnant women in LMICs), and treatment cost offsets are all excluded. Including them would increase CE estimate."),
]
for r, (a, b) in enumerate(notes_text, 1):
    ws_n.cell(row=r, column=1, value=a).font = Font(bold=bool(a and not a.startswith(" ")))
    ws_n.cell(row=r, column=2, value=b)
ws_n.column_dimensions["A"].width = 35
ws_n.column_dimensions["B"].width = 90
ws_n.cell(1, 1).font = Font(bold=True, size=12)

out_path = os.path.join(os.path.dirname(__file__), "rheumatic_heart_disease_prevention.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
