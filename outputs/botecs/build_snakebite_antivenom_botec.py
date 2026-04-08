import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

HAMZA_2016 = "https://journals.plos.org/plosntds/article?id=10.1371/journal.pntd.0004568"
GBD_2019 = "https://www.nature.com/articles/s41467-022-33627-9"
OPHIREX = "https://www.globenewswire.com/news-release/2025/12/11/3203856/0/en/Ophirex-Announces-Business-Update-Advances-Toward-Food-and-Drug-Administration-FDA-Approval-for-Varespladib-as-Novel-Oral-Snakebite-Treatment.html"
NATURE_NANOBODY = "https://www.nature.com/articles/s41586-025-09661-0"
ASEAN_2022 = "https://pmc.ncbi.nlm.nih.gov/articles/PMC9668136/"
GW_MORAL_WEIGHTS = "https://docs.google.com/spreadsheets/d/1YiZmHQb-JsDIDJF_tKKVte2x5NOL0CGUTndJNoB0iHo/"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# -- Costs ---------------------------------------------------------------
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Antivenom + direct treatment cost per patient")
ws.cell(3, 2, 124)
ws.cell(3, 2).number_format = "$#,##0.00"
c = add_link(3, 3, "West African EchiTAb average (Hamza et al. 2016)", HAMZA_2016)
c.comment = Comment(
    "Hamza et al. 2016 (PLOS NTDs, 16 West African countries): "
    "'The cost per death averted ranged from US$1,997 in Guinea-Bissau "
    "to US$6,205 in Sierra Leone and Liberia.' Average treatment cost "
    "~$124, based on EchiTAb-G (monospecific, ~1 vial for carpet viper) "
    "and EchiTAb Plus-ICP (~3 vials, ~$41/vial). This is dramatically "
    "lower than the $900/5-vial dose assumed in the original BOTEC, "
    "which used the more expensive polyvalent antivenoms.",
    "BOTEC"
)

ws.cell(4, 1, "Program overhead per patient (cold chain, distribution, training)")
ws.cell(4, 2, 51)
ws.cell(4, 2).number_format = "$#,##0.00"
c = ws.cell(4, 3, "Assumption: ~41% overhead on treatment cost")
c.comment = Comment(
    "Antivenom requires IV administration in a health facility, with "
    "cold chain storage, trained health workers, and observation periods. "
    "Program overhead includes: procurement and supply chain management, "
    "cold chain maintenance (antivenoms are temperature-sensitive), "
    "health worker training (species identification, dosing protocols, "
    "anaphylaxis management), and monitoring. I'm assuming $51/patient "
    "overhead (~41% of drug cost), giving $175 total. This is uncertain "
    "— the Hamza CEA included some but not all programmatic costs.",
    "BOTEC"
)

ws.cell(5, 1, "Total cost per treatment")
ws.cell(5, 2, "=B3+B4")
ws.cell(5, 2).number_format = "$#,##0.00"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "Patients treated")
ws.cell(6, 2, "=B2/B5")
ws.cell(6, 2).number_format = "#,##0"
ws.cell(6, 3, "Calculation")

# -- Burden & effect ------------------------------------------------------
ws.cell(7, 1, "")
ws.cell(8, 1, "Burden & effect").font = section_font

ws.cell(9, 1, "Weighted untreated case fatality rate")
ws.cell(9, 2, 0.10)
ws.cell(9, 2).number_format = "0%"
c = ws.cell(9, 3, "Weighted: 66% carpet viper (15% CFR) + 34% other (3% CFR)")
c.comment = Comment(
    "Carpet viper (Echis ocellatus) is responsible for ~66% of "
    "snakebite deaths in West Africa. Untreated envenomation CFR "
    "for carpet viper: ~15% (range 10-20% from hospital-based studies "
    "in Nigeria). For other species (cobra, mamba, puff adder), "
    "untreated envenomation CFR: ~3-5%. Weighted: 0.66 x 0.15 + "
    "0.34 x 0.03 = 10.9%, rounded to 10%. Note: this is for "
    "ENVENOMED patients only — 'dry bites' (no venom) are excluded.",
    "BOTEC"
)

ws.cell(10, 1, "Mortality reduction with antivenom treatment")
ws.cell(10, 2, 0.75)
ws.cell(10, 2).number_format = "0%"
c = ws.cell(10, 3, "OR 0.25 from meta-analysis of 4 observational studies (Habib & Warrell 2013)")
c.comment = Comment(
    "Habib & Warrell 2013 meta-analysis of 4 observational studies: "
    "OR 0.25 for mortality with antivenom vs. without. This corresponds "
    "to a ~75% reduction in mortality. The evidence is entirely "
    "observational (no RCTs exist and a placebo-controlled trial would "
    "be unethical), but with strong biological plausibility and "
    "consistent results across studies. The IV adjustment accounts "
    "for the non-randomized evidence base.",
    "BOTEC"
)

ws.cell(11, 1, "Absolute risk reduction")
ws.cell(11, 2, "=B9*B10")
ws.cell(11, 2).number_format = "0.00%"
ws.cell(11, 3, "Calculation")

ws.cell(12, 1, "Deaths before adjustments")
ws.cell(12, 2, "=B6*B11")
ws.cell(12, 2).number_format = "#,##0"
ws.cell(12, 3, "Calculation")

ws.cell(13, 1, "Internal validity adjustment")
ws.cell(13, 2, 0.70)
ws.cell(13, 2).number_format = "0%"
c = ws.cell(13, 3, "No RCTs; observational evidence only but biologically plausible")
c.comment = Comment(
    "All evidence on antivenom effectiveness comes from observational "
    "studies. No RCT exists and a placebo-controlled RCT would be "
    "unethical given antivenom's established biological mechanism. "
    "70% IV reflects the inherent limitation of observational evidence "
    "while crediting the strong biological plausibility (antivenom "
    "directly neutralizes circulating venom) and the consistency of "
    "the OR 0.25 across 4 studies.",
    "BOTEC"
)

ws.cell(14, 1, "External validity adjustment")
ws.cell(14, 2, 0.70)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "West African data; snake species and health systems vary across regions")
c.comment = Comment(
    "The West African CEA data (Hamza 2016) and the mortality "
    "reduction estimate (Habib & Warrell 2013) are both from West "
    "Africa, primarily Nigeria. Antivenom effectiveness varies by "
    "snake species — EchiTAb products are specific to Echis ocellatus "
    "(carpet viper). Different polyvalent antivenoms would be needed "
    "for South Asian species (e.g., Big 4 in India). Health system "
    "capacity for IV administration also varies. 70% EV reflects "
    "these generalizability concerns while acknowledging that the "
    "biological mechanism transfers across settings.",
    "BOTEC"
)

ws.cell(15, 1, "Wastage / inefficiency factor")
ws.cell(15, 2, 0.80)
ws.cell(15, 2).number_format = "0%"
c = ws.cell(15, 3, "20% waste from expiry, species mismatch, late presentation")
c.comment = Comment(
    "Not all procured antivenom is effectively used. Sources of "
    "wastage include: (1) antivenom expiry before use (limited shelf "
    "life + unpredictable demand); (2) species mismatch (patient "
    "bitten by species not covered by available antivenom); (3) "
    "patients arriving too late for antivenom to be effective; "
    "(4) anaphylactic reactions requiring treatment discontinuation. "
    "80% efficiency (20% wastage) is the same assumption as the "
    "original BOTEC.",
    "BOTEC"
)

ws.cell(16, 1, "Deaths averted (adjusted)")
ws.cell(16, 2, "=B12*B13*B14*B15")
ws.cell(16, 2).number_format = "#,##0"
ws.cell(16, 3, "Calculation")

# -- Benefits -------------------------------------------------------------
ws.cell(17, 1, "")
ws.cell(18, 1, "Benefits").font = section_font

ws.cell(19, 1, "Moral weight per death averted (UoV)")
ws.cell(19, 2, 95)
c = add_link(19, 3, "Age-weighted average using GiveWell 2020 moral weights", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "Snakebite deaths have a younger age distribution than assumed "
    "in the original BOTEC. From GBD 2019 India data (>80% of global "
    "burden): 28% under 15 (MW ~130), 17% ages 15-29 (MW ~120), "
    "47% ages 30-69 (MW ~40-106), 9% ages 70+ (MW ~21). "
    "Weighted average: 0.28 x 130 + 0.17 x 120 + 0.30 x 93 + "
    "0.17 x 55 + 0.09 x 21 = 95.4, rounded to 95. "
    "The original BOTEC implicitly used ~40 UoV (adult equivalent) "
    "which substantially undervalued snakebite deaths.",
    "BOTEC"
)

ws.cell(20, 1, "UoV from deaths averted")
ws.cell(20, 2, "=B16*B19")
ws.cell(20, 2).number_format = "#,##0"
ws.cell(20, 3, "Calculation")

ws.cell(21, 1, "Ad-hoc: morbidity uplift (chronic disability in survivors)")
ws.cell(21, 2, 0.10)
ws.cell(21, 2).number_format = "0%"
c = ws.cell(21, 3, "Assumption: 10% uplift for non-fatal benefits")
c.comment = Comment(
    "Snakebite envenomation causes significant chronic morbidity in "
    "survivors: tissue necrosis requiring amputation (~2-5% of carpet "
    "viper victims), chronic pain, renal damage, and psychological "
    "trauma. WHO estimates 400,000+ people live with permanent "
    "disabilities from snakebite. Antivenom reduces severity of "
    "these outcomes. 10% is a conservative morbidity uplift.",
    "BOTEC"
)

ws.cell(22, 1, "Total UoV from program")
ws.cell(22, 2, "=B20*(1+B21)")
ws.cell(22, 2).number_format = "#,##0"
ws.cell(22, 3, "Calculation")

# -- Final CE -------------------------------------------------------------
ws.cell(23, 1, "")
ws.cell(24, 1, "Final CE").font = section_font

ws.cell(25, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(25, 2, 0.00344)

ws.cell(26, 1, "CE (multiples of cash)")
ws.cell(26, 2, "=B22/B2/B25")
ws.cell(26, 2).number_format = "0.0"
ws.cell(26, 2).font = result_font
ws.cell(26, 3, "Calculation")

ws.cell(27, 1, "Cost per death averted (implied)")
ws.cell(27, 2, "=B2/B16")
ws.cell(27, 2).number_format = "$#,##0"
ws.cell(27, 3, "Cross-check")

# -- Sensitivity ----------------------------------------------------------
ws.cell(28, 1, "")
ws.cell(29, 1, "Sensitivity").font = section_font

# Sensitivity: antivenom cost
ws.cell(30, 1, "CE at $900/treatment (original BOTEC cost)")
ws.cell(30, 2, "=(B2/(900+B4)*B11*B13*B14*B15*B19*(1+B21))/(B2*B25)")
ws.cell(30, 2).number_format = "0.0"
ws.cell(30, 3, "Original: $180/vial x 5 vials polyvalent")

ws.cell(31, 1, "CE at $124/treatment (central, West African EchiTAb)")
ws.cell(31, 2, "=B26")
ws.cell(31, 2).number_format = "0.0"
ws.cell(31, 3, "= main BOTEC estimate")

ws.cell(32, 1, "CE at $50/treatment (Benin-like lower cost)")
ws.cell(32, 2, "=(B2/(50+B4)*B11*B13*B14*B15*B19*(1+B21))/(B2*B25)")
ws.cell(32, 2).number_format = "0.0"
c = add_link(32, 3, "Benin: $83/DALY averted (Hamza 2016)", HAMZA_2016)

ws.cell(33, 1, "CE at $20/treatment (varespladib/next-gen, speculative)")
ws.cell(33, 2, "=(B2/(20+B4)*B11*B13*B14*B15*B19*(1+B21))/(B2*B25)")
ws.cell(33, 2).number_format = "0.0"
c = add_link(33, 3, "Varespladib ~$10-20/course oral (Ophirex target)", OPHIREX)

# Sensitivity: IV/EV
ws.cell(34, 1, "CE at 50% IV / 50% EV (heavier observational discount)")
ws.cell(34, 2, "=(B2/B5*B11*0.50*0.50*B15*B19*(1+B21))/(B2*B25)")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "If observational evidence warrants deeper skepticism")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=34, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "snakebite_antivenom.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
