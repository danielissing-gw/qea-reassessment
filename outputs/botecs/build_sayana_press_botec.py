import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

GW_CONTRACEPTION_URL = "https://www.givewell.org/how-we-work/our-criteria/cost-effectiveness/valuing-contraception"
PATH_DMPASC_URL = "https://www.path.org/our-impact/articles/dmpa-sc/"
GUTTMACHER_URL = "https://www.guttmacher.org/sites/default/files/report_pdf/aiu-2012-estimates_0.pdf"
LAUNCH_SCALE_URL = "https://launchandscalefaster.org/intervention/sayanar-press"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 55
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"
ws.cell(2, 3, "Arbitrary anchor")

ws.cell(3, 1, "Commodity cost per dose (Sayana Press)")
ws.cell(3, 2, 1.00)
ws.cell(3, 2).number_format = "$#,##0.00"
c = add_link(3, 3, "Pfizer/Gates/CIFF agreement: $1/dose for 69 poorest countries", PATH_DMPASC_URL)
c.comment = Comment(
    'Under a novel pricing agreement negotiated by Pfizer Inc., the Bill & Melinda '
    'Gates Foundation, and the Children\'s Investment Fund Foundation (CIFF) in 2014, '
    'qualified purchasers in the world\'s poorest 69 countries can obtain Sayana Press '
    'for US$1 per dose.',
    "BOTEC"
)

ws.cell(4, 1, "Doses per year")
ws.cell(4, 2, 4)
ws.cell(4, 3, "Quarterly injection (every 3 months)")

ws.cell(5, 1, "Commodity cost per woman-year")
ws.cell(5, 2, "=B3*B4")
ws.cell(5, 2).number_format = "$#,##0.00"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "Delivery cost per woman-year (self-injection program)")
ws.cell(6, 2, 8.00)
ws.cell(6, 2).number_format = "$#,##0.00"
c = ws.cell(6, 3, "Estimate; original Guttmacher 2012 total was $9.14/year (incl. commodity)")
c.comment = Comment(
    'Original source: Singh and Darroch 2012, Table 3: injectable annualised cost '
    'for Africa = $9.14 (including commodity, supplies, and labour). With self-injection, '
    'delivery costs should be lower as no health worker needed at each injection. '
    '$8 assumes training, supply chain, and minimal demand generation. '
    'THIS IS AN ASSUMPTION I HAVE NOT VERIFIED against actual self-injection program budgets.',
    "BOTEC"
)

ws.cell(7, 1, "Total cost per woman-year of contraception")
ws.cell(7, 2, "=B5+B6")
ws.cell(7, 2).number_format = "$#,##0.00"
ws.cell(7, 3, "Calculation")

ws.cell(8, 1, "Woman-years of contraception provided by grant")
ws.cell(8, 2, "=B2/B7")
ws.cell(8, 2).number_format = "#,##0"
ws.cell(8, 3, "Calculation")

# ── Benefits (GW contraception valuation) ─────────────────────────────────
ws.cell(9, 1, "Benefits (GW contraception valuation)").font = section_font

ws.cell(10, 1, "UoV per woman-year of modern contraception (GW estimate)")
ws.cell(10, 2, 0.7)
c = add_link(10, 3, "GW April 2025 contraception valuation", GW_CONTRACEPTION_URL)
c.comment = Comment(
    'GiveWell estimates that providing one year of modern contraception in low- and '
    'middle-income countries generates approximately 0.7 units of value (with a '
    '25th-75th percentile range of 0.3-1.1 units). This encompasses: improved health '
    'for women/children (~0.2 UoV), increased earnings for women (~0.1 UoV), '
    'increased resources for existing children (~0.2 UoV), and improved subjective '
    'well-being (~0.2 UoV).',
    "BOTEC"
)

ws.cell(11, 1, "UoV per woman-year (low estimate, 25th percentile)")
ws.cell(11, 2, 0.3)
add_link(11, 3, "GW April 2025 contraception valuation", GW_CONTRACEPTION_URL)

ws.cell(12, 1, "UoV per woman-year (high estimate, 75th percentile)")
ws.cell(12, 2, 1.1)
add_link(12, 3, "GW April 2025 contraception valuation", GW_CONTRACEPTION_URL)

# ── Adjustments ───────────────────────────────────────────────────────────
ws.cell(13, 1, "Adjustments").font = section_font

ws.cell(14, 1, "Self-injection continuation rate uplift")
ws.cell(14, 2, 1.0)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "Meta-analysis: RR 1.27 for 12-month continuation. "
                    "Already partially captured in GW's UoV estimate; not double-counting.")
c.comment = Comment(
    'Meta-analysis of three RCTs found self-injection increases 12-month continuation '
    'by 27% (RR 1.27, 95% CI 1.16-1.39). However, GW\'s 0.7 UoV estimate is for '
    'a generic year of modern contraception. Self-injection\'s higher continuation '
    'could mean effective coverage is higher than 1 woman-year per woman-year purchased. '
    'I am NOT applying an uplift here to avoid double-counting, but note this is conservative.',
    "BOTEC"
)

ws.cell(15, 1, "IV adjustment")
ws.cell(15, 2, 1.0)
ws.cell(15, 2).number_format = "0%"
ws.cell(15, 3, "No IV adjustment: GW's valuation already incorporates evidence quality")

ws.cell(16, 1, "EV adjustment")
ws.cell(16, 2, 1.0)
ws.cell(16, 2).number_format = "0%"
ws.cell(16, 3, "No EV adjustment: GW's valuation is for LMICs broadly, matching target context")

# ── Total benefits ────────────────────────────────────────────────────────
ws.cell(17, 1, "Total benefits").font = section_font

ws.cell(18, 1, "Total UoV generated (best guess)")
ws.cell(18, 2, "=B8*B10*B14*B15*B16")
ws.cell(18, 2).number_format = "#,##0"
ws.cell(18, 3, "Calculation")

ws.cell(19, 1, "Total UoV generated (low)")
ws.cell(19, 2, "=B8*B11*B14*B15*B16")
ws.cell(19, 2).number_format = "#,##0"
ws.cell(19, 3, "Calculation (25th percentile UoV)")

ws.cell(20, 1, "Total UoV generated (high)")
ws.cell(20, 2, "=B8*B12*B14*B15*B16")
ws.cell(20, 2).number_format = "#,##0"
ws.cell(20, 3, "Calculation (75th percentile UoV)")

# ── Final CE ──────────────────────────────────────────────────────────────
ws.cell(21, 1, "Final CE").font = section_font

ws.cell(22, 1, "GW benchmark (UoV per dollar at 1x)")
ws.cell(22, 2, 0.003355)
ws.cell(22, 3, "0.003355 UoV/$ = 1x cash transfers (344 UoV per $100k)")

ws.cell(23, 1, "CE (best guess)")
ws.cell(23, 2, "=B18/B2/B22")
ws.cell(23, 2).number_format = "0.0"
ws.cell(23, 2).font = result_font
ws.cell(23, 3, "Calculation: total UoV / grant / benchmark")

ws.cell(24, 1, "CE (low)")
ws.cell(24, 2, "=B19/B2/B22")
ws.cell(24, 2).number_format = "0.0"
ws.cell(24, 3, "Calculation (25th percentile)")

ws.cell(25, 1, "CE (high)")
ws.cell(25, 2, "=B20/B2/B22")
ws.cell(25, 2).number_format = "0.0"
ws.cell(25, 3, "Calculation (75th percentile)")

# ── Sensitivity: cost per woman-year ──────────────────────────────────────
ws.cell(27, 1, "Sensitivity: CE at different costs per woman-year").font = section_font

ws.cell(28, 1, "CE at $6/woman-year (lean self-injection)")
ws.cell(28, 2, "=B2/$6*B10/B2/B22")
ws.cell(28, 2).number_format = "0.0"

ws.cell(29, 1, "CE at $12/woman-year (moderate self-injection)")
ws.cell(29, 2, "=B2/B7*B10/B2/B22")
ws.cell(29, 2).number_format = "0.0"

ws.cell(30, 1, "CE at $20/woman-year (provider-administered + demand gen)")
ws.cell(30, 2, "=B2/$20*B10/B2/B22")
ws.cell(30, 2).number_format = "0.0"

ws.cell(31, 1, "CE at $30/woman-year (comprehensive program)")
ws.cell(31, 2, "=B2/$30*B10/B2/B22")
ws.cell(31, 2).number_format = "0.0"

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=31, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "sayana_press.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
