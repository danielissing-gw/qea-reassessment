import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

UGANDA_PROGRAM = "https://pmc.ncbi.nlm.nih.gov/articles/PMC6328136/"
AMBITION = "https://www.nejm.org/doi/full/10.1056/NEJMoa2111904"
RAJASINGHAM_BURDEN = "https://www.thelancet.com/journals/laninf/article/PIIS1473-3099(22)00042-7/fulltext"
PLOS_UPTAKE = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11761098/"
SA_COST = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11538356/"
WHO_GUIDELINES = "https://www.who.int/publications/i/item/9789240052178/"
GW_MORAL_WEIGHTS = "https://docs.google.com/spreadsheets/d/1YiZmHQb-JsDIDJF_tKKVte2x5NOL0CGUTndJNoB0iHo/"
RAMACHANDRAN_CEA = "https://bmcinfectdis.biomedcentral.com/articles/10.1186/s12879-017-2325-9"
CM_AGE_REVIEW = "https://pmc.ncbi.nlm.nih.gov/articles/PMC9494297/"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Cost per person screened (CrAg test + delivery)")
ws.cell(3, 2, 10)
ws.cell(3, 2).number_format = "$#,##0"
c = ws.cell(3, 3, "Blended estimate: reflexed lab ($5-7) and POC ($15-20)")
c.comment = Comment(
    'South Africa NHLS reflexed screening: "$5.897 per result" (PMC 2024). '
    'Brazil CrAg-LFA: "$6.00 per test" (PMC 2021). '
    "POC in lower-resource settings would add health worker time, consumables, "
    "and sample transport: ~$15-20. Using $10 as blended estimate assuming "
    "philanthropic program would target countries with some existing lab infrastructure "
    "(reflexed model) while also supporting POC in some facilities.",
    "BOTEC"
)

ws.cell(4, 1, "Additional cost per CrAg-positive patient (pre-emptive Rx + monitoring)")
ws.cell(4, 2, 50)
ws.cell(4, 2).number_format = "$#,##0"
c = ws.cell(4, 3, "Pre-emptive fluconazole ($20) + follow-up visits ($15) + LP if symptomatic ($15)")
c.comment = Comment(
    "Pre-emptive fluconazole: 800mg/day for 2 weeks then 200mg/day for 8 weeks. "
    "Generic fluconazole is very cheap (~$0.05-0.10/150mg tablet in LMICs). "
    "Full course ~$10-20. Monitoring visits (2-3 over 10 weeks): ~$15. "
    "Lumbar puncture for symptomatic patients to confirm/exclude CM: ~$15 "
    "(not all CrAg+ patients need LP). Total ~$50/CrAg+ patient.",
    "BOTEC"
)

ws.cell(5, 1, "Additional cost per confirmed CM case (AMBITION regimen + hospitalization)")
ws.cell(5, 2, 200)
ws.cell(5, 2).number_format = "$#,##0"
c = ws.cell(5, 3, "Single-dose liposomal AmB ($100-150) + hospitalization ($50-100)")
c.comment = Comment(
    "AMBITION regimen: single high-dose liposomal amphotericin B (10 mg/kg) + "
    "14 days flucytosine + fluconazole. Liposomal AmB donated by Gilead (AmBisome "
    "Access Program) or procured at ~$100-150 for single dose. Flucytosine ~$20-40 "
    "for 14 days. Fluconazole ~$5. Hospitalization reduced from 7+ days to 1-3 days "
    "with single-dose regimen. Total ~$150-300, using $200.",
    "BOTEC"
)

ws.cell(6, 1, "CrAg prevalence among CD4 <100 patients")
ws.cell(6, 2, 0.05)
ws.cell(6, 2).number_format = "0%"
c = add_link(6, 3, "Midpoint: SA 6.2%, Uganda 1.4%, pooled SSA ~4-6%", SA_COST)
c.comment = Comment(
    'South Africa national CrAg detection rate: 6.2% at CD4 <100 (NHLS data 2024). '
    'Uganda national program: CrAg prevalence 1.4% (Michelow et al. 2019). '
    "Rajasingham et al. 2022 global estimates show wide variation by country. "
    "Using 5% as midpoint for high-burden SSA countries. This is a key sensitivity parameter.",
    "BOTEC"
)

ws.cell(7, 1, "% of CrAg+ with confirmed CM (needing AMBITION treatment)")
ws.cell(7, 2, 0.25)
ws.cell(7, 2).number_format = "0%"
c = ws.cell(7, 3, "~25% of CrAg+ patients have active CM at diagnosis")
c.comment = Comment(
    "Among CrAg+ patients, approximately 20-30% have confirmed CM on lumbar puncture "
    "at the time of screening. The remainder are subclinical/early stage and benefit "
    "from pre-emptive fluconazole. Using 25% as estimate.",
    "BOTEC"
)

ws.cell(8, 1, "Total cost per person screened (blended)")
ws.cell(8, 2, "=B3+B6*(B4+B7*B5)")
ws.cell(8, 2).number_format = "$#,##0.00"
ws.cell(8, 3, "Calculation: screening cost + CrAg prevalence × (treatment + CM% × CM treatment)")

ws.cell(9, 1, "Number of people screened")
ws.cell(9, 2, "=B2/B8")
ws.cell(9, 2).number_format = "#,##0"
ws.cell(9, 3, "Calculation")

ws.cell(10, 1, "CrAg-positive patients identified")
ws.cell(10, 2, "=B9*B6")
ws.cell(10, 2).number_format = "#,##0"
ws.cell(10, 3, "Calculation")

# ── Burden & effect ──────────────────────────────────────────────────────────
ws.cell(11, 1, "")
ws.cell(12, 1, "Burden & effect").font = section_font

ws.cell(13, 1, "1-year mortality among CrAg+ without pre-emptive treatment")
ws.cell(13, 2, 0.25)
ws.cell(13, 2).number_format = "0%"
c = ws.cell(13, 3, "Conservative estimate based on multiple sources")
c.comment = Comment(
    "AMBITION trial control arm 10-week mortality: 28.7% — but these patients already "
    "had confirmed CM (the sickest subgroup). For CrAg+ patients overall (including "
    "subclinical), 1-year mortality without treatment is ~20-40% depending on setting. "
    "REMSTART trial control: 11% at 12 weeks among all CD4<200 (not just CrAg+). "
    "Using 25% as estimate for CrAg+ specific 1-year mortality without pre-emptive Rx.",
    "BOTEC"
)

ws.cell(14, 1, "Mortality reduction from pre-emptive fluconazole treatment")
ws.cell(14, 2, 0.50)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "Based on modeling and observational data")
c.comment = Comment(
    "No RCT of CrAg screening + pre-emptive treatment vs. no screening for mortality. "
    "Uganda national program evaluation: 'CrAg screening averts 43% of deaths' among "
    "target population (Michelow et al. 2019). Ramachandran et al. 2017 modeling: CrAg "
    "screening 'cost-effective in 100% of scenarios.' REMSTART trial HR 0.65 for "
    "all-cause mortality with AHD screening package (including CrAg). Using 50% "
    "mortality reduction for CrAg+ patients specifically, which is conservative given "
    "the 43% figure applies to the entire screened population (mostly CrAg-negative).",
    "BOTEC"
)

ws.cell(15, 1, "Internal validity adjustment")
ws.cell(15, 2, 0.80)
ws.cell(15, 2).number_format = "0%"
c = ws.cell(15, 3, "No RCT of screening program; REMSTART HR 0.65 was non-significant")
c.comment = Comment(
    "The mortality reduction is based on modeling (Uganda evaluation), a single "
    "trial that was not statistically significant at 95% (REMSTART HR 0.65, "
    "95% CI 0.41-1.03), and observational/programmatic data. The biological "
    "mechanism is well-understood (CrAg identifies subclinical CM; pre-emptive "
    "fluconazole prevents progression). 80% IV discount reflects the lack of a "
    "definitive RCT while acknowledging the strong mechanistic basis.",
    "BOTEC"
)

ws.cell(16, 1, "External validity adjustment")
ws.cell(16, 2, 0.85)
ws.cell(16, 2).number_format = "0%"
c = ws.cell(16, 3, "Uganda/SA data extrapolated to other SSA settings")
c.comment = Comment(
    "CrAg screening has been implemented in multiple SSA countries with consistent "
    "results. CrAg prevalence varies by country (1.4% Uganda to 6.2% South Africa) "
    "but the screening mechanism is the same. Treatment protocols are WHO-standardized. "
    "85% EV reflects that a new program would need to establish lab/POC capacity and "
    "follow-up systems, which may be less effective than established programs.",
    "BOTEC"
)

ws.cell(17, 1, "Adjusted mortality reduction")
ws.cell(17, 2, "=B14*B15*B16")
ws.cell(17, 2).number_format = "0.0%"
ws.cell(17, 3, "Calculation: mortality reduction × IV × EV")

ws.cell(18, 1, "Deaths averted")
ws.cell(18, 2, "=B10*B13*B17")
ws.cell(18, 2).number_format = "#,##0.0"
ws.cell(18, 3, "Calculation: CrAg+ patients × baseline mortality × adjusted reduction")

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(19, 1, "")
ws.cell(20, 1, "Benefits").font = section_font

ws.cell(21, 1, "Moral weight per death averted (UoV) — age-weighted")
ws.cell(21, 2, 98)
c = add_link(21, 3, "GiveWell 2020 moral weights, weighted for CM age distribution", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "Median age of CM patients in SSA: 33-36 years (Frontiers in Medicine 2022 "
    "systematic review). Age distribution estimate: 5% ages 20-24 (MW 118), "
    "15% ages 25-29 (MW 112.7), 25% ages 30-34 (MW 106.3), 25% ages 35-39 "
    "(MW 99.7), 20% ages 40-44 (MW 86.4), 10% ages 45-49 (MW 75.7). "
    "Weighted average: 0.05×118 + 0.15×112.7 + 0.25×106.3 + 0.25×99.7 "
    "+ 0.20×86.4 + 0.10×75.7 = 98.8, rounded to 98.",
    "BOTEC"
)

ws.cell(22, 1, "UoV from deaths averted")
ws.cell(22, 2, "=B18*B21")
ws.cell(22, 2).number_format = "#,##0"
ws.cell(22, 3, "Calculation")

ws.cell(23, 1, "Ad-hoc: morbidity averted (CM-related disability, DALYs from survivors)")
ws.cell(23, 2, 0.15)
ws.cell(23, 2).number_format = "0%"
c = ws.cell(23, 3, "Assumption: 15% uplift for non-fatal benefits")
c.comment = Comment(
    "CM survivors often suffer long-term sequelae: hearing loss (~25%), visual "
    "impairment (~10%), cognitive deficits, and ongoing need for ART adherence "
    "support. Pre-emptive treatment prevents CM entirely in most cases, averting "
    "both the acute episode and long-term disability. Uganda evaluation found "
    "20.6 DALYs averted per death prevented, suggesting substantial morbidity "
    "component. 15% is a conservative estimate.",
    "BOTEC"
)

ws.cell(24, 1, "Total UoV from program")
ws.cell(24, 2, "=B22*(1+B23)")
ws.cell(24, 2).number_format = "#,##0"
ws.cell(24, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(25, 1, "")
ws.cell(26, 1, "Final CE").font = section_font

ws.cell(27, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(27, 2, 0.00344)

ws.cell(28, 1, "CE (multiples of cash)")
ws.cell(28, 2, "=B24/B2/B27")
ws.cell(28, 2).number_format = "0.0"
ws.cell(28, 2).font = result_font
ws.cell(28, 3, "Calculation")

ws.cell(29, 1, "Cost per death averted (implied)")
ws.cell(29, 2, "=B2/B18")
ws.cell(29, 2).number_format = "$#,##0"
c = ws.cell(29, 3, "Cross-check: Uganda program found $459/death averted")
c.comment = Comment(
    'Michelow et al. 2019: "cost of $459 per life saved" in the Uganda national '
    "program. Our estimate will be higher because we include philanthropic delivery "
    "costs and IV/EV discounts that the Uganda evaluation did not apply.",
    "BOTEC"
)

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(30, 1, "")
ws.cell(31, 1, "Sensitivity").font = section_font

# Sensitivity: screening cost
ws.cell(32, 1, "CE at $5/person screened (reflexed lab model)")
ws.cell(32, 2, "=(B2/(5+B6*(B4+B7*B5))*B6*B13*B14*B15*B16*B21*(1+B23))/(B2*B27)")
ws.cell(32, 2).number_format = "0.0"
add_link(32, 3, "Reflexed screening: marginal cost ~$5/result (SA NHLS data)", SA_COST)

ws.cell(33, 1, "CE at $10/person screened (best guess)")
ws.cell(33, 2, "=B28")
ws.cell(33, 2).number_format = "0.0"
ws.cell(33, 3, "= main BOTEC estimate")

ws.cell(34, 1, "CE at $20/person screened (new POC program)")
ws.cell(34, 2, "=(B2/(20+B6*(B4+B7*B5))*B6*B13*B14*B15*B16*B21*(1+B23))/(B2*B27)")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "Point-of-care screening in lower-resource settings")

# Sensitivity: CrAg prevalence
ws.cell(35, 1, "CE at 2% CrAg prevalence (low-burden setting)")
ws.cell(35, 2, "=(B2/(B3+0.02*(B4+B7*B5))*0.02*B13*B14*B15*B16*B21*(1+B23))/(B2*B27)")
ws.cell(35, 2).number_format = "0.0"
ws.cell(35, 3, "E.g., East Africa with better ART coverage")

ws.cell(36, 1, "CE at 8% CrAg prevalence (high-burden setting)")
ws.cell(36, 2, "=(B2/(B3+0.08*(B4+B7*B5))*0.08*B13*B14*B15*B16*B21*(1+B23))/(B2*B27)")
ws.cell(36, 2).number_format = "0.0"
ws.cell(36, 3, "E.g., Southern Africa, West/Central Africa")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=36, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "cryptococcal_antigen_testing.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
