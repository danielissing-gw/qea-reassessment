import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

REACH_TRIAL = "https://www.nejm.org/doi/full/10.1056/NEJMoa1813598"
REACH_FOLLOWUP = "https://www.thelancet.com/journals/lanhae/article/PIIS2352-3026(24)00078-4/abstract"
GBD_2021 = "https://www.thelancet.com/journals/lanhae/article/PIIS2352-3026(23)00118-7/fulltext"
UGANDA_CEA = "https://link.springer.com/article/10.1007/s40273-023-01294-3"
WHO_EML = "https://list.essentialmeds.org/medicines/93"
GW_MORAL_WEIGHTS = "https://docs.google.com/spreadsheets/d/1YiZmHQb-JsDIDJF_tKKVte2x5NOL0CGUTndJNoB0iHo/"
CHAI_BOTEC = "Local file: 230724 SCD BOTECvShare.xlsx"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    if url.startswith("http"):
        c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 65
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Cost per child-year of SCD treatment (full program)")
ws.cell(3, 2, 60)
ws.cell(3, 2).number_format = "$#,##0"
c = ws.cell(3, 3, "Includes screening, HU, other therapies, TA, M&E")
c.comment = Comment(
    "Derived from CHAI Nigeria BOTEC (July 2023): total 5-year cost ~$29M for ~472K cumulative "
    "child-years of treatment = ~$62/child-year average. Year 1 is higher (~$96) due to fixed "
    "costs; year 5 is lower (~$47) as fixed costs amortize. Using $60 as central estimate. "
    "A treatment-only program (where screening exists, e.g. Uganda/Ghana) could be $35-40/child-year. "
    "HU drug cost alone is only ~$18/year for children >1yr.",
    "BOTEC"
)

ws.cell(4, 1, "Child-years of treatment per $1M")
ws.cell(4, 2, "=B2/B3")
ws.cell(4, 2).number_format = "#,##0"
ws.cell(4, 3, "Calculation")

# ── Burden ────────────────────────────────────────────────────────────────────
ws.cell(5, 1, "")
ws.cell(6, 1, "Burden (SCD in children under 5)").font = section_font

ws.cell(7, 1, "Case fatality rate per 1,000 person-years (SCD, under 5)")
ws.cell(7, 2, 29.9)
c = add_link(7, 3, "IHME GBD 2021: ~24,400 under-5 SCD deaths / ~817K under-5 SCD person-years in Nigeria", GBD_2021)
c.comment = Comment(
    "GBD 2021 (Lancet Haematology 2023): 81,100 under-5 SCD deaths globally in 2021. "
    "Nigeria accounts for ~30% of global SCD burden. CHAI BOTEC uses 29.85/1000 person-years "
    "from IHME data. Note: GBD 2021 total-mortality methodology suggests true burden may be "
    "~11x higher than cause-specific coding — the 29.9 figure is the cause-specific rate.",
    "BOTEC"
)

ws.cell(8, 1, "Expected deaths without treatment in treated cohort")
ws.cell(8, 2, "=B4*B7/1000")
ws.cell(8, 2).number_format = "#,##0.0"
ws.cell(8, 3, "Calculation")

# ── Effect of hydroxyurea ────────────────────────────────────────────────────
ws.cell(9, 1, "")
ws.cell(10, 1, "Effect of hydroxyurea treatment").font = section_font

ws.cell(11, 1, "Mortality reduction from HU (from REACH trial)")
ws.cell(11, 2, 0.70)
ws.cell(11, 2).number_format = "0%"
c = add_link(11, 3, "REACH: IRR 0.30 → 70% mortality reduction (pre-post, n=606)", REACH_TRIAL)
c.comment = Comment(
    "Tshilolo et al. 2019 NEJM (REACH trial): hydroxyurea in children with SCA in SSA "
    "(DRC, Uganda, Kenya, Angola). IRR for mortality 0.30. However, this is a pre-post "
    "comparison — children received substantially more monitoring and care in the treatment "
    "period. The 8-year follow-up (Lancet Haematology 2024) confirms sustained benefit at "
    "MTD dosing with 4,340 patient-years of treatment. No RCT exists.",
    "BOTEC"
)

ws.cell(12, 1, "Internal validity adjustment")
ws.cell(12, 2, 0.65)
ws.cell(12, 2).number_format = "0%"
c = ws.cell(12, 3, "Pre-post design with increased monitoring; no RCT available")
c.comment = Comment(
    "The REACH trial is a pre-post study, not an RCT. GW reviewers flagged that children "
    "received more care and monitoring in the post-treatment period, which could drive part "
    "of the observed effect. CHAI used 89% IV which GW considered too high. I'm using 65% — "
    "the effect is consistent across 4 SSA countries and 8 years of follow-up, but the "
    "confounding from increased care is a genuine concern. Meta-analysis of 3 small RCTs "
    "(Narra X 2024, n=423) confirms HbF improvement but doesn't directly measure mortality.",
    "BOTEC"
)

ws.cell(13, 1, "External validity adjustment")
ws.cell(13, 2, 0.80)
ws.cell(13, 2).number_format = "0%"
c = ws.cell(13, 3, "REACH was at research sites; national implementation would differ")
c.comment = Comment(
    "REACH trial was conducted at research hospitals in DRC, Uganda, Kenya, and Angola. "
    "Scaling to routine health facilities in Nigeria (or other settings) introduces "
    "challenges: adherence without research monitoring, drug supply chain reliability, "
    "loss to follow-up. CONSA data shows only 38.5% of screened SCD babies had a documented "
    "clinical visit — suggesting major care cascade problems at scale. "
    "CHAI used 97.5% EV which is too high for a program that hasn't been implemented at scale.",
    "BOTEC"
)

ws.cell(14, 1, "Adherence adjustment")
ws.cell(14, 2, 0.80)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "HU requires daily administration by caregivers; not modeled in CHAI BOTEC")
c.comment = Comment(
    "HU is a daily oral medication administered at home by caregivers. No child-friendly "
    "liquid formulation widely available (Siklos approved in Ghana 2022 but not widely "
    "distributed). Caregivers crush tablets or dissolve capsule powder. US studies show "
    "adherence ~60-80% among children. Education of caregivers is critical, especially for "
    "asymptomatic children. GW flagged this as a missing adjustment in the CHAI model.",
    "BOTEC"
)

ws.cell(15, 1, "Adjusted mortality reduction")
ws.cell(15, 2, "=B11*B12*B13*B14")
ws.cell(15, 2).number_format = "0.0%"
ws.cell(15, 3, "Calculation: effect × IV × EV × adherence")

ws.cell(16, 1, "Deaths averted per $1M (face value)")
ws.cell(16, 2, "=B8*B15")
ws.cell(16, 2).number_format = "#,##0.0"
ws.cell(16, 3, "Calculation")

# ── Sustainability adjustment ────────────────────────────────────────────────
ws.cell(17, 1, "")
ws.cell(18, 1, "Sustainability adjustment (critical assumption)").font = section_font

ws.cell(19, 1, "Sustainability discount on moral weight")
ws.cell(19, 2, 0.50)
ws.cell(19, 2).number_format = "0%"
c = ws.cell(19, 3, "SCD requires lifelong treatment; program covers only ages 0-5")
c.comment = Comment(
    "This is the most important and most uncertain parameter in the BOTEC. "
    "SCD requires lifelong HU treatment, but the proposed program only covers children 0-5. "
    "A death averted at age 2 has full moral weight only if the child lives to normal life "
    "expectancy. If treatment stops at 5 and the child dies at 10-15, we've delayed the death "
    "rather than averted it. 50% discount assumes a mix of outcomes: ~30% chance of lifelong "
    "treatment access (government uptake, endowment), ~30% chance of survival to ~25 without "
    "treatment (milder SCD genotypes), ~40% chance of death before age 20. "
    "This discount does NOT apply to morbidity benefits, which are realized during the program period.",
    "BOTEC"
)

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(20, 1, "")
ws.cell(21, 1, "Benefits").font = section_font

ws.cell(22, 1, "Moral weight per under-5 death averted (UoV)")
ws.cell(22, 2, 117)
c = add_link(22, 3, "GiveWell 2020 moral weights: Post Neonatal 100.7, 1-4yr 127.3", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "From GiveWell's 2020 moral weights tool. Most SCD deaths in under-5s occur in "
    "the first 2 years of life (50% in year 1 per IHME). Weighted average of "
    "Post Neonatal (100.7) and 1-4yr (127.3), roughly 40/60 split = 117.",
    "BOTEC"
)

ws.cell(23, 1, "UoV from mortality (sustainability-adjusted)")
ws.cell(23, 2, "=B16*B22*B19")
ws.cell(23, 2).number_format = "#,##0"
ws.cell(23, 3, "Calculation: deaths averted × moral weight × sustainability discount")

ws.cell(24, 1, "Ad-hoc: morbidity averted (pain crises, stroke, organ damage)")
ws.cell(24, 2, 0.15)
ws.cell(24, 2).number_format = "0%"
c = ws.cell(24, 3, "Assumption: 15% uplift for morbidity benefits realized during program")
c.comment = Comment(
    "SCD causes severe morbidity beyond mortality: vaso-occlusive pain crises, acute chest "
    "syndrome, stroke, splenic sequestration, chronic organ damage, and cognitive impairment. "
    "HU reduces pain crises by ~40-50% and transfusion needs by 75% (REACH 8yr follow-up). "
    "15% uplift is conservative — the true morbidity benefit may be larger. "
    "Unlike the mortality benefit, morbidity benefits are fully realized during the program period "
    "and do not need a sustainability discount.",
    "BOTEC"
)

ws.cell(25, 1, "UoV from morbidity (no sustainability discount needed)")
ws.cell(25, 2, "=B16*B22*B24")
ws.cell(25, 2).number_format = "#,##0"
ws.cell(25, 3, "Calculation")

ws.cell(26, 1, "Total UoV from program")
ws.cell(26, 2, "=B23+B25")
ws.cell(26, 2).number_format = "#,##0"
ws.cell(26, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(27, 1, "")
ws.cell(28, 1, "Final CE").font = section_font

ws.cell(29, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(29, 2, 0.00344)

ws.cell(30, 1, "CE (multiples of cash)")
ws.cell(30, 2, "=B26/B2/B29")
ws.cell(30, 2).number_format = "0.0"
ws.cell(30, 2).font = result_font
ws.cell(30, 3, "Calculation")

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(31, 1, "")
ws.cell(32, 1, "Sensitivity").font = section_font

# CE without sustainability discount
ws.cell(33, 1, "CE without sustainability discount (face value)")
ws.cell(33, 2, "=B16*B22*(1+B24)/B2/B29")
ws.cell(33, 2).number_format = "0.0"
ws.cell(33, 3, "If lifelong treatment is ensured (endowment/govt uptake)")

# Treatment-only program (lower cost)
ws.cell(34, 1, "CE at $35/child-year (treatment-only, e.g. Uganda/Ghana)")
ws.cell(34, 2, "=(B2/35*B7/1000*B15*B22*(B19+B24))/B2/B29")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "Countries with existing screening; lower overhead")

# Treatment-only + no sustainability discount
ws.cell(35, 1, "CE at $35/child-year + no sustainability discount")
ws.cell(35, 2, "=(B2/35*B7/1000*B15*B22*(1+B24))/B2/B29")
ws.cell(35, 2).number_format = "0.0"
ws.cell(35, 3, "Best case: existing screening + lifelong treatment ensured")

# Pessimistic
ws.cell(36, 1, "CE at $80/child-year + 70% sustainability discount")
ws.cell(36, 2, "=(B2/80*B7/1000*B15*B22*(0.30+B24))/B2/B29")
ws.cell(36, 2).number_format = "0.0"
ws.cell(36, 3, "Pessimistic: high costs, poor treatment sustainability")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=36, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "sickle_cell_disease.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
