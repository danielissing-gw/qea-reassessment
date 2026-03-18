import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

ACTION_I_NEJM = "https://www.nejm.org/doi/full/10.1056/NEJMoa2022398"
ACTION_I_CEA = "https://www.sciencedirect.com/science/article/pii/S2214109X22003400"
WHO_ACS_2022 = "https://www.who.int/publications/i/item/9789240057296"


def add_link(row, col, text, url):
    """Write a cell with a hyperlink."""
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Drug cost per course (dexamethasone, 4x6mg IM)")
ws.cell(3, 2, 0.51)
ws.cell(3, 2).number_format = "$#,##0.00"
c = ws.cell(3, 3, "Save the Children 2013: 'as little as 51 cents per treatment'")
c.comment = Comment(
    "Dexamethasone is on the WHO Essential Medicines List and is widely available. "
    "The $0.51 figure covers only the drug cost. The full regimen is four 6mg IM "
    "injections at 12-hour intervals.",
    "BOTEC"
)

ws.cell(4, 1, "Incremental program cost per woman treated (training, ultrasound, supervision)")
ws.cell(4, 2, 29.49)
ws.cell(4, 2).number_format = "$#,##0.00"
c = ws.cell(4, 3, "Assumption: ~$30 total cost per woman in hospitals with adequate neonatal care.")
c.comment = Comment(
    "This is the key uncertain parameter. Components:\n"
    "- Drug: $0.51 (included separately)\n"
    "- Training costs amortized per woman: ~$5-10\n"
    "- Ultrasound dating (if not already available): ~$5-10 amortized\n"
    "- Program management/supervision: ~$5-10\n"
    "Total ~$30/woman is my best guess for programs in hospitals that already have "
    "adequate neonatal care capacity. If hospitals need capacity building, costs "
    "could be $50-100+. See sensitivity analysis below.",
    "BOTEC"
)

ws.cell(5, 1, "Total cost per woman treated")
ws.cell(5, 2, "=B3+B4")
ws.cell(5, 2).number_format = "$#,##0.00"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "Number of women treated")
ws.cell(6, 2, "=B2/B5")
ws.cell(6, 2).number_format = "#,##0"
ws.cell(6, 3, "Calculation")

# ── Burden & effect ───────────────────────────────────────────────────────────
ws.cell(7, 1, "").font = section_font
ws.cell(8, 1, "Burden & effect").font = section_font

ws.cell(9, 1, "Baseline neonatal mortality among women at risk of early preterm birth")
ws.cell(9, 2, 0.235)
ws.cell(9, 2).number_format = "0.0%"
c = add_link(9, 3, "ACTION-I placebo arm: 331/1406 = 23.5%", ACTION_I_NEJM)
c.comment = Comment(
    'From ACTION-I trial (NEJM 2020): "Neonatal death occurred in [...] 331 of '
    '1406 infants (23.5%) in the placebo group." These are women at risk of early '
    'preterm birth (26-34 weeks) presenting at secondary/tertiary hospitals in '
    'Bangladesh, India, Kenya, Nigeria, and Pakistan. 90% actually delivered preterm.',
    "BOTEC"
)

ws.cell(10, 1, "RR for neonatal death (dexamethasone vs placebo)")
ws.cell(10, 2, 0.84)
c = add_link(10, 3, "ACTION-I trial (NEJM 2020): RR 0.84 (95% CI 0.72-0.97)", ACTION_I_NEJM)
c.comment = Comment(
    '"Neonatal death occurred in 278 of 1417 infants (19.6%) in the dexamethasone '
    'group, as compared with 331 of 1406 infants (23.5%) in the placebo group '
    '(relative risk, 0.84; 95% CI, 0.72 to 0.97; P=0.03)."\n'
    "https://www.nejm.org/doi/full/10.1056/NEJMoa2022398",
    "BOTEC"
)

ws.cell(11, 1, "Mortality reduction (1 - RR)")
ws.cell(11, 2, "=1-B10")
ws.cell(11, 2).number_format = "0%"
ws.cell(11, 3, "Calculation")

ws.cell(12, 1, "Baseline neonatal deaths in cohort")
ws.cell(12, 2, "=B6*B9")
ws.cell(12, 2).number_format = "#,##0.0"
ws.cell(12, 3, "Calculation")

ws.cell(13, 1, "Internal validity adjustment")
ws.cell(13, 2, 0.90)
ws.cell(13, 2).number_format = "0%"
c = ws.cell(13, 3, "Single large RCT (n=2852), stopped early for benefit. Minor discount for early stopping bias.")
c.comment = Comment(
    "The ACTION-I trial was a well-designed, double-blind, placebo-controlled RCT "
    "across 29 hospitals in 5 LMICs. However, it was stopped early for benefit at "
    "the second interim analysis. Trials stopped early for benefit tend to "
    "overestimate treatment effects (Bassler et al., JAMA 2010). A 10% IV discount "
    "accounts for this and for the fact that this is a single trial.",
    "BOTEC"
)

ws.cell(14, 1, "External validity adjustment")
ws.cell(14, 2, 0.80)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "Trial in selected hospitals with adequate neonatal care; real-world implementation may be less targeted.")
c.comment = Comment(
    "The ACTION-I trial was conducted in secondary/tertiary hospitals selected for "
    "having adequate neonatal care. Real-world ACS programs may:\n"
    "- Include hospitals with less adequate neonatal care (reducing effectiveness)\n"
    "- Have less rigorous gestational age assessment (increasing overtreatment risk)\n"
    "- Achieve lower compliance with the 4-dose regimen\n"
    "A 20% EV discount accounts for these implementation realities.",
    "BOTEC"
)

ws.cell(15, 1, "Adjusted neonatal deaths averted")
ws.cell(15, 2, "=B12*B11*B13*B14")
ws.cell(15, 2).number_format = "#,##0.0"
ws.cell(15, 3, "Calculation")

# ── Benefits ──────────────────────────────────────────────────────────────────
ws.cell(16, 1, "").font = section_font
ws.cell(17, 1, "Benefits").font = section_font

ws.cell(18, 1, "Moral weight per neonatal death averted (UoV)")
ws.cell(18, 2, 52.5)
ws.cell(18, 3, "GiveWell standard: under-5 death averted")

ws.cell(19, 1, "UoV from neonatal deaths averted")
ws.cell(19, 2, "=B15*B18")
ws.cell(19, 2).number_format = "#,##0"
ws.cell(19, 3, "Calculation")

ws.cell(20, 1, "Ad-hoc: morbidity reduction (reduced RDS, IVH, NEC among survivors)")
ws.cell(20, 2, 0.15)
ws.cell(20, 2).number_format = "0%"
c = ws.cell(20, 3, "ACS reduces RDS (RR 0.66), IVH (RR 0.55), NEC (RR 0.50) per Cochrane 2017.")
c.comment = Comment(
    "Beyond mortality, ACS substantially reduces neonatal morbidity:\n"
    "- RDS: RR 0.66 (Cochrane 2017)\n"
    "- IVH: RR 0.55\n"
    "- NEC: RR 0.50\n"
    "- Need for mechanical ventilation: RR 0.68\n"
    "These are major drivers of long-term disability and healthcare costs. "
    "15% morbidity uplift is conservative given the magnitude of these effects.",
    "BOTEC"
)

ws.cell(21, 1, "Ad-hoc: reduced healthcare costs (shorter NICU stays)")
ws.cell(21, 2, 0.10)
ws.cell(21, 2).number_format = "0%"
c = ws.cell(21, 3, "ACTION-I CEA: intervention was cost-saving in all 5 countries.")
c.comment = Comment(
    "The ACTION-I cost-effectiveness analysis found ACS was cost-saving in all "
    "5 trial countries, with savings of $1,778-$53,681 per 1,000 woman-baby units "
    "from reduced neonatal intensive care costs. 10% is a conservative credit "
    "for these healthcare cost savings.",
    "BOTEC"
)

ws.cell(22, 1, "Total UoV from program")
ws.cell(22, 2, "=B19*(1+B20+B21)")
ws.cell(22, 2).number_format = "#,##0"
ws.cell(22, 3, "Calculation")

# ── Final CE ──────────────────────────────────────────────────────────────────
ws.cell(23, 1, "").font = section_font
ws.cell(24, 1, "Final CE").font = section_font

ws.cell(25, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(25, 2, 0.00344)

ws.cell(26, 1, "CE (multiples of cash)")
ws.cell(26, 2, "=B22/B2/B25")
ws.cell(26, 2).number_format = "0.0"
ws.cell(26, 2).font = result_font
ws.cell(26, 3, "Calculation")

# ── Sensitivity ───────────────────────────────────────────────────────────────
ws.cell(27, 1, "").font = section_font
ws.cell(28, 1, "Sensitivity (CE at different program costs per woman)").font = section_font

ws.cell(29, 1, "CE at $10/woman (drug + minimal training overhead)")
ws.cell(29, 2, "=(B9*(1-B10)*B13*B14*B18*(1+B20+B21))/(10*B25)")
ws.cell(29, 2).number_format = "0.0"
ws.cell(29, 3, "Scenario: hospitals already have ultrasound and neonatal care; only need drug supply + minimal training")

ws.cell(30, 1, "CE at $30/woman (best guess)")
ws.cell(30, 2, "=(B9*(1-B10)*B13*B14*B18*(1+B20+B21))/(30*B25)")
ws.cell(30, 2).number_format = "0.0"
ws.cell(30, 3, "Scenario: moderate program (training + drug supply + supervision in equipped hospitals)")

ws.cell(31, 1, "CE at $50/woman (includes some capacity building)")
ws.cell(31, 2, "=(B9*(1-B10)*B13*B14*B18*(1+B20+B21))/(50*B25)")
ws.cell(31, 2).number_format = "0.0"
ws.cell(31, 3, "Scenario: hospitals need some additional neonatal care capacity")

ws.cell(32, 1, "CE at $100/woman (substantial capacity building)")
ws.cell(32, 2, "=(B9*(1-B10)*B13*B14*B18*(1+B20+B21))/(100*B25)")
ws.cell(32, 2).number_format = "0.0"
ws.cell(32, 3, "Scenario: hospitals need significant neonatal care upgrades + training + equipment")

ws.cell(33, 1, "CE at $30/woman with ACTION-III benefit (late preterm also effective, 3x eligible population)")
ws.cell(33, 2, "=B30")
ws.cell(33, 2).number_format = "0.0"
c = ws.cell(33, 3, "If ACTION-III shows benefit for 34-36 week births, CE unchanged per woman but eligible pool triples.")
c.comment = Comment(
    "If ACS is also effective for late preterm births (34-36 weeks), the number "
    "of eligible women roughly triples. The CE per woman treated doesn't change, "
    "but the total addressable market and RFMF increase substantially. The "
    "ACTION-III trial is testing this and results are expected ~2027.",
    "BOTEC"
)

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=33, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "antenatal_corticosteroids_preterm_birth.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
