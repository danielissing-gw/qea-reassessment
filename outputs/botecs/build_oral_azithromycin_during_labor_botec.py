import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

APLUS_NEJM = "https://www.nejm.org/doi/full/10.1056/NEJMoa2212111"
LANCET_CEA = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11950424/"
ORIGINAL_BOTEC = "https://docs.google.com/spreadsheets/d/171kRw2c0D5glTLNrGDGLuQTlM0gEMXxNnP7ibj1lEp0/edit"


def add_link(row, col, text, url):
    """Write a cell with a hyperlink."""
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Drug cost per dose (2g oral azithromycin)")
ws.cell(3, 2, 0.91)
ws.cell(3, 2).number_format = "$#,##0.00"
add_link(3, 3, "Lancet Global Health 2025 CEA: $0.91 per 2g dose", LANCET_CEA)

ws.cell(4, 1, "Incremental delivery cost per pregnancy (integration into facility care)")
ws.cell(4, 2, 2.09)
ws.cell(4, 2).number_format = "$#,##0.00"
c = ws.cell(4, 3, "Assumption: ~$2 incremental cost for adding one oral tablet to existing delivery care. "
    "Original QEA used $7 based on analogy to other facility-based interventions, but this "
    "overstates cost since no new patient visits are required.")
c.comment = Comment(
    "The drug cost is $0.91. The original QEA assumed $7/pregnancy total by analogy to "
    "dual HIV/syphilis testing and IPTi. However, since the woman is already in the facility "
    "for delivery, the incremental cost of adding one oral tablet should be much lower. "
    "I'm assuming ~$2 for minor supply chain/training costs on top of $0.91 drug cost, "
    "giving $3/pregnancy total. This is a key assumption that needs programmatic data to validate.",
    "BOTEC"
)

ws.cell(5, 1, "Total cost per pregnancy covered")
ws.cell(5, 2, "=B3+B4")
ws.cell(5, 2).number_format = "$#,##0.00"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "Number of pregnancies covered")
ws.cell(6, 2, "=B2/B5")
ws.cell(6, 2).number_format = "#,##0"
ws.cell(6, 3, "Calculation")

# ── Maternal mortality (Nigeria — high-MMR setting) ──────────────────────────
ws.cell(7, 1, "").font = section_font
ws.cell(8, 1, "Burden & effect (modeled for Nigeria)").font = section_font

ws.cell(9, 1, "Baseline facility-birth maternal mortality ratio (per 100,000)")
ws.cell(9, 2, 735.1)
add_link(9, 3, "From original QEA BOTEC, adjusted for facility births", ORIGINAL_BOTEC)

ws.cell(10, 1, "Maternal deaths in cohort (baseline)")
ws.cell(10, 2, "=B6*B9/100000")
ws.cell(10, 2).number_format = "#,##0.0"
ws.cell(10, 3, "Calculation")

ws.cell(11, 1, "RR for maternal sepsis/death (azithromycin vs placebo)")
ws.cell(11, 2, 0.67)
c = ws.cell(11, 3, "A-PLUS trial (NEJM 2023)")
c.hyperlink = APLUS_NEJM
c.font = link_font
c.comment = Comment(
    'Tita et al., NEJM 2023: "The incidence of the primary outcome of maternal sepsis '
    'or death was lower in the azithromycin group than in the placebo group (1.6% vs. '
    '2.4%; relative risk, 0.67; 95% CI, 0.56 to 0.79; P<0.001)."\n'
    'https://www.nejm.org/doi/full/10.1056/NEJMoa2212111',
    "BOTEC"
)

ws.cell(12, 1, "Mortality reduction (1 - RR)")
ws.cell(12, 2, "=1-B11")
ws.cell(12, 2).number_format = "0%"
ws.cell(12, 3, "Calculation")

ws.cell(13, 1, "Internal validity adjustment")
ws.cell(13, 2, 0.95)
ws.cell(13, 2).number_format = "0%"
ws.cell(13, 3, "Large multicenter RCT (n=29,278), 7 LMIC sites. Minor discount for single trial.")

ws.cell(14, 1, "External validity adjustment")
ws.cell(14, 2, 0.85)
ws.cell(14, 2).number_format = "0%"
ws.cell(14, 3, "Trial covered 7 LMICs including Nigeria. Discount for variation in facility quality and MMR composition.")

ws.cell(15, 1, "Adjusted maternal deaths averted")
ws.cell(15, 2, "=B10*B12*B13*B14")
ws.cell(15, 2).number_format = "#,##0.0"
ws.cell(15, 3, "Calculation")

# ── Neonatal mortality ───────────────────────────────────────────────────────
ws.cell(16, 1, "").font = section_font
ws.cell(17, 1, "Neonatal mortality (null result)").font = section_font

ws.cell(18, 1, "Neonatal deaths averted")
ws.cell(18, 2, 0)
c = ws.cell(18, 3, "A-PLUS: RR 1.02 (null). PregnAnZI-2: also null. No neonatal benefit.")
c.comment = Comment(
    'A-PLUS: "The incidence of the co-primary outcome of neonatal sepsis, stillbirth, '
    'or death was similar in the azithromycin group and placebo group (10.5% vs. 10.3%; '
    'relative risk, 1.02; 95% CI, 0.95 to 1.09)."\n'
    'PregnAnZI-2: "The overall incidence of the primary composite endpoint was similar '
    'in the azithromycin and in the placebo arm (2.0% vs 1.9%, p=0.70)"',
    "BOTEC"
)

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(19, 1, "").font = section_font
ws.cell(20, 1, "Benefits").font = section_font

ws.cell(21, 1, "Moral weight per maternal death averted (UoV)")
ws.cell(21, 2, 125)
ws.cell(21, 3, "From original QEA BOTEC (GiveWell standard)")

ws.cell(22, 1, "UoV from maternal deaths averted")
ws.cell(22, 2, "=B15*B21")
ws.cell(22, 2).number_format = "#,##0"
ws.cell(22, 3, "Calculation")

ws.cell(23, 1, "Ad-hoc: morbidity averted (maternal infections reduced)")
ws.cell(23, 2, 0.15)
ws.cell(23, 2).number_format = "0%"
c = ws.cell(23, 3, "A-PLUS also found RR 0.75 for maternal infection. Morbidity uplift for non-fatal infections averted.")
c.comment = Comment(
    "A-PLUS found reduced maternal infections (RR 0.75) beyond sepsis/death. "
    "This captures YLDs from non-fatal infections, reduced hospital readmissions, "
    "and reduced treatment costs. 15% is a rough estimate.",
    "BOTEC"
)

ws.cell(24, 1, "Ad-hoc: reduced healthcare costs (facility readmissions averted)")
ws.cell(24, 2, 0.05)
ws.cell(24, 2).number_format = "0%"
ws.cell(24, 3, "Lancet CEA: 248.5 readmissions averted per 100k pregnancies. Small offset.")

ws.cell(25, 1, "Total UoV from program")
ws.cell(25, 2, "=B22*(1+B23+B24)")
ws.cell(25, 2).number_format = "#,##0"
ws.cell(25, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(26, 1, "").font = section_font
ws.cell(27, 1, "Final CE").font = section_font

ws.cell(28, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(28, 2, 0.00344)

ws.cell(29, 1, "CE (multiples of cash)")
ws.cell(29, 2, "=B25/B2/B28")
ws.cell(29, 2).number_format = "0.0"
ws.cell(29, 2).font = result_font
ws.cell(29, 3, "Calculation")

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(30, 1, "").font = section_font
ws.cell(31, 1, "Sensitivity (CE at different costs per pregnancy)").font = section_font

ws.cell(32, 1, "CE at $1.50/pregnancy (drug + minimal integration)")
ws.cell(32, 2, "=(B9/100000*(1-B11)*B13*B14*B21*(1+B23+B24))/(1.5*B28)")
ws.cell(32, 2).number_format = "0.0"
ws.cell(32, 3, "Scenario: drug cost + very low integration overhead")

ws.cell(33, 1, "CE at $3/pregnancy (best guess)")
ws.cell(33, 2, "=(B9/100000*(1-B11)*B13*B14*B21*(1+B23+B24))/(3*B28)")
ws.cell(33, 2).number_format = "0.0"
ws.cell(33, 3, "Scenario: moderate integration cost (= main BOTEC)")

ws.cell(34, 1, "CE at $7/pregnancy (original QEA assumption)")
ws.cell(34, 2, "=(B9/100000*(1-B11)*B13*B14*B21*(1+B23+B24))/(7*B28)")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "Scenario: full original QEA cost assumption")

ws.cell(35, 1, "CE at $3/pregnancy in Pakistan (MMR 124.7)")
ws.cell(35, 2, "=(124.7/100000*(1-B11)*B13*B14*B21*(1+B23+B24))/(3*B28)")
ws.cell(35, 2).number_format = "0.0"
ws.cell(35, 3, "Lower-MMR setting: Pakistan (from original BOTEC)")

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=35, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "oral_azithromycin_during_labor.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
