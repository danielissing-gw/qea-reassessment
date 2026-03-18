"""Build single-sheet BOTEC for Rheumatic Heart Disease Prevention.

Best-guess CE estimate with:
  - Subjective 3x cost multiplier on Cuba data (original acknowledged as
    "possibly a gross underestimate" for India)
  - Ad-hoc uplifts for missing benefit streams (morbidity, perinatal)
  - Updated GBD 2021 burden for India

Produces outputs/botecs/rheumatic_heart_disease_prevention.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

# Source URLs
ORIGINAL_QEA = "https://docs.google.com/document/d/1siZVFAA3E3SLWroQ_oQqTuxUBMfEYvM_z8tL6eOp2Y8/edit"
ORIGINAL_BOTEC = "https://docs.google.com/spreadsheets/d/1FVr6y-dke7kJUIDbEt5Bqah4zPsc7gaBEqFki08DKaA/edit"
AHA_ABSTRACT = "https://www.ahajournals.org/doi/10.1161/circ.150.suppl_1.4144731"
WHO_GUIDELINE = "https://www.ncbi.nlm.nih.gov/books/NBK609692/"
BLS_CPI = "https://www.bls.gov/data/inflation_calculator.htm"
WHO_RHD_FACT = "https://www.who.int/news-room/fact-sheets/detail/rheumatic-heart-disease"


def add_link(row, col, text, url):
    """Write a cell with a hyperlink."""
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 65
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Burden & effect ──────────────────────────────────────────────────────────
ws.cell(1, 1, "Burden & effect").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Effect size: % reduction in RHD risk with prevention program")
ws.cell(2, 2, 0.89)
ws.cell(2, 2).number_format = "0%"
add_link(2, 3, "From Cuba observational study (Pinar del Rio 1986-1996)", ORIGINAL_QEA)
ws.cell(2, 3).comment = Comment(
    'Original paper abstract: "a decline in RHD prevalence from 2.27 patients '
    'per 1 000 children in 1986 to 0.24 per 1 000 in 1996" — implies ~89% '
    'reduction over 10 years, assuming decline is fully attributable to the '
    'intervention.\nSource: cited in original QEA.',
    "BOTEC"
)

ws.cell(3, 1, "Internal validity (IV) adjustment")
ws.cell(3, 2, 0.70)
ws.cell(3, 2).number_format = "0%"
ws.cell(3, 3, "Evidence base is observational, not RCT. Unchanged from original.")

ws.cell(4, 1, "External validity (EV) adjustment")
ws.cell(4, 2, 0.70)
ws.cell(4, 2).number_format = "0%"
add_link(4, 3, "Cuba→India extrapolation. Cuba has universal free healthcare; "
               "India does not. Unchanged from original.", ORIGINAL_QEA)

ws.cell(5, 1, "Combined validity-adjusted effect")
ws.cell(5, 2, "=B2*B3*B4")
ws.cell(5, 2).number_format = "0.0%"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "% of over-20 deaths caused by RHD in India")
ws.cell(6, 2, 0.0166)
ws.cell(6, 2).number_format = "0.00%"
add_link(6, 3, "GBD 2021: ~166,017 India RHD deaths / ~10M all-cause deaths. "
               "Original used 1.50% (129k deaths).", AHA_ABSTRACT)
ws.cell(6, 3).comment = Comment(
    'AHA Circulation 2023 abstract reports India had an estimated 166,017 RHD '
    'deaths in 2021 using GBD 2021 data. I could not verify this against the '
    'raw GBD database — medium confidence.\n' + AHA_ABSTRACT,
    "BOTEC"
)

ws.cell(7, 1, "Lifetime RHD deaths per cohort of 100,000")
ws.cell(7, 2, "=B6*100000")
ws.cell(7, 2).number_format = "#,##0"
ws.cell(7, 3, "Calculation")

ws.cell(8, 1, "Lifetime deaths averted per 100,000 with prevention program")
ws.cell(8, 2, "=B5*B7")
ws.cell(8, 2).number_format = "#,##0.0"
ws.cell(8, 3, "Calculation")

# ── Mortality benefit ────────────────────────────────────────────────────────
ws.cell(9, 1, "Mortality benefit").font = section_font

ws.cell(10, 1, "Units of value per RHD death averted (GW moral weights)")
ws.cell(10, 2, 52.5)
add_link(10, 3, "GiveWell moral weights for LMIC adult death. Unchanged.", ORIGINAL_BOTEC)

ws.cell(11, 1, "UoV from mortality per 100,000 cohort")
ws.cell(11, 2, "=B8*B10")
ws.cell(11, 2).number_format = "#,##0"
ws.cell(11, 3, "Calculation")

# ── Ad-hoc benefit adjustments ───────────────────────────────────────────────
ws.cell(12, 1, "Ad-hoc benefit adjustments (subjective)").font = section_font

ws.cell(13, 1, "Morbidity uplift: YLDs averted from RHD (% of mortality UoV)")
ws.cell(13, 2, 0.30)
ws.cell(13, 2).number_format = "0%"
ws.cell(13, 3, "I'm assuming +30%. RHD causes chronic heart failure, "
              "reduced exercise tolerance, and need for lifelong secondary "
              "prophylaxis. This is a rough guess — not verified against "
              "GBD YLD data.")

ws.cell(14, 1, "Perinatal outcomes uplift (% of mortality UoV)")
ws.cell(14, 2, 0.15)
ws.cell(14, 2).number_format = "0%"
add_link(14, 3, "I'm assuming +15%. RHD is 'the principal heart disease "
                "seen in pregnant women' per WHO; causes maternal and "
                "perinatal morbidity/mortality. Rough guess.", ORIGINAL_QEA)

ws.cell(15, 1, "Total UoV per 100,000 cohort (with adjustments)")
ws.cell(15, 2, "=B11*(1+B13+B14)")
ws.cell(15, 2).number_format = "#,##0"
ws.cell(15, 3, "Calculation: mortality UoV x (1 + morbidity + perinatal)")

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(16, 1, "Costs").font = section_font

ws.cell(17, 1, "Annual program cost in 2010 USD (Watkins et al. 2015)")
ws.cell(17, 2, 20289)
ws.cell(17, 2).number_format = "$#,##0"
add_link(17, 3, "From Cuba study. No new cost data found. "
                "GiveWell considered this 'possibly a gross underestimate'.", ORIGINAL_QEA)

ws.cell(18, 1, "CPI inflation factor (2010 → 2026)")
ws.cell(18, 2, 1.46)
add_link(18, 3, "I'm estimating ~1.46 for 2010→2026. "
                "Should verify with BLS CPI calculator.", BLS_CPI)

ws.cell(19, 1, "Annual program cost in 2026 USD (pre-adjustment)")
ws.cell(19, 2, "=B17*B18")
ws.cell(19, 2).number_format = "$#,##0"
ws.cell(19, 3, "Calculation")

ws.cell(20, 1, "Subjective cost multiplier")
ws.cell(20, 2, 3)
ws.cell(20, 3, "I'm assuming 3x. Cuba has universal free healthcare with "
               "existing primary care infrastructure; India would need to build "
               "screening, treatment, and referral systems. 3x is moderate — "
               "at 5x the CE is still ~16x (above bar). "
               "At 1x (original) CE is ~54x (almost certainly too high).")
ws.cell(20, 3).comment = Comment(
    'The original QEA stated: "I am concerned that the cost estimate (from '
    'the academic lit) may be an underestimate, possibly a gross one." and '
    'noted the intervention would still be 10x cash at 6x cost increase.\n'
    + ORIGINAL_QEA,
    "BOTEC"
)

ws.cell(21, 1, "Adjusted annual program cost in 2026 USD")
ws.cell(21, 2, "=B19*B20")
ws.cell(21, 2).number_format = "$#,##0"
ws.cell(21, 3, "Calculation")

ws.cell(22, 1, "Number covered by program annually (Watkins et al. 2015)")
ws.cell(22, 2, 273933)
ws.cell(22, 2).number_format = "#,##0"
add_link(22, 3, "From original Cuba study", ORIGINAL_BOTEC)

ws.cell(23, 1, "Adjusted annual cost per beneficiary")
ws.cell(23, 2, "=B21/B22")
ws.cell(23, 2).number_format = "$#,##0.00"
ws.cell(23, 3, "Calculation")

ws.cell(24, 1, "Duration of intervention per beneficiary (years)")
ws.cell(24, 2, 19)
ws.cell(24, 3, "Covers ages 5-24. Unchanged from original.")

ws.cell(25, 1, "Total cost per beneficiary (lifetime)")
ws.cell(25, 2, "=B23*B24")
ws.cell(25, 2).number_format = "$#,##0.00"
ws.cell(25, 3, "Calculation")

ws.cell(26, 1, "Total cost per 100,000 cohort")
ws.cell(26, 2, "=B25*100000")
ws.cell(26, 2).number_format = "$#,##0"
ws.cell(26, 3, "Calculation")

# ── Final CE ──────────────────────────────────────────────────────────────────
ws.cell(27, 1, "Final CE estimate").font = section_font

ws.cell(28, 1, "Units of value per $100,000 donated")
ws.cell(28, 2, "=B15/B26*100000")
ws.cell(28, 2).number_format = "#,##0"
ws.cell(28, 3, "Calculation")

ws.cell(29, 1, "GiveDirectly units of value per $100,000")
ws.cell(29, 2, 344)
add_link(29, 3, "GW benchmark", ORIGINAL_BOTEC)

ws.cell(30, 1, "Cost per death averted")
ws.cell(30, 2, "=B26/B8")
ws.cell(30, 2).number_format = "$#,##0"
ws.cell(30, 3, "Calculation (mortality only; does not include morbidity/perinatal)")

ws.cell(31, 1, "CE (multiples of cash)")
ws.cell(31, 2, "=B28/B29")
ws.cell(31, 2).number_format = "0.0"
ws.cell(31, 2).font = result_font
ws.cell(31, 3, "Best-guess estimate. Includes 3x cost multiplier and "
               "ad-hoc benefit adjustments. At 5x cost multiplier → ~16x. "
               "Confidence: Low (cost is pivotal unknown).")

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=31, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "rheumatic_heart_disease_prevention.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
