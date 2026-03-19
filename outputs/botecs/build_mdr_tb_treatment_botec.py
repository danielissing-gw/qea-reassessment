import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

TB_PRACTECAL = "https://www.thelancet.com/journals/lanres/article/PIIS2213-2600(23)00389-2/fulltext"
ENDTB = "https://endtb.org/endtb-clinical-trial-results"
NIX_TB = "https://www.nejm.org/doi/full/10.1056/NEJMoa1901814"
ZENIX = "https://www.nejm.org/doi/full/10.1056/NEJMoa2119430"
WHO_2022 = "https://www.who.int/publications/i/item/9789240063129"
WHO_TB_REPORT = "https://www.who.int/teams/global-programme-on-tuberculosis-and-lung-health/tb-reports/global-tuberculosis-report-2024/tb-diagnosis-and-treatment/2-4-drug-resistant-tb-treatment"
GDF_PRICING = "https://www.stoptb.org/news/stop-tbs-global-drug-facility-announces-38-price-decrease-bpalm-regimen-and-newly-available"
COST_SAVINGS = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11257096/"
GBD_MDR = "https://bmcmedicine.biomedcentral.com/articles/10.1186/s12916-025-04269-7"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Cost per patient treated (all-in: drugs + diagnostics + monitoring)")
ws.cell(3, 2, 1500)
ws.cell(3, 2).number_format = "$#,##0"
c = ws.cell(3, 3, "Mid-range of $996-$2,573 across 4 LMICs (Pakistan, Philippines, SA, Ukraine)")
c.comment = Comment(
    "Drug cost alone is $310/course (BPaLM via Global Drug Facility). "
    "Total treatment costs including health system costs: $996 (Pakistan) to $2,573 (Ukraine). "
    "Using $1,500 as mid-range estimate. If philanthropic funding covers drugs only ($310-750), "
    "CE would be much higher. See sensitivity analysis.",
    "BOTEC"
)

ws.cell(4, 1, "Patients treated")
ws.cell(4, 2, "=B2/B3")
ws.cell(4, 2).number_format = "#,##0"
ws.cell(4, 3, "Calculation")

# ── Treatment effect ─────────────────────────────────────────────────────────
ws.cell(5, 1, "")
ws.cell(6, 1, "Treatment effect").font = section_font

ws.cell(7, 1, "Treatment success rate (BPaLM)")
ws.cell(7, 2, 0.89)
ws.cell(7, 2).number_format = "0%"
c = add_link(7, 3, "TB-PRACTECAL trial (Lancet Resp Med 2023)", TB_PRACTECAL)
c.comment = Comment(
    '"Treatment success was achieved in 89% of participants randomised to BPaLM" '
    "(TB-PRACTECAL). Consistent with endTB Regimen 2 (90.4%), Nix-TB (90%), "
    "ZeNix 600mg/26wk (91%), and STREAM Stage 2 (91%). Using 89% as "
    "conservative central estimate (TB-PRACTECAL is the largest multi-country trial).",
    "BOTEC"
)

ws.cell(8, 1, "Mortality during treatment (1 - success - other failures)")
ws.cell(8, 2, 0.06)
ws.cell(8, 2).number_format = "0%"
c = ws.cell(8, 3, "Assumption: ~6% die during treatment (remainder relapse/lost to follow-up)")
c.comment = Comment(
    "In TB-PRACTECAL: 89% success, ~5% lost to follow-up/not evaluated, ~6% died. "
    "In endTB: unfavorable outcomes included death (primary), relapse, and treatment failure. "
    "6% in-treatment mortality is conservative for programmatic conditions "
    "(trial settings have better adherence support).",
    "BOTEC"
)

ws.cell(9, 1, "Counterfactual 5-year mortality without MDR-TB-specific treatment")
ws.cell(9, 2, 0.50)
ws.cell(9, 2).number_format = "0%"
c = ws.cell(9, 3, "Conservative estimate; Tiemersma 2011 found 70% 10-year TB mortality")
c.comment = Comment(
    "Tiemersma et al. 2011 estimated 70% of TB patients die within 10 years without treatment. "
    "MDR-TB patients specifically may have worse outcomes. However, some patients in the "
    "counterfactual would receive standard TB treatment (which is partially effective even "
    "for MDR-TB, yielding ~40-50% cure in some settings). Using 50% as a conservative "
    "estimate of 5-year mortality without MDR-TB-specific treatment.",
    "BOTEC"
)

ws.cell(10, 1, "Marginal lives saved per patient treated")
ws.cell(10, 2, "=B9-B8")
ws.cell(10, 2).number_format = "0.0%"
c = ws.cell(10, 3, "Calculation: counterfactual mortality - treatment mortality")
c.comment = Comment(
    "This is the net mortality reduction: 50% would die without treatment, 6% die with "
    "treatment, so each treated patient saves 0.44 lives on average. This is a simplification "
    "that ignores the timing of deaths and discount rates.",
    "BOTEC"
)

# ── Adjustments ──────────────────────────────────────────────────────────────
ws.cell(11, 1, "")
ws.cell(12, 1, "Adjustments").font = section_font

ws.cell(13, 1, "Internal validity adjustment")
ws.cell(13, 2, 0.90)
ws.cell(13, 2).number_format = "0%"
c = ws.cell(13, 3, "Multiple Phase 3 RCTs — high IV, minor discount for open-label designs")
c.comment = Comment(
    "TB-PRACTECAL, endTB, Nix-TB, ZeNix, and STREAM Stage 2 all demonstrate ~89-91% "
    "treatment success. Multiple independent trials across diverse settings. Minor discount "
    "(10%) because: (1) most trials were open-label; (2) 'treatment success' includes "
    "microbiological cure, not just mortality; (3) adherence support in trials exceeds "
    "typical programmatic conditions.",
    "BOTEC"
)

ws.cell(14, 1, "External validity adjustment")
ws.cell(14, 2, 0.80)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "Trial sites (SA, Uzbekistan, India) may differ from scaled programs")
c.comment = Comment(
    "TB-PRACTECAL sites were in Uzbekistan, Belarus, and South Africa — settings with "
    "relatively strong TB infrastructure. In countries with weaker health systems (e.g., DRC, "
    "Myanmar), treatment success may be lower due to: (1) drug supply interruptions; "
    "(2) less adherence support; (3) higher loss to follow-up; (4) fewer trained clinicians. "
    "Global actual treatment success rate is 68% (2021 cohort) vs. 89% in trials, "
    "suggesting a meaningful implementation gap. Using 80% EV.",
    "BOTEC"
)

ws.cell(15, 1, "Adjusted lives saved per patient treated")
ws.cell(15, 2, "=B10*B13*B14")
ws.cell(15, 2).number_format = "0.0%"
ws.cell(15, 3, "Calculation")

ws.cell(16, 1, "Total deaths averted")
ws.cell(16, 2, "=B4*B15")
ws.cell(16, 2).number_format = "#,##0.0"
ws.cell(16, 3, "Calculation")

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(17, 1, "")
ws.cell(18, 1, "Benefits").font = section_font

ws.cell(19, 1, "Mean age of MDR-TB death")
ws.cell(19, 2, 39)
c = add_link(19, 3, "GBD 2021 MDR-TB demographics: mean age ~39, 62.6% under 40", GBD_MDR)
c.comment = Comment(
    "GBD 2021 study: MDR-TB patients have mean age ~39-42, with 62.6% under 40. "
    "Peak incidence at ages 35-44. Significantly younger than non-MDR-TB patients.",
    "BOTEC"
)

ws.cell(20, 1, "Moral weight per MDR-TB death averted (UoV)")
ws.cell(20, 2, 40)
c = ws.cell(20, 3, "Working-age adults dying at ~39; remaining LE ~30-35 years in LMIC")
c.comment = Comment(
    "At age 39 in a typical LMIC (LE ~65-70), remaining life expectancy is ~30-35 years. "
    "GiveWell standard: 52.5 UoV for under-5 deaths. For adults dying at ~39, I'm estimating "
    "~40 UoV — lower than under-5 deaths but higher than elderly CVD deaths (~20-30 UoV). "
    "This is between the original aspirin BOTEC's 30.7 (older adults) and under-5 standard. "
    "76% of MDR-TB patients are male, which slightly reduces the moral weight per GW's "
    "framework (women have longer remaining LE).",
    "BOTEC"
)

ws.cell(21, 1, "UoV from deaths averted")
ws.cell(21, 2, "=B16*B20")
ws.cell(21, 2).number_format = "#,##0"
ws.cell(21, 3, "Calculation")

ws.cell(22, 1, "Ad-hoc: morbidity averted (reduced treatment duration, fewer side effects)")
ws.cell(22, 2, 0.10)
ws.cell(22, 2).number_format = "0%"
c = ws.cell(22, 3, "Assumption: 10% uplift for non-mortality benefits")
c.comment = Comment(
    "BPaLM is 6 months vs. 18-24 months for older regimens. Shorter treatment means: "
    "(1) less lost income during treatment; (2) fewer severe side effects (no injectable "
    "agents — eliminates hearing loss risk); (3) better quality of life during treatment; "
    "(4) reduced transmission due to faster cure. 10% is a conservative uplift for these "
    "non-mortality benefits.",
    "BOTEC"
)

ws.cell(23, 1, "Ad-hoc: transmission prevention (cured patients don't spread MDR-TB)")
ws.cell(23, 2, 0.15)
ws.cell(23, 2).number_format = "0%"
c = ws.cell(23, 3, "Assumption: 15% uplift for averted secondary cases")
c.comment = Comment(
    "Each untreated MDR-TB patient can infect 10-15 people over time, of whom ~5-10% "
    "develop active TB. A cured patient stops transmitting. Faster cure (6 months vs. "
    "18-24 months) also reduces the infectious period. 15% uplift is a rough estimate "
    "for the value of preventing secondary MDR-TB cases. This is likely conservative — "
    "some models estimate that 30-50% of the total value of TB treatment comes from "
    "preventing onward transmission.",
    "BOTEC"
)

ws.cell(24, 1, "Total UoV from program")
ws.cell(24, 2, "=B21*(1+B22+B23)")
ws.cell(24, 2).number_format = "#,##0"
ws.cell(24, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(25, 1, "")
ws.cell(26, 1, "Final CE").font = section_font

ws.cell(27, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(27, 2, 0.00344)

ws.cell(28, 1, "CE (multiples of cash)")
ws.cell(28, 2, "=B24/B2/B27")
ws.cell(28, 2).number_format = "0.0"
ws.cell(28, 2).font = result_font
ws.cell(28, 3, "Calculation")

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(29, 1, "")
ws.cell(30, 1, "Sensitivity").font = section_font

# Helper formula for sensitivity: all parameters same except cost per patient
# lives_per_patient = (B9-B8)*B13*B14
# UoV = (grant/cost) * lives_per_patient * B20 * (1+B22+B23)
# CE = UoV / (grant * 0.00344)

ws.cell(31, 1, "CE at $750/patient (drug procurement + basic support)")
ws.cell(31, 2, "=(B2/750)*B15*B20*(1+B22+B23)/(B2*B27)")
ws.cell(31, 2).number_format = "0.0"
ws.cell(31, 3, "Scenario: drugs ($310) + GeneXpert ($20) + monitoring ($420)")

ws.cell(32, 1, "CE at $1,000/patient (lean program)")
ws.cell(32, 2, "=(B2/1000)*B15*B20*(1+B22+B23)/(B2*B27)")
ws.cell(32, 2).number_format = "0.0"
ws.cell(32, 3, "Scenario: low-end of country-level cost estimates (Pakistan)")

ws.cell(33, 1, "CE at $1,500/patient (best guess)")
ws.cell(33, 2, "=B28")
ws.cell(33, 2).number_format = "0.0"
ws.cell(33, 3, "= main BOTEC estimate")

ws.cell(34, 1, "CE at $2,500/patient (full program with diagnostics)")
ws.cell(34, 2, "=(B2/2500)*B15*B20*(1+B22+B23)/(B2*B27)")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "Scenario: high-end country costs (Ukraine/South Africa)")

ws.cell(35, 1, "CE at $310/patient (drug procurement only)")
ws.cell(35, 2, "=(B2/310)*B15*B20*(1+B22+B23)/(B2*B27)")
ws.cell(35, 2).number_format = "0.0"
c = ws.cell(35, 3, "Scenario: filling acute drug stockouts from US funding cuts")
c.comment = Comment(
    "If the philanthropic role is purely drug procurement to prevent stockouts — "
    "with existing health systems covering diagnostics, monitoring, and clinical care — "
    "the per-patient cost is just $310 (BPaLM drug cost via GDF). This is the highest-leverage "
    "scenario and may be realistic in the context of the 2025 US funding crisis, where "
    "drug supply chains are disrupted but health infrastructure remains intact.",
    "BOTEC"
)

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=35, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "mdr_tb_treatment.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
