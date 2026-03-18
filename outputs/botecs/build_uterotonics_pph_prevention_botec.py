"""Build single-sheet BOTEC for Uterotonics for PPH Prevention/Treatment.

Standard GW CE model structure:
  1. $1M grant → births covered
  2. HSC prevention component: deaths averted
  3. TXA treatment component: deaths averted
  4. Total deaths averted → UoV via moral weights
  5. Ad-hoc morbidity/treatment-cost adjustments
  6. Final CE as multiples of benchmark

Produces outputs/botecs/uterotonics_pph_prevention.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

# Source URLs
ORIGINAL_QEA = "https://docs.google.com/document/d/1euCoBUuJQ8V8l7Mgl5SFP6UpcT2H7BioX8uRxoohpoM/edit"
WOMAN_TRIAL = "http://www.thelancet.com/journals/lancet/article/PIIS0140-6736(17)30638-4/fulltext"
LANCET_META = "https://www.thelancet.com/journals/lancet/article/PIIS0140-6736(24)02102-0/fulltext"
WHO_PPH_2025 = "https://www.ncbi.nlm.nih.gov/books/NBK619236/"
FERRING_HSC = ("https://www.ferring.com/ferring-statement-on-subsidised-pricing-of-"
               "heat-stable-carbetocin-for-the-prevention-of-postpartum-haemorrhage-"
               "in-low-and-lower-middle-income-countries/")
PMC_REVIEW = "https://pmc.ncbi.nlm.nih.gov/articles/PMC12145113/"
UNITAID = ("https://unitaid.org/project/expanding-access-to-recently-recommended-"
           "drugs-to-prevent-and-treat-postpartum-haemorrhage-pph/")
EMOTIVE = "https://www.nejm.org/doi/full/10.1056/NEJMoa2303966"


def add_link(row, col, text, url):
    """Write a cell with a hyperlink."""
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 70
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Grant & coverage ─────────────────────────────────────────────────────────
ws.cell(1, 1, "Grant & coverage").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"
ws.cell(2, 3, "Arbitrary anchor for calculation")

ws.cell(3, 1, "Cost per facility birth (bundled PPH program)")
ws.cell(3, 2, 2.00)
ws.cell(3, 2).number_format = "$#,##0.00"
ws.cell(3, 3, "I'm assuming $2/birth. Includes: HSC drug ($0.31), prorated TXA "
              "drug ($0.18), training increment ($0.70), supply chain ($0.35), "
              "monitoring ($0.35). Highly uncertain — at $1/birth CE roughly "
              "doubles; at $3/birth CE drops to ~2x. AMPLI-PPHI cost data "
              "(expected mid-2026) would resolve this.")
ws.cell(3, 3).comment = Comment(
    'Cost breakdown assumption:\n'
    '  HSC drug: $0.31 (Ferring subsidized price)\n'
    '  TXA drug (prorated): $3.00 x 6% PPH rate = $0.18\n'
    '  Training increment: ~$0.70 (short module for existing facility staff)\n'
    '  Supply chain: ~$0.35 (HSC is heat-stable; minimal cold chain)\n'
    '  Monitoring/supervision: ~$0.35\n'
    'Total: ~$1.89, rounded to $2.00.\n\n'
    'This assumes facilities and birth attendants already exist.\n'
    'For a standalone program (building facility capacity), costs '
    'would be much higher ($5-10/birth).',
    "BOTEC"
)

ws.cell(4, 1, "Facility births covered by grant")
ws.cell(4, 2, "=B2/B3")
ws.cell(4, 2).number_format = "#,##0"
ws.cell(4, 3, "Calculation")

# ── HSC prevention component ─────────────────────────────────────────────────
ws.cell(6, 1, "HSC prevention: deaths averted").font = section_font

ws.cell(7, 1, "Maternal lives saved per 100,000 facility births with HSC")
ws.cell(7, 2, 5)
add_link(7, 3, "India modeling study (2023), cited in PMC review. "
               "Compares HSC to settings without reliable uterotonic "
               "(cold-chain failure). I could not access the original study.",
         PMC_REVIEW)
ws.cell(7, 3).comment = Comment(
    'From PMC review: "prevent about 5,500 additional PPH cases and save '
    'five maternal lives per 100,000 births when priced comparably to '
    'oxytocin."\n' + PMC_REVIEW,
    "BOTEC"
)

ws.cell(8, 1, "IV adjustment (HSC)")
ws.cell(8, 2, 0.85)
ws.cell(8, 2).number_format = "0%"
ws.cell(8, 3, "Modeling study, not direct RCT in the target setting. "
              "Based on CHAMPION trial (HSC non-inferior to oxytocin) "
              "plus assumptions about cold-chain failure rates.")

ws.cell(9, 1, "EV adjustment (HSC)")
ws.cell(9, 2, 0.90)
ws.cell(9, 2).number_format = "0%"
ws.cell(9, 3, "India model applied to broader LMIC settings. "
              "Cold-chain reliability varies across settings.")

ws.cell(10, 1, "Adjusted lives saved per 100,000 births")
ws.cell(10, 2, "=B7*B8*B9")
ws.cell(10, 2).number_format = "0.00"
ws.cell(10, 3, "Calculation")

ws.cell(11, 1, "Deaths averted from HSC")
ws.cell(11, 2, "=B4/100000*B10")
ws.cell(11, 2).number_format = "0.0"
ws.cell(11, 3, "Calculation")

# ── TXA treatment component ──────────────────────────────────────────────────
ws.cell(13, 1, "TXA treatment: deaths averted").font = section_font

ws.cell(14, 1, "PPH rate among facility births")
ws.cell(14, 2, 0.06)
ws.cell(14, 2).number_format = "0%"
ws.cell(14, 3, "~6% is standard WHO estimate for PPH incidence "
               "(blood loss >= 500ml after birth).")

ws.cell(15, 1, "PPH cases from grant")
ws.cell(15, 2, "=B4*B14")
ws.cell(15, 2).number_format = "#,##0"
ws.cell(15, 3, "Calculation")

ws.cell(16, 1, "Mortality rate among PPH cases (standard care)")
ws.cell(16, 2, 0.019)
ws.cell(16, 2).number_format = "0.0%"
add_link(16, 3, "WOMAN trial placebo arm: 191/9,985 = 1.91%", WOMAN_TRIAL)
ws.cell(16, 3).comment = Comment(
    'WOMAN trial (2017): "Death due to bleeding was significantly reduced '
    'in women given tranexamic acid (155 [1.5%] of 10 036 patients vs '
    '191 [1.9%] of 9985 in the placebo group, risk ratio [RR] 0.81, '
    '95% CI 0.65-1.00; p=0.045)."\n' + WOMAN_TRIAL,
    "BOTEC"
)

ws.cell(17, 1, "Relative risk reduction from TXA (1 - RR)")
ws.cell(17, 2, 0.19)
ws.cell(17, 2).number_format = "0%"
add_link(17, 3, "WOMAN trial: RR 0.81 (95% CI 0.65-1.00) for death from "
               "bleeding. Using WOMAN trial as primary source for mortality. "
               "2024 Lancet IPD meta-analysis (OR 0.77 for life-threatening "
               "bleeding composite) is confirmatory.",
         WOMAN_TRIAL)
ws.cell(17, 3).comment = Comment(
    '2024 Lancet IPD meta-analysis (54,404 women, 5 RCTs): "Life-threatening '
    'bleeding occurred in 178 (0.65%) of 27,300 women in the tranexamic acid '
    'group versus 230 (0.85%) of 27,093 women in the placebo group (pooled '
    'odds ratio [OR] 0.77 [95% CI 0.63-0.93]; p=0.008)."\n\n'
    'Note: meta-analysis endpoint is composite (death + surgical interventions). '
    'WOMAN trial gives mortality-specific data; used as primary effect size.\n'
    + LANCET_META,
    "BOTEC"
)

ws.cell(18, 1, "IV adjustment (TXA)")
ws.cell(18, 2, 0.95)
ws.cell(18, 2).number_format = "0%"
ws.cell(18, 3, "Large, high-quality multi-country RCT (WOMAN trial). "
               "Confirmed by 2024 IPD meta-analysis. Small discount for "
               "field vs. trial conditions.")

ws.cell(19, 1, "EV adjustment (TXA)")
ws.cell(19, 2, 0.85)
ws.cell(19, 2).number_format = "0%"
ws.cell(19, 3, "WOMAN trial included LMIC sites, but TXA requires "
               "rapid PPH diagnosis and IV/IM administration within 3 hours. "
               "Effectiveness in field conditions may be lower than trial.")

ws.cell(20, 1, "Adjusted deaths averted per PPH case treated")
ws.cell(20, 2, "=B16*B17*B18*B19")
ws.cell(20, 2).number_format = "0.0000%"
ws.cell(20, 3, "Calculation: CFR x RRR x IV x EV")

ws.cell(21, 1, "Deaths averted from TXA")
ws.cell(21, 2, "=B15*B20")
ws.cell(21, 2).number_format = "0.0"
ws.cell(21, 3, "Calculation")

# ── Total benefits ────────────────────────────────────────────────────────────
ws.cell(23, 1, "Total benefits").font = section_font

ws.cell(24, 1, "Total deaths averted (HSC + TXA)")
ws.cell(24, 2, "=B11+B21")
ws.cell(24, 2).number_format = "0.0"
ws.cell(24, 3, "Calculation. Slight overcount: HSC prevents some PPH, "
               "reducing TXA treatment pool. Effect is small (~10%).")

ws.cell(25, 1, "UoV per maternal death averted (GW moral weights)")
ws.cell(25, 2, 52.5)
ws.cell(25, 3, "GiveWell standard moral weight for LMIC adult death. "
               "Maternal deaths are women aged 15-49.")

ws.cell(26, 1, "UoV from mortality")
ws.cell(26, 2, "=B24*B25")
ws.cell(26, 2).number_format = "#,##0"
ws.cell(26, 3, "Calculation")

ws.cell(27, 1, "Morbidity uplift (% of mortality UoV)")
ws.cell(27, 2, 0.75)
ws.cell(27, 2).number_format = "0%"
ws.cell(27, 3, "I'm assuming +75%. PPH causes severe anemia, need for "
               "blood transfusion, surgical interventions (hysterectomy, "
               "laparotomy), and long-term disability. For every PPH death, "
               "many women survive with significant morbidity. Rough guess.")

ws.cell(28, 1, "Treatment costs averted uplift (% of mortality UoV)")
ws.cell(28, 2, 0.25)
ws.cell(28, 2).number_format = "0%"
ws.cell(28, 3, "I'm assuming +25%. Blood transfusions are scarce and "
               "expensive in LMICs; surgical interventions carry high cost. "
               "Rough guess.")

ws.cell(29, 1, "Total UoV from grant")
ws.cell(29, 2, "=B26*(1+B27+B28)")
ws.cell(29, 2).number_format = "#,##0"
ws.cell(29, 3, "Calculation: mortality UoV x (1 + morbidity + treatment)")

# ── Final CE ──────────────────────────────────────────────────────────────────
ws.cell(31, 1, "Final CE estimate").font = section_font

ws.cell(32, 1, "UoV per $100,000 donated")
ws.cell(32, 2, "=B29/B2*100000")
ws.cell(32, 2).number_format = "#,##0"
ws.cell(32, 3, "Calculation")

ws.cell(33, 1, "GiveDirectly units of value per $100,000")
ws.cell(33, 2, 344)
ws.cell(33, 3, "GW benchmark (1x cash transfers)")

ws.cell(34, 1, "Cost per death averted")
ws.cell(34, 2, "=B2/B24")
ws.cell(34, 2).number_format = "$#,##0"
ws.cell(34, 3, "Calculation (mortality only; morbidity benefits additional)")

ws.cell(35, 1, "CE (multiples of cash)")
ws.cell(35, 2, "=B32/B33")
ws.cell(35, 2).number_format = "0.0"
ws.cell(35, 2).font = result_font
ws.cell(35, 3, "Best-guess estimate. Highly sensitive to cost/birth: "
               "at $1/birth -> ~6.5x; at $3/birth -> ~2.2x. "
               "Confidence: Low.")

# ── Context ───────────────────────────────────────────────────────────────────
ws.cell(37, 1, "Context & notes").font = section_font

ws.cell(38, 1, "E-MOTIVE trial (bundled detection + treatment)")
add_link(38, 3, "NEJM 2023: bundled PPH early detection + treatment reduced "
                "composite outcome (severe PPH, laparotomy, or death) by 60% "
                "(RR 0.40, 95% CI 0.32-0.50). Supports the bundled approach "
                "modeled above but uses a different intervention package.",
         EMOTIVE)
ws.cell(38, 3).comment = Comment(
    'E-MOTIVE (Gallos et al., NEJM 2023): "A primary-outcome event occurred '
    'in 1.6% of the patients in the intervention group, as compared with '
    '4.3% of those in the usual-care group (risk ratio, 0.40; 95% confidence '
    'interval [CI], 0.32 to 0.50; P<0.001)." 210,132 women across 80 '
    'hospitals in Kenya, Nigeria, South Africa, and Tanzania.\n' + EMOTIVE,
    "BOTEC"
)

ws.cell(39, 1, "Misoprostol: off the table")
add_link(39, 3, "WHO 2025: misoprostol is 'last resort' when no injectable "
                "options are available", WHO_PPH_2025)

ws.cell(40, 1, "Key cost data source")
add_link(40, 3, "AMPLI-PPHI ($26M, 6 countries, 2022-2026) is generating "
                "cost-effectiveness evidence. Results expected by mid-2026.",
         UNITAID)

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=40, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__),
                        "uterotonics_pph_prevention.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
