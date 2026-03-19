import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

HEALS_2010 = "https://pubmed.ncbi.nlm.nih.gov/20646756/"
HEALS_2025 = "https://www.nih.gov/news-events/nih-research-matters/arsenic-reduction-linked-lower-risk-death"
VAN_GEEN = "https://pmc.ncbi.nlm.nih.gov/articles/PMC6535723/"
INFANT_META = "https://pmc.ncbi.nlm.nih.gov/articles/PMC4421764/"
FLANAGAN_WHO = "https://pmc.ncbi.nlm.nih.gov/articles/PMC3506399/"
WORLD_BANK_BD = "https://data.worldbank.org/indicator/SP.DYN.CBRT.IN?locations=BD"
ORIGINAL_BOTEC = "https://docs.google.com/spreadsheets/d/1ATbZ61dmczX3GB3Mo-nwITfsMcTikpGmlzG-D83xu5Y/"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 65
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Cost per well tested (all-in: kit + labor + education + placard)")
ws.cell(3, 2, 5)
ws.cell(3, 2).number_format = "$#,##0"
c = ws.cell(3, 3, "Assumption: field kit $0.17-1.00 + labor/logistics/education ~$4-5")
c.comment = Comment(
    "Field kit costs: $0.17/test bulk UNICEF price, $0.60/test retail. "
    "Full informational intervention: <$10/household (Barnwal et al. 2023). "
    "Van Geen: $0.90/person whose exposure was reduced (~$10/well). "
    "$5/well is a mid-range assumption between bare-minimum testing "
    "and a comprehensive education program. Key assumption.",
    "BOTEC"
)

ws.cell(4, 1, "Wells tested")
ws.cell(4, 2, "=B2/B3")
ws.cell(4, 2).number_format = "#,##0"
ws.cell(4, 3, "Calculation")

ws.cell(5, 1, "People per well")
ws.cell(5, 2, 11)
add_link(5, 3, "From original BOTEC", ORIGINAL_BOTEC)

ws.cell(6, 1, "Total people covered by testing")
ws.cell(6, 2, "=B4*B5")
ws.cell(6, 2).number_format = "#,##0"
ws.cell(6, 3, "Calculation")

# ── Contamination ────────────────────────────────────────────────────────────
ws.cell(7, 1, "")
ws.cell(8, 1, "Contamination & behavioral response").font = section_font

ws.cell(9, 1, "% of wells with As 50-150 ug/L")
ws.cell(9, 2, 0.086)
ws.cell(9, 2).number_format = "0.0%"
add_link(9, 3, "From original BOTEC (Bangladesh distribution)", ORIGINAL_BOTEC)

ws.cell(10, 1, "% of wells with As 150+ ug/L")
ws.cell(10, 2, 0.048)
ws.cell(10, 2).number_format = "0.0%"
ws.cell(10, 3, "From original BOTEC")

ws.cell(11, 1, "People exposed at 50-150 ug/L")
ws.cell(11, 2, "=B6*B9")
ws.cell(11, 2).number_format = "#,##0"
ws.cell(11, 3, "Calculation")

ws.cell(12, 1, "People exposed at 150+ ug/L")
ws.cell(12, 2, "=B6*B10")
ws.cell(12, 2).number_format = "#,##0"
ws.cell(12, 3, "Calculation")

ws.cell(13, 1, "Total people at high-exposure wells")
ws.cell(13, 2, "=B11+B12")
ws.cell(13, 2).number_format = "#,##0"
ws.cell(13, 3, "Calculation")

ws.cell(14, 1, "% of high-exposure population induced to switch to safe well")
ws.cell(14, 2, 0.45)
ws.cell(14, 2).number_format = "0%"
c = ws.cell(14, 3, "Updated from 35% — multiple studies find 37-60%")
c.comment = Comment(
    "Araihazar (with placards/counseling): 60% within 1 year. "
    "Singair cluster RCT: 53%. Barnwal 2023 informational intervention: 37%. "
    "BAMWSP national program (without sustained follow-up): 29%. "
    "45% is a conservative central estimate given education is included in the program.",
    "BOTEC"
)

# ── Adult mortality (ages 15-49) ─────────────────────────────────────────────
ws.cell(15, 1, "")
ws.cell(16, 1, "Adult mortality (ages 15-49)").font = section_font

ws.cell(17, 1, "Annual mortality rate, ages 15-49 (Bangladesh)")
ws.cell(17, 2, 0.0013)
ws.cell(17, 2).number_format = "0.00%"
add_link(17, 3, "From original BOTEC (Bangladesh vital statistics)", ORIGINAL_BOTEC)

ws.cell(18, 1, "HR for As 50-150 ug/L (vs. <10 ug/L)")
ws.cell(18, 2, 1.09)
ws.cell(18, 2).number_format = "0.00"
c = add_link(18, 3, "HEALS cohort (Argos et al., Lancet 2010)", HEALS_2010)
c.comment = Comment(
    '"Multivariate adjusted hazard ratios for all-cause mortality comparing '
    'arsenic at concentrations of [...] 50.1-150.0 ug/L [...] with a reference '
    'of at least 10.0 ug/L in well water were [...] 1.09 (0.81-1.47)." '
    "Note: 95% CI crosses 1.0 — this HR is not statistically significant.",
    "BOTEC"
)

ws.cell(19, 1, "HR for As 150+ ug/L (vs. <10 ug/L)")
ws.cell(19, 2, 1.68)
ws.cell(19, 2).number_format = "0.00"
c = add_link(19, 3, "HEALS cohort (Argos et al., Lancet 2010)", HEALS_2010)
c.comment = Comment(
    '"Multivariate adjusted hazard ratios [...] 150.1-864.0 ug/L [...] '
    '1.68 (1.26-2.23)." Statistically significant.',
    "BOTEC"
)

ws.cell(20, 1, "IV adjustment (observational evidence only)")
ws.cell(20, 2, 0.70)
ws.cell(20, 2).number_format = "0%"
c = ws.cell(20, 3, "No RCT on mortality; HEALS is strong observational (20yr, n=10,977)")
c.comment = Comment(
    "HEALS is a high-quality prospective cohort (20-year follow-up, 10,977 adults, "
    "1,401 chronic disease deaths). The 2025 JAMA paper strengthens the causal case "
    "by showing dose-response and temporal consistency. But confounding cannot be "
    "fully excluded — healthier/wealthier people may be more likely to switch wells. "
    "70% IV reflects strong-but-not-experimental evidence.",
    "BOTEC"
)

ws.cell(21, 1, "Annual adult (15-49) deaths averted from switching (50-150 group)")
ws.cell(21, 2, "=B11*B14*B17*(B18-1)*B20")
ws.cell(21, 2).number_format = "#,##0.0"
ws.cell(21, 3, "= N_exposed × switch_rate × mortality × (HR-1) × IV")

ws.cell(22, 1, "Annual adult (15-49) deaths averted from switching (150+ group)")
ws.cell(22, 2, "=B12*B14*B17*(B19-1)*B20")
ws.cell(22, 2).number_format = "#,##0.0"
ws.cell(22, 3, "Calculation")

ws.cell(23, 1, "Total annual adult (15-49) deaths averted")
ws.cell(23, 2, "=B21+B22")
ws.cell(23, 2).number_format = "#,##0.0"
ws.cell(23, 3, "Calculation")

# ── Adult mortality (ages 50+) ───────────────────────────────────────────────
ws.cell(24, 1, "")
ws.cell(25, 1, "Adult chronic disease mortality (ages 50+) — new benefit stream").font = section_font

ws.cell(26, 1, "% of exposed population aged 50+")
ws.cell(26, 2, 0.15)
ws.cell(26, 2).number_format = "0%"
ws.cell(26, 3, "Assumption: ~15% of Bangladesh population is 50+")

ws.cell(27, 1, "People aged 50+ at high-exposure wells")
ws.cell(27, 2, "=B13*B26")
ws.cell(27, 2).number_format = "#,##0"
ws.cell(27, 3, "Calculation")

ws.cell(28, 1, "Annual chronic disease mortality rate (ages 50+, Bangladesh)")
ws.cell(28, 2, 0.012)
ws.cell(28, 2).number_format = "0.0%"
ws.cell(28, 3, "Assumption: ~1.2% (60% of ~2% all-cause mortality is chronic disease)")

ws.cell(29, 1, "Reduction in chronic disease mortality from switching (from HEALS 2025)")
ws.cell(29, 2, 0.22)
ws.cell(29, 2).number_format = "0%"
c = add_link(29, 3, "HEALS 20-year follow-up (JAMA 2025)", HEALS_2025)
c.comment = Comment(
    '"For every 200 ug/gram creatinine decrease in urinary arsenic levels: '
    '22% reduced chronic disease mortality, 20% reduced cancer mortality, '
    '23% reduced heart disease mortality." '
    "Using 22% as a conservative estimate; high-to-low switchers saw 54% reduction.",
    "BOTEC"
)

ws.cell(30, 1, "IV adjustment (same as above)")
ws.cell(30, 2, 0.70)
ws.cell(30, 2).number_format = "0%"
ws.cell(30, 3, "Same 70% observational evidence discount")

ws.cell(31, 1, "Annual deaths averted (50+ chronic disease)")
ws.cell(31, 2, "=B27*B14*B28*B29*B30")
ws.cell(31, 2).number_format = "#,##0.0"
ws.cell(31, 3, "= N_50+ × switch_rate × mortality × reduction × IV")

# ── Infant mortality ─────────────────────────────────────────────────────────
ws.cell(32, 1, "")
ws.cell(33, 1, "Infant mortality").font = section_font

ws.cell(34, 1, "Birth rate per 1,000 people (Bangladesh, annual)")
ws.cell(34, 2, 18.2)
add_link(34, 3, "World Bank", WORLD_BANK_BD)

ws.cell(35, 1, "Births among high-exposure population (annual)")
ws.cell(35, 2, "=B13*B34/1000")
ws.cell(35, 2).number_format = "#,##0"
ws.cell(35, 3, "Calculation")

ws.cell(36, 1, "Infant mortality rate (Bangladesh)")
ws.cell(36, 2, 0.0256)
ws.cell(36, 2).number_format = "0.00%"
ws.cell(36, 3, "From original BOTEC (World Bank 2019)")

ws.cell(37, 1, "OR for infant mortality with arsenic exposure (>=50 ug/L)")
ws.cell(37, 2, 1.35)
c = add_link(37, 3, "Meta-analysis (Rahman et al.)", INFANT_META)
c.comment = Comment(
    '"Arsenic in groundwater (>=50 ug/L) associated with [...] '
    'infant mortality OR = 1.35." From original BOTEC.',
    "BOTEC"
)

ws.cell(38, 1, "IV adjustment for infant mortality (observational)")
ws.cell(38, 2, 0.70)
ws.cell(38, 2).number_format = "0%"
ws.cell(38, 3, "Same 70% discount as adult mortality")

ws.cell(39, 1, "Annual infant deaths averted from switching")
ws.cell(39, 2, "=B35*B14*B36*(B37-1)*B38")
ws.cell(39, 2).number_format = "#,##0.0"
ws.cell(39, 3, "= births × switch_rate × IMR × (OR-1) × IV")

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(40, 1, "")
ws.cell(41, 1, "Benefits").font = section_font

ws.cell(42, 1, "Moral weight per adult (15-49) death averted (UoV)")
ws.cell(42, 2, 50)
c = ws.cell(42, 3, "Adults dying at age ~30-50; remaining LE ~25-40 years")
c.comment = Comment(
    "Original BOTEC used 106 UoV (labeled 'guess'). GW standard for adult deaths "
    "is typically 30-50 UoV depending on age. Using 50 as these are relatively "
    "young adults (15-49) dying prematurely.",
    "BOTEC"
)

ws.cell(43, 1, "Moral weight per adult (50+) death averted (UoV)")
ws.cell(43, 2, 20)
c = ws.cell(43, 3, "Older adults; remaining LE ~10-20 years")
c.comment = Comment(
    "Adults dying of chronic disease at 50-70 have shorter remaining life "
    "expectancy. GW typically values these deaths at 15-30 UoV. Using 20.",
    "BOTEC"
)

ws.cell(44, 1, "Moral weight per infant death averted (UoV)")
ws.cell(44, 2, 52.5)
ws.cell(44, 3, "GiveWell standard for under-5 deaths")

ws.cell(45, 1, "Annual UoV from adult (15-49) deaths averted")
ws.cell(45, 2, "=B23*B42")
ws.cell(45, 2).number_format = "#,##0"
ws.cell(45, 3, "Calculation")

ws.cell(46, 1, "Annual UoV from adult (50+) deaths averted")
ws.cell(46, 2, "=B31*B43")
ws.cell(46, 2).number_format = "#,##0"
ws.cell(46, 3, "Calculation")

ws.cell(47, 1, "Annual UoV from infant deaths averted")
ws.cell(47, 2, "=B39*B44")
ws.cell(47, 2).number_format = "#,##0"
ws.cell(47, 3, "Calculation")

ws.cell(48, 1, "Duration of benefits (years)")
ws.cell(48, 2, 8)
c = ws.cell(48, 3, "Updated from 5 — Araihazar data shows persistent switching over 8-15 years")
c.comment = Comment(
    "Pfaff et al. 2017: switching from 2003-2005 was 'highly persistent' through "
    "2015. New switching continued to occur years after testing. However, "
    "22% of households forgot results by year 8, and new untested wells may be "
    "installed. 8 years is a moderate estimate.",
    "BOTEC"
)

ws.cell(49, 1, "Total lifetime UoV from program")
ws.cell(49, 2, "=(B45+B46+B47)*B48")
ws.cell(49, 2).number_format = "#,##0"
ws.cell(49, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(50, 1, "")
ws.cell(51, 1, "Final CE").font = section_font

ws.cell(52, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(52, 2, 0.00344)

ws.cell(53, 1, "CE (multiples of cash)")
ws.cell(53, 2, "=B49/B2/B52")
ws.cell(53, 2).number_format = "0.0"
ws.cell(53, 2).font = result_font
ws.cell(53, 3, "Calculation")

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(54, 1, "")
ws.cell(55, 1, "Sensitivity").font = section_font

# For sensitivity at different costs, I need to compute the full chain
# UoV_annual = UoV_adult_1549 + UoV_adult_50 + UoV_infant
# All scale linearly with B4 (wells tested) = B2/B3
# So CE = (annual_UoV_per_well * wells * duration) / (grant * 0.00344)
# CE scales as 1/cost_per_well

ws.cell(56, 1, "CE at $3/well (bare-minimum testing)")
ws.cell(56, 2, "=B53*B3/3")
ws.cell(56, 2).number_format = "0.0"
ws.cell(56, 3, "Scenario: kit cost + minimal labor only")

ws.cell(57, 1, "CE at $5/well (best guess)")
ws.cell(57, 2, "=B53")
ws.cell(57, 2).number_format = "0.0"
ws.cell(57, 3, "= main BOTEC estimate")

ws.cell(58, 1, "CE at $10/well (full education program)")
ws.cell(58, 2, "=B53*B3/10")
ws.cell(58, 2).number_format = "0.0"
ws.cell(58, 3, "Scenario: comprehensive intervention per Barnwal et al.")

ws.cell(59, 1, "CE at $10/well, IV 50% (pessimistic)")
ws.cell(59, 2, "=B53*B3/10*0.5/B20")
ws.cell(59, 2).number_format = "0.0"
ws.cell(59, 3, "Scenario: higher cost + heavier observational discount")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=59, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "arsenic_in_wells.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
