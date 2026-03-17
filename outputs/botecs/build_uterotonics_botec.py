import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

wb = openpyxl.Workbook()

# ── PARAMETERS SHEET ──────────────────────────────────────────────────────────
ws_p = wb.active
ws_p.title = "Parameters"

header_font = Font(bold=True)
input_fill  = PatternFill("solid", fgColor="DDEEFF")
calc_fill   = PatternFill("solid", fgColor="F0F0F0")
note_fill   = PatternFill("solid", fgColor="FFF8DC")

ws_p.cell(row=1, column=1, value="Parameter").font = header_font
ws_p.cell(row=1, column=2, value="Original value").font = header_font
ws_p.cell(row=1, column=3, value="Updated value").font = header_font
ws_p.cell(row=1, column=4, value="Source").font = header_font
ws_p.cell(row=1, column=5, value="Notes").font = header_font

# NOTE: original BOTEC was incomplete (only one cost row for TXA).
# This BOTEC models TXA treatment for PPH, not misoprostol prevention.
# A separate model for heat-stable carbetocin is noted but not built here.

rows = [
    # Section A: TXA treatment for PPH
    ("── SECTION A: Tranexamic Acid (TXA) for PPH Treatment ──", None, None, "", ""),

    ("PPH mortality rate with standard care (placebo arm, WOMAN trial)",
     0.019, 0.019,
     "https://pmc.ncbi.nlm.nih.gov/articles/PMC2864262/",
     "WOMAN trial placebo: 191/9,985 = 1.91%. Rounded to 1.9%. Unchanged."),

    ("Relative risk of death from bleeding with TXA vs placebo (WOMAN trial)",
     None, 0.81,
     "http://www.thelancet.com/journals/lancet/article/PIIS0140-6736(17)30638-4/fulltext",
     "Original BOTEC did not model this. WOMAN trial RR = 0.81 (95% CI 0.65–1.00). "
     "2024 Lancet meta-analysis reports 31% mortality reduction within 3 hours "
     "(source: secondary; could not access full Lancet 2024 paper)."),

    ("Absolute risk reduction in death per PPH case treated (= row 2 × (1 - row 3))",
     None, None,
     "calculation",
     ""),

    ("TXA drug cost per patient (generic, LMIC, 2024 estimate)",
     17.48, 3.00,
     "assumption",
     "Original: $17.48 from Guerriero et al. 2011 (Tanzania, includes hospital stay and nurse labor — not drug alone). "
     "Updated: I'm estimating $1–5 for TXA drug alone (generic, off-patent); $3 is a rough midpoint. "
     "No specific current LMIC price source found — this needs verification."),

    ("Program cost multiplier (training, supply chain, monitoring, over drug cost)",
     None, 10.0,
     "assumption",
     "I'm assuming program delivery costs 10x the drug cost, as a rough midpoint. "
     "This is highly uncertain — AMPLI-PPHI program may have real data by 2026."),

    ("Total program cost per PPH case treated",
     None, None,
     "calculation",
     "= drug cost × program multiplier"),

    ("Cost per death averted (= total cost per case / absolute risk reduction)",
     None, None,
     "calculation",
     ""),

    ("GiveWell CE benchmark: cost per death averted equivalent for GiveDirectly",
     None, 4500,
     "assumption",
     "I'm assuming GiveDirectly is ~$4,500/death averted equivalent (rough estimate; "
     "GiveWell's exact figure depends on moral weights and year). Needs verification."),

    ("Multiples of GiveDirectly (= benchmark / cost per death averted)",
     None, None,
     "calculation",
     ""),

    ("── SECTION B: Heat-Stable Carbetocin (HSC) for PPH Prevention ──", None, None, "", ""),

    ("HSC drug cost per birth (LMIC public sector, Ferring subsidized price)",
     None, 0.31,
     "https://www.ghsupplychain.org/sites/default/files/2023-07/HSC_20230630.pdf",
     "Ferring provides at $0.31 ± 10% per ampoule for public sector in LMICs. Not in original QEA."),

    ("Additional PPH cases prevented per 100,000 facility births (vs standard care)",
     None, 5500,
     "assumption",
     "From a 2023 India modeling study cited in PMC review (https://pmc.ncbi.nlm.nih.gov/articles/PMC12145113/). "
     "I could not access the full India study — treating with medium confidence."),

    ("Maternal lives saved per 100,000 births with HSC (vs standard care)",
     None, 5,
     "assumption",
     "Same India modeling study. Highly context-dependent."),

    ("Drug cost per life saved (= drug cost × 100,000 births / lives saved)",
     None, None,
     "calculation",
     "= row 13 × 100,000 / row 15"),

    ("Estimated full program cost per life saved (assuming program = 20x drug cost)",
     None, None,
     "calculation",
     "= row 16 × 20. I'm assuming 20x drug cost for full program; highly uncertain."),

    ("Multiples of GiveDirectly (= benchmark / full program cost per life saved)",
     None, None,
     "calculation",
     "= row 9 / row 17"),
]

for r, (param, orig, upd, src, notes) in enumerate(rows, 2):
    ws_p.cell(row=r, column=1, value=param)
    if param.startswith("──"):
        ws_p.cell(row=r, column=1).font = Font(bold=True)
        continue
    if orig is not None:
        ws_p.cell(row=r, column=2, value=orig).fill = input_fill
    if upd is not None and src != "calculation":
        ws_p.cell(row=r, column=3, value=upd).fill = input_fill
    elif src == "calculation":
        ws_p.cell(row=r, column=3, value="see Calculation sheet").fill = calc_fill
    ws_p.cell(row=r, column=4, value=src)
    ws_p.cell(row=r, column=5, value=notes)

ws_p.column_dimensions["A"].width = 65
ws_p.column_dimensions["B"].width = 18
ws_p.column_dimensions["C"].width = 22
ws_p.column_dimensions["D"].width = 60
ws_p.column_dimensions["E"].width = 80

# ── CALCULATION SHEET ──────────────────────────────────────────────────────────
ws_c = wb.create_sheet("Calculation")

# Section A: TXA
ws_c["A1"] = "SECTION A: TXA for PPH Treatment"
ws_c["A1"].font = Font(bold=True, size=12)

txA_rows = [
    ("Description", "Formula / Value", "Notes"),
    ("PPH mortality rate (placebo)", "=Parameters!C3", "1.9%"),
    ("RR of death with TXA", "=Parameters!C4", "0.81"),
    ("Absolute risk reduction (deaths per PPH case)", "=C3*(1-C4)", "= mortality rate × (1 - RR)"),
    ("TXA drug cost per patient ($)", "=Parameters!C6", ""),
    ("Program cost multiplier", "=Parameters!C7", ""),
    ("Total cost per PPH case treated ($)", "=C6*C7", "drug × multiplier"),
    ("Cost per death averted ($)", "=C8/C5", "total cost / ARR"),
    ("GiveDirectly benchmark ($/death averted)", "=Parameters!C10", ""),
    ("MULTIPLES OF CASH — TXA", "=C10/C9", ""),
]

for r, row_data in enumerate(txA_rows, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws_c.cell(row=r, column=c, value=val)
        if r == 2:
            cell.font = Font(bold=True)
        if r == len(txA_rows) + 1 and c == 2:
            cell.font = Font(bold=True, size=12)

# Section B: HSC
offset = len(txA_rows) + 4
ws_c.cell(row=offset, column=1, value="SECTION B: Heat-Stable Carbetocin for PPH Prevention").font = Font(bold=True, size=12)

hsc_rows = [
    ("Description", "Formula / Value", "Notes"),
    ("HSC drug cost per birth ($)", "=Parameters!C13", "$0.31"),
    ("Lives saved per 100,000 births", "=Parameters!C15", "5"),
    ("Drug cost per life saved ($)", "=Parameters!C13*100000/Parameters!C15", ""),
    ("Program cost per life saved (drug × 20x)", "=C" + str(offset+3) + "*20", "Very rough; assuming 20x drug cost for program"),
    ("GiveDirectly benchmark ($/death averted)", "=Parameters!C10", ""),
    ("MULTIPLES OF CASH — HSC", "=C" + str(offset+5) + "/C" + str(offset+4), ""),
]

for r, row_data in enumerate(hsc_rows, offset + 1):
    for c, val in enumerate(row_data, 1):
        cell = ws_c.cell(row=r, column=c, value=val)
        if r == offset + 1:
            cell.font = Font(bold=True)
        if r == offset + len(hsc_rows) and c == 2:
            cell.font = Font(bold=True, size=12)

ws_c.column_dimensions["A"].width = 45
ws_c.column_dimensions["B"].width = 30
ws_c.column_dimensions["C"].width = 45

# ── NOTES SHEET ───────────────────────────────────────────────────────────────
ws_n = wb.create_sheet("Notes")
notes = [
    ("Key caveats",),
    ("",),
    ("1. Misoprostol for PPH prevention is NOT modeled here.",
     "The original QEA focused on misoprostol (PSI program). WHO 2025 now designates misoprostol as a 'last resort' "
     "when no injectable options are available. No mortality evidence exists. This intervention is effectively off the table."),
    ("",),
    ("2. TXA model has very high uncertainty.",
     "The $3 drug cost and 10x program multiplier are rough assumptions. "
     "AMPLI-PPHI (Jhpiego/Unitaid, 2022-2026) should have real cost data; that program is generating CE evidence."),
    ("",),
    ("3. Heat-stable carbetocin was not in the original QEA.",
     "HSC is the most important new development in the PPH space since 2017. "
     "At $0.31/ampoule and WHO-recommended, it should be modeled separately with full program cost data."),
    ("",),
    ("4. Major implementers are now active.",
     "Unitaid AMPLI-PPHI ($26M, 2022-2026) is scaling up exactly these interventions in 6 countries. "
     "Room for more funding depends on geographic gaps and whether this program addresses the whole opportunity."),
]
for r, row in enumerate(notes, 1):
    ws_n.cell(row=r, column=1, value=row[0]).font = Font(bold=(r == 1 or (len(row[0]) > 2 and row[0][0].isdigit())))
    if len(row) > 1:
        ws_n.cell(row=r, column=2, value=row[1])
ws_n.column_dimensions["A"].width = 40
ws_n.column_dimensions["B"].width = 90

out_path = os.path.join(os.path.dirname(__file__), "uterotonics_pph_prevention.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
