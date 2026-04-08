import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

PANIGRAHI = "https://www.nature.com/articles/nature23480"
FLEISCHMANN = "https://pmc.ncbi.nlm.nih.gov/articles/PMC8311109/"
BARNARDS = "https://pubmed.ncbi.nlm.nih.gov/35427523/"
PROSYNK_COST = "https://www.medrxiv.org/content/10.1101/2024.12.17.24319142v1.full"
PELL_BANGLADESH = "https://pubmed.ncbi.nlm.nih.gov/39992135/"
GBD_SEPSIS = "https://www.frontiersin.org/journals/medicine/articles/10.3389/fmed.2025.1536948/full"
GW_MORAL_WEIGHTS = "https://docs.google.com/spreadsheets/d/1YiZmHQb-JsDIDJF_tKKVte2x5NOL0CGUTndJNoB0iHo/"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"

ws.cell(3, 1, "Drug cost per synbiotic course")
ws.cell(3, 2, 1)
ws.cell(3, 2).number_format = "$#,##0.00"
c = add_link(3, 3, "Panigrahi trial: ~$1/course for 7-day regimen", PANIGRAHI)
c.comment = Comment(
    "The synbiotic consists of Lactobacillus plantarum (ATCC 202195) + "
    "fructooligosaccharide, given orally for 7 days starting in the first "
    "days of life. Drug cost ~$1 per course based on Panigrahi et al. 2017 "
    "media coverage. Novonesis (formerly Chr. Hansen) manufactures LP202195 "
    "(product name PPLP-217) and supplies for research trials.",
    "BOTEC"
)

ws.cell(4, 1, "Delivery/overhead costs per course")
ws.cell(4, 2, 4)
ws.cell(4, 2).number_format = "$#,##0.00"
c = ws.cell(4, 3, "Assumption: $4 delivery overhead for a scaled 7-day regimen")
c.comment = Comment(
    "PROSYNK Kenya trial found programmatic cost of $9.15/infant for a "
    "32-dose/6-month regimen — but this is a much longer protocol than "
    "Panigrahi's 7-day course. For a 7-day regimen delivered through "
    "existing community health worker platforms (e.g., postnatal home "
    "visits), delivery costs would be lower. I'm assuming $4/course "
    "overhead ($5 total), which is between the original $1 overhead "
    "and PROSYNK's ~$8 overhead. This is uncertain — at scale through "
    "government platforms, could be lower; in a new standalone program, "
    "could be higher.",
    "BOTEC"
)

ws.cell(5, 1, "Total cost per synbiotic course")
ws.cell(5, 2, "=B3+B4")
ws.cell(5, 2).number_format = "$#,##0.00"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "Neonates treated")
ws.cell(6, 2, "=B2/B5")
ws.cell(6, 2).number_format = "#,##0"
ws.cell(6, 3, "Calculation")

# ── Burden & effect ──────────────────────────────────────────────────────────
ws.cell(7, 1, "")
ws.cell(8, 1, "Burden & effect").font = section_font

ws.cell(9, 1, "Baseline risk of death or sepsis (composite)")
ws.cell(9, 2, 0.0904)
ws.cell(9, 2).number_format = "0.00%"
c = add_link(9, 3, "Panigrahi 2017: 206/2,278 in control group", PANIGRAHI)
c.comment = Comment(
    'Panigrahi et al. 2017 (Nature, n=4,556): "206 [death or culture-positive '
    'sepsis events] in the control group [n=2,278]." Baseline risk = 206/2,278 '
    "= 9.04%. Note: the original BOTEC author flagged that 'Neonatal mortality "
    "rates were very high in the study setting. The baseline risk of sepsis "
    "and/or death may be much lower in other areas.' The EV adjustment "
    "accounts for this.",
    "BOTEC"
)

ws.cell(10, 1, "RR for death or sepsis (synbiotic vs. placebo)")
ws.cell(10, 2, 0.60)
ws.cell(10, 2).number_format = "0.00"
c = add_link(10, 3, "Panigrahi 2017: RR 0.60 (95% CI 0.48-0.74)", PANIGRAHI)
c.comment = Comment(
    'Panigrahi et al. 2017: "The relative risk of the primary outcome '
    "[sepsis or death] was 0.60 (95% CI 0.48-0.74).\" Based on a single "
    "large RCT (n=4,556). The trial also found significant reduction in "
    "lower respiratory infections (RR 0.66) and local infections (RR 0.70). "
    "A corrigendum was published (Nature 2017) correcting statements about "
    "culture-positive vs. culture-negative cases, but the primary endpoint "
    "result was unchanged.",
    "BOTEC"
)

ws.cell(11, 1, "Absolute risk reduction")
ws.cell(11, 2, "=B9*(1-B10)")
ws.cell(11, 2).number_format = "0.00%"
ws.cell(11, 3, "Calculation")

ws.cell(12, 1, "Composite death/sepsis events prevented")
ws.cell(12, 2, "=B6*B11")
ws.cell(12, 2).number_format = "#,##0"
ws.cell(12, 3, "Calculation")

ws.cell(13, 1, "Neonatal sepsis case fatality rate (CFR)")
ws.cell(13, 2, 0.15)
ws.cell(13, 2).number_format = "0%"
c = add_link(13, 3, "Midpoint of systematic review estimates (10-20%)", FLEISCHMANN)
c.comment = Comment(
    "The original BOTEC used 5% as an acknowledged guess. Fleischmann et al. "
    "2021 (systematic review, 26 studies, 2.8M live births): pooled CFR "
    '17.6% (95% CI 10.3-28.6%). BARNARDS study (Lancet Global Health 2022, '
    "7 LMICs): 13.3% for lab-confirmed sepsis. pSBI pooled estimate: 9.8%. "
    "Using 15% as a central estimate, roughly the midpoint of BARNARDS "
    "(13.3%) and the global pooled (17.6%). This is the key correction from "
    "the original BOTEC — the 5% figure was 2-3x too low.",
    "BOTEC"
)

ws.cell(14, 1, "Deaths before adjustments")
ws.cell(14, 2, "=B12*B13")
ws.cell(14, 2).number_format = "#,##0"
ws.cell(14, 3, "Calculation: composite events prevented × CFR")

ws.cell(15, 1, "Internal validity adjustment")
ws.cell(15, 2, 0.70)
ws.cell(15, 2).number_format = "0%"
c = ws.cell(15, 3, "Single large RCT (Nature, n=4,556) but unreplicated; corrigendum")
c.comment = Comment(
    "The Panigrahi trial is well-powered and published in Nature, which "
    "supports high IV. But: (1) single trial with no clinical efficacy "
    "replication in 8+ years; (2) a corrigendum was published; (3) the "
    "Bangladesh Phase 2 trial (Pell et al. 2025) failed to replicate "
    "persistent gut colonization — raising doubt about the mechanism "
    "in other settings. 70% IV reflects these concerns while still "
    "crediting the trial's RCT design and strong statistical significance "
    "(95% CI 0.48-0.74).",
    "BOTEC"
)

ws.cell(16, 1, "External validity adjustment")
ws.cell(16, 2, 0.60)
ws.cell(16, 2).number_format = "0%"
c = ws.cell(16, 3, "Rural Odisha, India → other LMIC settings; colonization persistence concern")
c.comment = Comment(
    "The trial was in rural Odisha, India — a high neonatal mortality "
    "setting. Key EV concerns: (1) the Bangladesh trial found LP202195 "
    "colonization 'did not persist beyond 2 months,' inconsistent with "
    "6-month persistence in India; (2) baseline sepsis risk (9.04%) may "
    "be higher than in other LMIC settings (global average ~3.8%); "
    "(3) gut microbiome composition varies across populations and may "
    "affect colonization success. 60% EV reflects the serious "
    "uncertainty about whether the clinical effect generalizes.",
    "BOTEC"
)

ws.cell(17, 1, "Deaths averted (adjusted)")
ws.cell(17, 2, "=B14*B15*B16")
ws.cell(17, 2).number_format = "#,##0"
ws.cell(17, 3, "Calculation")

# ── Benefits ─────────────────────────────────────────────────────────────────
ws.cell(18, 1, "")
ws.cell(19, 1, "Benefits").font = section_font

ws.cell(20, 1, "Moral weight per neonatal death averted (UoV)")
ws.cell(20, 2, 84)
c = add_link(20, 3, "GiveWell 2020 moral weights, neonatal age bracket", GW_MORAL_WEIGHTS)
c.comment = Comment(
    "Most neonatal sepsis deaths occur in the first 7 days (early neonatal: "
    "MW 83.6) with some in days 7-28 (late neonatal: MW 84.5). Using 84 "
    "as weighted average (heavier weight on early neonatal given that "
    "early-onset sepsis accounts for ~66% of neonatal sepsis deaths per "
    "GBD 2021: 136k early-onset vs 70k late-onset).",
    "BOTEC"
)

ws.cell(21, 1, "UoV from deaths averted")
ws.cell(21, 2, "=B17*B20")
ws.cell(21, 2).number_format = "#,##0"
ws.cell(21, 3, "Calculation")

ws.cell(22, 1, "Ad-hoc: morbidity uplift (developmental sequelae in sepsis survivors)")
ws.cell(22, 2, 0.10)
ws.cell(22, 2).number_format = "0%"
c = ws.cell(22, 3, "Assumption: 10% uplift for non-fatal benefits")
c.comment = Comment(
    "Neonatal sepsis survivors have elevated risk of neurodevelopmental "
    "impairment, hearing loss, and growth faltering. A meta-analysis found "
    "~23% of neonatal sepsis survivors have at least one major "
    "neurodevelopmental impairment. Preventing sepsis cases averts these "
    "long-term sequelae. 10% is a conservative morbidity uplift "
    "acknowledging that most BOTEC value comes from deaths averted.",
    "BOTEC"
)

ws.cell(23, 1, "Total UoV from program")
ws.cell(23, 2, "=B21*(1+B22)")
ws.cell(23, 2).number_format = "#,##0"
ws.cell(23, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────────
ws.cell(24, 1, "")
ws.cell(25, 1, "Final CE").font = section_font

ws.cell(26, 1, "Value per dollar spent on benchmark (GiveDirectly)")
ws.cell(26, 2, 0.00344)

ws.cell(27, 1, "CE (multiples of cash)")
ws.cell(27, 2, "=B23/B2/B26")
ws.cell(27, 2).number_format = "0.0"
ws.cell(27, 2).font = result_font
ws.cell(27, 3, "Calculation")

ws.cell(28, 1, "Cost per death averted (implied)")
ws.cell(28, 2, "=B2/B17")
ws.cell(28, 2).number_format = "$#,##0"
ws.cell(28, 3, "Cross-check")

# ── Sensitivity ──────────────────────────────────────────────────────────────
ws.cell(29, 1, "")
ws.cell(30, 1, "Sensitivity").font = section_font

# Sensitivity: cost per course
ws.cell(31, 1, "CE at $2/course (original Indian cost)")
ws.cell(31, 2, "=(B2/2*B11*B13*B15*B16*B20*(1+B22))/(B2*B26)")
ws.cell(31, 2).number_format = "0.0"
ws.cell(31, 3, "Drug $1 + minimal overhead $1")

ws.cell(32, 1, "CE at $5/course (central estimate)")
ws.cell(32, 2, "=B27")
ws.cell(32, 2).number_format = "0.0"
ws.cell(32, 3, "= main BOTEC estimate")

ws.cell(33, 1, "CE at $9/course (PROSYNK Kenya programmatic cost)")
ws.cell(33, 2, "=(B2/9*B11*B13*B15*B16*B20*(1+B22))/(B2*B26)")
ws.cell(33, 2).number_format = "0.0"
add_link(33, 3, "PROSYNK cost study (32-dose/6-month regimen)", PROSYNK_COST)

# Sensitivity: CFR
ws.cell(34, 1, "CE at 10% CFR (conservative)")
ws.cell(34, 2, "=(B2/B5*B11*0.10*B15*B16*B20*(1+B22))/(B2*B26)")
ws.cell(34, 2).number_format = "0.0"
ws.cell(34, 3, "Lower bound of systematic review range")

ws.cell(35, 1, "CE at 15% CFR (central estimate)")
ws.cell(35, 2, "=B27")
ws.cell(35, 2).number_format = "0.0"
ws.cell(35, 3, "= main BOTEC estimate")

ws.cell(36, 1, "CE at 20% CFR (higher estimate)")
ws.cell(36, 2, "=(B2/B5*B11*0.20*B15*B16*B20*(1+B22))/(B2*B26)")
ws.cell(36, 2).number_format = "0.0"
ws.cell(36, 3, "Upper-central range (Fleischmann pooled: 17.6%)")

# Sensitivity: heavier discounts for replication failure
ws.cell(37, 1, "CE at 50% IV / 50% EV (heavy replication discount)")
ws.cell(37, 2, "=(B2/B5*B11*B13*0.50*0.50*B20*(1+B22))/(B2*B26)")
ws.cell(37, 2).number_format = "0.0"
ws.cell(37, 3, "If Bangladesh colonization failure warrants deeper skepticism")

# Wrap text
for row in ws.iter_rows(min_row=1, max_row=37, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "synbiotics_neonatal_sepsis.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
