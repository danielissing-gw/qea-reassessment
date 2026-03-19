import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

COCHRANE_URL = "https://www.cochranelibrary.com/cdsr/doi/10.1002/14651858.CD004905.pub6/full"
WHO_MMS_URL = "https://www.who.int/publications-detail-redirect/9789240007789"
KIRK_URL = "https://kirkhumanitarian.org/unimmap-mms-product-attributes/"
COPENHAGEN_URL = "https://hmhbconsortium.org/brief-mms-copenhagen-consensus/"
GROWTH_META_URL = "https://www.sciencedirect.com/science/article/pii/S0002916525002382"
ADHERENCE_META_URL = "https://www.sciencedirect.com/science/article/pii/S2161831325000912"


def add_link(row, col, text, url):
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "Grant amount")
ws.cell(2, 2, 1000000)
ws.cell(2, 2).number_format = "$#,##0"
ws.cell(2, 3, "Arbitrary anchor")

ws.cell(3, 1, "MMS tablet cost per pregnancy (180 tablets)")
ws.cell(3, 2, 2.13)
ws.cell(3, 2).number_format = "$#,##0.00"
c = add_link(3, 3, "Kirk Humanitarian procurement: $0.0118/dose x 180 tablets", KIRK_URL)
c.comment = Comment(
    'UNIMMAP MMS is produced at $0.0118 per dose or $2.13 per woman per pregnancy '
    'when packaged in a 180-count bottle.',
    "BOTEC"
)

ws.cell(4, 1, "IFA tablet cost per pregnancy (180 tablets)")
ws.cell(4, 2, 1.80)
ws.cell(4, 2).number_format = "$#,##0.00"
c = ws.cell(4, 3, "Estimate; IFA typically ~$0.01/tablet")
c.comment = Comment(
    'IFA costs vary by source. UNICEF Supply Catalogue lists various formulations '
    'in the range of $0.005-0.015/tablet. Using $0.01/tablet x 180 = $1.80 as a '
    'reasonable estimate. The original BOTEC used an incremental cost of $0.004878/tablet '
    'from Adams et al. 2019 academic CEA.',
    "BOTEC"
)

ws.cell(5, 1, "Incremental commodity cost per pregnancy (MMS - IFA)")
ws.cell(5, 2, "=B3-B4")
ws.cell(5, 2).number_format = "$#,##0.00"
ws.cell(5, 3, "Calculation")

ws.cell(6, 1, "Programmatic/transition costs per pregnancy")
ws.cell(6, 2, 2.67)
ws.cell(6, 2).number_format = "$#,##0.00"
c = ws.cell(6, 3, "Estimate; original QEA used $3. Reduced given Kirk scale-up experience.")
c.comment = Comment(
    'The original BOTEC used $3/person for "Programmatic costs of transition/TA per '
    'person, tablet waste, etc." and labeled it a guess. Kirk Humanitarian has now '
    'scaled MMS to 75M women in 111 LMICs, suggesting transition costs may be lower '
    'than feared. I am reducing from $3 to $2.67 but note this is still an assumption '
    'I have not verified against actual program cost data.',
    "BOTEC"
)

ws.cell(7, 1, "Total incremental cost per pregnancy (MMS vs IFA)")
ws.cell(7, 2, "=B5+B6")
ws.cell(7, 2).number_format = "$#,##0.00"
ws.cell(7, 3, "Calculation")

ws.cell(8, 1, "Pregnancies covered by grant")
ws.cell(8, 2, "=B2/B7")
ws.cell(8, 2).number_format = "#,##0"
ws.cell(8, 3, "Calculation")

# ── Burden & effect ──────────────────────────────────────────────────────
ws.cell(10, 1, "Burden & effect").font = section_font

ws.cell(11, 1, "Baseline LBW rate (South Asia)")
ws.cell(11, 2, 0.28)
ws.cell(11, 2).number_format = "0%"
c = ws.cell(11, 3, "South Asia has the highest LBW rates globally (~28%)")
c.comment = Comment(
    'From original BOTEC, sourced to WHO Global Health Observatory data. South Asia '
    'is the region with the highest rates of LBW, making it the natural target for '
    'this intervention.',
    "BOTEC"
)

ws.cell(12, 1, "Relative risk of LBW with MMS vs IFA")
ws.cell(12, 2, 0.88)
c = add_link(12, 3, "2019 Cochrane review: RR 0.88 (95% CI 0.85-0.91), 18 trials, I²=0%", COCHRANE_URL)
c.comment = Comment(
    '"MMN reduced the number of newborn infants identified as low birthweight (LBW) '
    '(average RR 0.88, 95% CI 0.85 to 0.91; 18 trials, 68,801 participants; '
    'high-quality evidence)." Confirmed by 2024-2025 meta-analyses.',
    "BOTEC"
)

ws.cell(13, 1, "Percent reduction in LBW")
ws.cell(13, 2, "=1-B12")
ws.cell(13, 2).number_format = "0%"
ws.cell(13, 3, "Calculation")

ws.cell(14, 1, "Expected LBW births averted")
ws.cell(14, 2, "=B8*B11*B13")
ws.cell(14, 2).number_format = "#,##0"
ws.cell(14, 3, "Calculation")

# ── Benefits ─────────────────────────────────────────────────────────────
ws.cell(16, 1, "Benefits").font = section_font

ws.cell(17, 1, "Moral weight per LBW averted (UoV)")
ws.cell(17, 2, 3)
c = ws.cell(17, 3, "Rough guess from original QEA (borrowed from syphilis CEA)")
c.comment = Comment(
    'The original BOTEC used MW=3 as a "super rough guess, from syphilis CEA." '
    'Value comes mainly through birthweight\'s effect on development. The 2025 '
    'meta-analysis showing MMS reduces stunting (RR 0.86) through 24 months '
    'provides direct evidence for a developmental benefit channel, supporting '
    'this assumption. However, MW=3 remains uncertain and is the second-largest '
    'driver of CE after cost.',
    "BOTEC"
)

ws.cell(18, 1, "UoV from LBW reduction")
ws.cell(18, 2, "=B14*B17")
ws.cell(18, 2).number_format = "#,##0"
ws.cell(18, 3, "Calculation")

ws.cell(19, 1, "Adjustment for excluded child effects")
ws.cell(19, 2, 1.3)
c = ws.cell(19, 3, "Lower morbidity, developmental effects beyond LBW (original estimate)")
c.comment = Comment(
    'From original BOTEC. "This might include lower morbidity both early and later '
    'in life, adult disability, etc." Unchanged from original. The new stunting evidence '
    'supports there being real developmental effects, but these are likely already '
    'captured in the MW for LBW of 3.',
    "BOTEC"
)

ws.cell(20, 1, "UoV before IV/EV adjustments")
ws.cell(20, 2, "=B18*B19")
ws.cell(20, 2).number_format = "#,##0"
ws.cell(20, 3, "Calculation")

# ── Adjustments ──────────────────────────────────────────────────────────
ws.cell(22, 1, "Adjustments").font = section_font

ws.cell(23, 1, "Internal validity adjustment")
ws.cell(23, 2, 1.0)
ws.cell(23, 2).number_format = "0%"
c = ws.cell(23, 3, "High-quality evidence: 18 RCTs, I²=0%, Cochrane rated high quality")
c.comment = Comment(
    'The Cochrane review rated the LBW evidence as "high quality" with I²=0% '
    'across 18 trials and 68,801 participants. No IV discount warranted.',
    "BOTEC"
)

ws.cell(24, 1, "External validity adjustment")
ws.cell(24, 2, 0.75)
ws.cell(24, 2).number_format = "0%"
c = ws.cell(24, 3, "Adherence concern partially addressed by scale-up (original: 0.7)")
c.comment = Comment(
    'The original QEA used EV=0.7: "My guess would be that, given this requires '
    'taking 180 pills, adherence would be lower in field implementation than in '
    'real-world settings." Kirk Humanitarian\'s scale-up to 75M women in 111 LMICs '
    'partially addresses this concern by demonstrating feasibility at scale. The '
    '2025 IPD meta-analysis confirms that adherence matters (>=90% adherence needed '
    'for full benefit). I am slightly increasing from 0.7 to 0.75 but note this is '
    'still an assumption.',
    "BOTEC"
)

ws.cell(25, 1, "UoV after adjustments")
ws.cell(25, 2, "=B20*B23*B24")
ws.cell(25, 2).number_format = "#,##0"
ws.cell(25, 3, "Calculation")

# ── Final CE ─────────────────────────────────────────────────────────────
ws.cell(27, 1, "Final CE").font = section_font

ws.cell(28, 1, "GW benchmark (UoV per $100k at 1x cash)")
ws.cell(28, 2, 344)
ws.cell(28, 3, "344 UoV per $100k = 1x cash transfers")

ws.cell(29, 1, "CE (best guess)")
ws.cell(29, 2, "=B25/(B2/100000*B28)")
ws.cell(29, 2).number_format = "0.0"
ws.cell(29, 2).font = result_font
ws.cell(29, 3, "Calculation: total UoV / (grant/100k * benchmark)")

# ── Sensitivity: cost per pregnancy ──────────────────────────────────────
ws.cell(31, 1, "Sensitivity: CE at different costs per pregnancy").font = section_font

ws.cell(32, 1, "CE at $2/pregnancy (efficient IFA→MMS switch)")
ws.cell(32, 2, "=(B2/2)*B11*B13*B17*B19*B23*B24/(B2/100000*B28)")
ws.cell(32, 2).number_format = "0.0"

ws.cell(33, 1, "CE at $3/pregnancy (best guess)")
ws.cell(33, 2, "=(B2/3)*B11*B13*B17*B19*B23*B24/(B2/100000*B28)")
ws.cell(33, 2).number_format = "0.0"

ws.cell(34, 1, "CE at $4/pregnancy (original QEA estimate)")
ws.cell(34, 2, "=(B2/4)*B11*B13*B17*B19*B23*B24/(B2/100000*B28)")
ws.cell(34, 2).number_format = "0.0"

ws.cell(35, 1, "CE at $6/pregnancy (comprehensive new program)")
ws.cell(35, 2, "=(B2/6)*B11*B13*B17*B19*B23*B24/(B2/100000*B28)")
ws.cell(35, 2).number_format = "0.0"

# ── Sensitivity: moral weight for LBW ───────────────────────────────────
ws.cell(37, 1, "Sensitivity: CE at different moral weights for LBW").font = section_font

ws.cell(38, 1, "CE at MW=2 for LBW")
ws.cell(38, 2, "=B8*B11*B13*2*B19*B23*B24/(B2/100000*B28)")
ws.cell(38, 2).number_format = "0.0"

ws.cell(39, 1, "CE at MW=3 for LBW (best guess)")
ws.cell(39, 2, "=B29")
ws.cell(39, 2).number_format = "0.0"

ws.cell(40, 1, "CE at MW=4 for LBW")
ws.cell(40, 2, "=B8*B11*B13*4*B19*B23*B24/(B2/100000*B28)")
ws.cell(40, 2).number_format = "0.0"

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=40, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "mms_during_pregnancy.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
