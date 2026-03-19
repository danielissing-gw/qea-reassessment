import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
header = Font(bold=True, size=12)
wrap = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions["A"].width = 55
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 60

r = 1

def add(label, value, source, is_header=False, is_calc=False):
    global r
    ws.cell(r, 1, label).font = header if is_header else Font()
    ws.cell(r, 1).alignment = wrap
    if value is not None:
        ws.cell(r, 2, value)
    if source:
        ws.cell(r, 3, source).alignment = wrap
    r += 1

# ── Section: Burden & effect ──
add("Burden & effect", "Value", "Source / notes", is_header=True)

add("Counterfactual mortality rate among women with obstetric hemorrhage",
    0.016,
    "RCT control group mortality 2.3% (Miller et al. 2013), adjusted down 31% for country-level maternal hemorrhage mortality reductions between 2010-2019 (GBD). Result: 1.6%.")

add("Mortality reduction from NASG (pooled non-experimental)",
    0.48,
    "Pooled RR 0.52 (95% CI 0.36-0.77) from random effects meta-analysis of 5 observational/pre-post studies (Pileggi-Castro et al. 2015 systematic review).")

add("Internal validity (IV) adjustment",
    0.66,
    "Evidence is mostly observational (4 pre/post + 1 fully observational study). One underpowered RCT (OR 0.36, NS). Potential confounding from differences in care quality between study phases. Original QEA applied 66%.")

add("External validity (EV) adjustment",
    0.75,
    "Implementation likely better quality in study settings than real-world. Scale-up data from Zimbabwe (90% usage) is encouraging. Original QEA applied 75%.")

add("Combined IV/EV-adjusted mortality reduction",
    "=B2*B4*B5",
    "Calculation: 48% × 66% × 75% = ~24%")

add("Deaths averted per woman treated",
    "=B2*B6",
    "Calculation: 1.6% mortality × 24% adjusted reduction = ~0.38%")

# ── Section: Mortality benefit ──
add("", None, None)
add("Mortality benefit", None, None, is_header=True)

add("Units of value per maternal death averted (GW moral weights)",
    109.5,
    "GiveWell moral weights for adult female death at average age 27-30. Used in original BOTEC.")

add("UoV from mortality per woman treated",
    "=B7*B10",
    "Calculation")

# ── Section: Ad-hoc adjustments ──
add("", None, None)
add("Ad-hoc benefit adjustments (subjective)", None, None, is_header=True)

add("Morbidity uplift (% of mortality UoV)",
    0.15,
    "I'm assuming +15%. NASG reduces severe anemia at discharge and speeds recovery from shock (165 min vs 209 min in RCT). Also reduces need for emergency hysterectomy. Rough guess.")

add("Infant mortality uplift (% of mortality UoV)",
    0.10,
    "I'm assuming +10%. Maternal death during/after delivery substantially increases infant mortality. Observational evidence suggests 4-25x higher infant mortality when mothers die. This is a rough guess — not formally modeled.")

add("Total UoV per woman treated (with adjustments)",
    "=B11*(1+B14+B15)",
    "Calculation: mortality UoV × (1 + morbidity + infant)")

# ── Section: Costs ──
add("", None, None)
add("Costs", None, None, is_header=True)

add("Grant size ($)",
    1000000,
    "Standard GW hypothetical grant")

add("Cost per woman treated — best guess ($)",
    12,
    "I'm assuming $12 as midpoint. Device cost ~$1/use (amortized over 72 uses). Training ~$1.62/use. Additional logistics, supervision, return-and-exchange system. Original QEA guessed $8; CHAI Zimbabwe suggests higher total costs when including broader programmatic activities.")

add("Women with obstetric hemorrhage treated",
    "=B19/B20",
    "Calculation")

# ── Section: Results ──
add("", None, None)
add("Results", None, None, is_header=True)

add("Deaths averted (adjusted)",
    "=B21*B7",
    "Calculation")

add("Total UoV from grant",
    "=B21*B16",
    "Calculation")

add("GiveWell benchmark (UoV per $ spent)",
    0.00344,
    "1x cash = 0.003355 UoV/$ (from CE bar of 8x). I'm using 0.00344 as approximate benchmark.")

add("CE as multiples of GiveDirectly",
    "=B25/(B19*B26)",
    "Calculation")

# ── Section: Sensitivity ──
add("", None, None)
add("Sensitivity analysis", None, None, is_header=True)

add("CE at $8/woman (optimistic)",
    "=(B19/$8)*B16/(B19*B26)",
    "If original guess of $8/woman is correct")

add("CE at $12/woman (best guess)",
    "=B27",
    "Same as main estimate")

add("CE at $25/woman (conservative)",
    "=(B19/$25)*B16/(B19*B26)",
    "If programmatic costs are substantially higher than device + training costs alone")

add("CE at $50/woman (pessimistic — full treatment cost)",
    "=(B19/$50)*B16/(B19*B26)",
    "Original BOTEC noted mean treatment cost of $54.56/woman across both RCT groups (including blood transfusion, uterotonics). This would represent NASG as part of full hemorrhage treatment, not incremental NASG cost.")

# ── Section: Key assumptions ──
add("", None, None)
add("Key assumptions and caveats", None, None, is_header=True)

add("1. Mortality rate uses RCT control group",
    None,
    "The 1.6% counterfactual mortality is based on the RCT control group in Zambia/Zimbabwe, adjusted for time trends. Actual mortality among hemorrhaging women varies substantially by setting (higher in Nigeria, lower in India).")

add("2. NASG reaches all hemorrhaging women at facilities",
    None,
    "BOTEC assumes 100% coverage of women with obstetric hemorrhage at participating facilities. Zimbabwe pilot achieved 90%. Women hemorrhaging before facility arrival are not covered.")

add("3. Evidence is observational",
    None,
    "The pooled RR 0.52 comes from non-experimental studies. The 50% IV/EV adjustment attempts to account for this, but the true effect is uncertain. The RCT point estimate (OR 0.36) is actually larger but is not significant (n=15 deaths).")

add("4. Cost per woman excludes ART-like ongoing costs",
    None,
    "Unlike HIV interventions, NASG is a one-time device application with no ongoing treatment costs. The garment itself is reusable. This makes the cost-effectiveness calculation simpler but the upfront cost ($57-100/garment) must be amortized over actual (not theoretical) reuses.")


out_path = os.path.join(os.path.dirname(__file__), "non_pneumatic_anti_shock_garment.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

from export_csv import save_csv
save_csv(ws, out_path.replace(".xlsx", ".csv"))
