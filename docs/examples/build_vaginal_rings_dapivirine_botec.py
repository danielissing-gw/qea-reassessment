import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOTEC"

bold = Font(bold=True)
section_font = Font(bold=True, size=11)
result_font = Font(bold=True, size=12)
link_font = Font(color="0000FF", underline="single")

# Commonly referenced URLs
PREP_CEA_SUMMARY = "https://docs.google.com/spreadsheets/d/1vH1Fh9TUotkKvIeX7uR9-hVHKqrGJJwBQLugb8rkDdA/edit?gid=1203937207#gid=1203937207"
PREP_CEA_DETAIL = "https://docs.google.com/spreadsheets/d/1vH1Fh9TUotkKvIeX7uR9-hVHKqrGJJwBQLugb8rkDdA/edit?gid=2020638865#gid=2020638865"
PREP_CEA_INCIDENCE = PREP_CEA_DETAIL + "&range=23:23"
PREP_CEA_BENCHMARK = PREP_CEA_SUMMARY + "&range=A7:L7"


def add_link(row, col, text, url):
    """Write a cell with a hyperlink."""
    c = ws.cell(row, col, text)
    c.hyperlink = url
    c.font = link_font
    return c


# Column widths
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 80

# ── Costs ─────────────────────────────────────────────────────────────────────
ws.cell(1, 1, "Costs").font = section_font
ws.cell(1, 2, "Value").font = bold
ws.cell(1, 3, "Source / notes").font = bold

ws.cell(2, 1, "High-risk cohort who begins using PrEP (arbitrary)")
ws.cell(2, 2, 1000)

ws.cell(3, 1, "Drug costs per person-year of protection")
ws.cell(3, 2, "=12*5.9")
ws.cell(3, 2).number_format = "$#,##0.00"
add_link(3, 3, "Pop Council estimate",
         "https://popcouncil.org/media/dapivirine-vaginal-ring-price-drops-by-over-50-bringing-more-affordable-hiv-prevention-options-to-women-and-adolescent-girls-in-africa-and-beyond/")

ws.cell(4, 1, "Delivery costs per person-year of protection")
ws.cell(4, 2, 50)
add_link(4, 3, 'Stegman 2025 estimates that "The cost to provide a full year of '
               'PrEP ring was US$206; 76% (US$156) of that cost was attributed '
               'to the PrEP rings."',
         "https://academiccommons.columbia.edu/doi/10.7916/wqj1-q879")

ws.cell(5, 1, "Average time that women will use dapivirine ring (years)")
ws.cell(5, 2, 1)
add_link(5, 3, "Benchmarking on oral PrEP", PREP_CEA_BENCHMARK)

ws.cell(6, 1, "Total lifetime costs of dap ring")
ws.cell(6, 2, "=SUM(B3:B4)*B5")
ws.cell(6, 2).number_format = "$#,##0.00"
ws.cell(6, 3, "Calculation")

ws.cell(7, 1, "Total program cost")
ws.cell(7, 2, "=B6*B2")
ws.cell(7, 2).number_format = "$#,##0"
ws.cell(7, 3, "Calculation")

# ── Infections prevented ──────────────────────────────────────────────────────
ws.cell(8, 1, "Infections prevented").font = section_font

ws.cell(9, 1, "Baseline rate of new HIV incidence")
ws.cell(9, 2, 0.05)
ws.cell(9, 2).number_format = "0%"
add_link(9, 3, "For KP programs, we model 1.9 - 9.6% in SSA. "
               "Using a rough midpoint here.",
         PREP_CEA_INCIDENCE)

ws.cell(10, 1, "Risk reduction for HIV acquisition, "
               "dapivirine ring compared with placebo")
ws.cell(10, 2, "=AVERAGE(0.39,0.62)")
c = ws.cell(10, 3, "Average of Baeten 2022 and Nel 2021")
c.hyperlink = "https://pmc.ncbi.nlm.nih.gov/articles/PMC8038210/"
c.font = link_font
c.comment = Comment(
    'Nel 2021 (DREAM): "18 (1.9%) HIV-1 infections were confirmed during '
    'DVR use, resulting in an incidence of 1.8 (95% CI 1.1\u20132.6) per 100 '
    'person-years, 62% lower than the simulated placebo rate."\n'
    'https://www.thelancet.com/journals/lanhiv/article/PIIS2352-3018(20)30300-3/fulltext\n\n'
    'Baeten 2022 (HOPE): "HIV-1 incidence was 2.7 per 100 person-years '
    '(95% CI 1.9\u20133.8, n=35 infections), compared to an expected incidence '
    'of 4.4 per 100 person-years (95% CI 3.2\u20135.8) among a population '
    'matched on age, site, and presence of a sexually transmitted infection '
    'from the placebo group of ASPIRE." 1 \u2212 2.7/4.4 \u2248 39%.\n'
    'https://pmc.ncbi.nlm.nih.gov/articles/PMC8038210/',
    "BOTEC"
)

ws.cell(11, 1, "IV adjustment")
ws.cell(11, 2, 0.95)
ws.cell(11, 2).number_format = "0%"
add_link(11, 3, "Use same values as for LEN", PREP_CEA_SUMMARY)

ws.cell(12, 1, "EV adjustment")
ws.cell(12, 2, 0.9)
ws.cell(12, 2).number_format = "0%"
add_link(12, 3, "Use same values as for LEN", PREP_CEA_SUMMARY)

ws.cell(13, 1, "Penalty for repetitive saving (infections merely delayed)")
ws.cell(13, 2, -0.25)
ws.cell(13, 2).number_format = "0%"
add_link(13, 3, "See estimate for MOZ, where incidence is estimated at 4.7%, "
                "similar to what we assume here. "
                "Might be lower for non-KPs (e.g. AGYW)",
         PREP_CEA_INCIDENCE)

ws.cell(14, 1, "Additional infections averted per infection "
               "averted directly by PrEP")
ws.cell(14, 2, 0.6)
add_link(14, 3, "Average guess for SSA, see here", PREP_CEA_INCIDENCE)

ws.cell(15, 1, "Infections prevented by Dapivirine ring (lifetime)")
ws.cell(15, 2, "=(B2*B9*B10*B11*B12)*(1+B13)*(1+B14)")
ws.cell(15, 2).number_format = "0"
ws.cell(15, 3, "Calculation")

# ── Benefits ──────────────────────────────────────────────────────────────────
ws.cell(16, 1, "Benefits (all per infection prevented)").font = section_font

ws.cell(17, 1, "Units of value from YLLs gained")
ws.cell(17, 2, 34)
add_link(17, 3, "HIV Pre-Exposure Prophylaxis CEA", PREP_CEA_DETAIL)

ws.cell(18, 1, "Units of value from YLDs averted")
ws.cell(18, 2, 6)
add_link(18, 3, "HIV Pre-Exposure Prophylaxis CEA", PREP_CEA_DETAIL)

ws.cell(19, 1, "Units of value for income benefits")
ws.cell(19, 2, 8)
add_link(19, 3, "HIV Pre-Exposure Prophylaxis CEA", PREP_CEA_DETAIL)

ws.cell(20, 1, "Units of value from treatment costs averted")
ws.cell(20, 2, 16)
add_link(20, 3, "value for SSA", PREP_CEA_DETAIL)

ws.cell(21, 1, "Units of value from subjective well-being benefits")
ws.cell(21, 2, 11)
add_link(21, 3, "value for SSA", PREP_CEA_DETAIL)

ws.cell(22, 1, "Total UoV generated by program")
ws.cell(22, 2, "=SUM(B17:B21)*B15")
ws.cell(22, 2).number_format = "#,##0"
ws.cell(22, 3, "Calculation")

# ── Final CE estimates ────────────────────────────────────────────────────────
ws.cell(23, 1, "Final CE estimates").font = section_font

ws.cell(24, 1, "Cost per HIV infection prevented by PrEP")
ws.cell(24, 2, "=B6*B2/B15")
ws.cell(24, 2).number_format = "$#,##0"
ws.cell(24, 3, "Calculation")

ws.cell(25, 1, "Value per dollar spent on benchmark")
ws.cell(25, 2, 0.00335)
add_link(25, 3, "GW moral weight", PREP_CEA_SUMMARY)

ws.cell(26, 1, "CE")
ws.cell(26, 2, "=B22/B7/B25")
ws.cell(26, 2).number_format = "0.0"
ws.cell(26, 2).font = result_font
ws.cell(26, 3, "Calculation")

# Wrap text in source/notes column
for row in ws.iter_rows(min_row=1, max_row=26, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

out_path = os.path.join(os.path.dirname(__file__), "vaginal_rings_dapivirine_botec.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
